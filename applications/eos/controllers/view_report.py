# -*- coding: utf-8 -*-
###############################################################################
# Author : 
# Date   : 
# 
# Description : 
# 
###############################################################################
import logging

from applications.eos.modules import common, manh_test
from applications.eos.common import helper
from gluon import current
from datetime import datetime, timedelta
from w2pex import date_util
from datetime import datetime, timedelta
import collections
from operator import attrgetter

if False:
    from gluon import *

    request = current.request
    response = current.response
    session = current.session
    cache = current.cache
    T = current.T
    db = current.db
    pydb = pydb


@auth.requires_login()
def call():
    return service()


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def index():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    areas = db(db.areas.id > 0).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces, areas=areas)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_exceed_indicator_1():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station_for_envisoft()
    areas = db(db.areas.id > 0).select()
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
            row = db(db.stations.id.belongs(station_ids)).select(db.stations.area_id, distinct=True)
            areas_ids = [str(it.area_id) for it in row]
            areas = db(db.areas.id.belongs(areas_ids)).select(db.areas.area_name,
                                                              db.areas.id)
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces, areas=areas)


###############################################################################
@service.json
def get_list_report_exceed_indicator_1_bak(*args, **kwargs):
    try:
        from datetime import datetime, timedelta
        from w2pex import date_util
        iDisplayStart = int(request.vars.iDisplayStart)
        iDisplayLength = int(request.vars.iDisplayLength)
        station_id = request.vars.station_id
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        data_type = request.vars.data_type
        if data_type:
            data_type = int(data_type)

        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        aaData = []
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        table = 'data_min'
        if data_type == 2:
            table = 'data_adjust'
        conditions = {'station_id': station_id}
        if from_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            # conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
            conditions['get_time']['$gte'] = date_util.string_to_datetime(from_date)
        if to_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            # conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
            conditions['get_time']['$lte'] = date_util.string_to_datetime(to_date) + timedelta(days=1)
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions['get_time'] = {'$gte': from_date, '$lte': to_date}

        # conditions['is_exceed'] = True
        attrs = {'get_time': 1, 'data': 1}

        list_data = pydb[table].find(conditions, attrs).sort('get_time', -1) \
            .skip(iDisplayStart).limit(iDisplayLength)
        # Tong so ban ghi khong thuc hien phan trang
        iTotalRecords = pydb[table].count(conditions)

        rows = db(db.station_indicator.station_id == station_id).select(db.station_indicator.ALL,
                                                                        orderby=db.station_indicator.station_id)

        dic_station_indicator = dict()
        indicators = common.get_indicator_dict()

        for row in rows:
            if not dic_station_indicator.has_key(row['station_id']):
                dic_station_indicator[str(row['station_id'])] = dict()
            if indicators.has_key(row['indicator_id']):
                indicator_name = indicators[row['indicator_id']]
                dic_station_indicator[str(row['station_id'])][indicator_name] = row.as_dict()

        qcvn_dict = dic_station_indicator[station_id]

        for c, item in enumerate(list_data):
            row = [
                c + 1,
                item['get_time'].strftime("%H:%M %d/%m/%Y")
            ]
            for indicator in added_columns:
                # i_name_decode = u'{}'.format(indicator)
                i_name_decode = indicator.encode('utf-8')
                x = str(item['data'][indicator])
                if x == 'None' or x == 'NULL' or x == '-' or x.find(',') != - \
                        1:
                    row.append(SPAN('-'))
                else:
                    z = float(x)
                    name_decode = indicator.encode('utf-8')
                    qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                        name_decode) else None
                    qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                        name_decode) else None
                    check_qcvn = True
                    if z or z == 0:
                        if qcvn_max or qcvn_min:
                            check_qcvn = True
                            title = 'QCVN Min: %s' % qcvn_min + ' - QCVN Max: %s' % qcvn_max
                        if qcvn_max and z > qcvn_max:
                            check_qcvn = False
                            title = 'QCVN Max: %s' % qcvn_max
                        if qcvn_min and z < qcvn_min:
                            check_qcvn = False
                            title = 'QCVN Min: %s' % qcvn_min
                        if not check_qcvn:
                            row.append(SPAN(z, _style="background:#EA3223; color:white", _class="badge", _title=title))
                        else:
                            row.append(SPAN(z, _style="background:#1dce6c; color:white", _class="badge"))
            aaData.append(row)

        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


###############################################################################
@service.json
def get_list_report_exceed_indicator_1(*args, **kwargs):
    try:
        from datetime import datetime, timedelta
        from w2pex import date_util
        iDisplayStart = int(request.vars.iDisplayStart)
        iDisplayLength = int(request.vars.iDisplayLength)
        station_id = request.vars.station_id
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        data_type = request.vars.data_type
        exeed_type_data = request.vars.exeed_type_data
        if data_type:
            data_type = int(data_type)

        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        aaData = []
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        table = exeed_type_data
        if data_type == 2:
            if exeed_type_data == 'data_min':
                table = 'data_adjust'
            else:
                table = exeed_type_data + '_adjust'

        conditions = {'station_id': station_id}
        time_range = request.vars.time_range
        if time_range:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=int(time_range))
        if from_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            # conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
            conditions['get_time']['$gte'] = date_util.string_to_datetime(from_date)
        if to_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            # conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
            conditions['get_time']['$lte'] = date_util.string_to_datetime(to_date) + timedelta(days=1)
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions['get_time'] = {'$gte': from_date, '$lte': to_date}

        attrs = {'get_time': 1, 'data': 1}

        rows = db(db.station_indicator.station_id == station_id).select(db.station_indicator.ALL,
                                                                        orderby=db.station_indicator.station_id)

        dic_station_indicator = dict()
        indicators = common.get_indicator_dict()

        for row in rows:
            if not dic_station_indicator.has_key(row['station_id']):
                dic_station_indicator[str(row['station_id'])] = dict()
            if indicators.has_key(row['indicator_id']):
                indicator_name = indicators[row['indicator_id']]
                dic_station_indicator[str(row['station_id'])][indicator_name] = row.as_dict()

        qcvn_dict = dic_station_indicator[station_id]

        query = common.where_is_exceed(added_columns, qcvn_dict)
        if len(query) > 0:
            conditions["$or"] = query
        list_data = pydb[table].find(conditions, attrs).sort('get_time', -1) \
            .skip(iDisplayStart).limit(iDisplayLength)
        # Tong so ban ghi khong thuc hien phan trang
        iTotalRecords = pydb[table].count(conditions)
        for i, item in enumerate(list_data):
            row = [
                str(iDisplayStart + 1 + i),
                item['get_time']
            ]
            added_item = dict()
            if item['data']:
                if added_columns:
                    for data_key in added_columns:
                        i_name = str(data_key).encode('utf-8')
                        current_indicator = item['data'][i_name] if item['data'].has_key(i_name) else ''
                        if not current_indicator:
                            added_item[i_name] = SPAN('-')
                        else:
                            try:
                                z = float(current_indicator)
                                qcvn_min = qcvn_dict[data_key]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                                    data_key) else None
                                qcvn_max = qcvn_dict[data_key]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                                    data_key) else None
                                check_qcvn = True
                                if z or z == 0:
                                    if qcvn_min or qcvn_max:
                                        check_qcvn = True
                                    if qcvn_max and z > qcvn_max:
                                        check_qcvn = False
                                    if qcvn_min and z < qcvn_min:
                                        check_qcvn = False
                                    if not check_qcvn:
                                        added_item[data_key] = common.convert_data(z)
                                    else:
                                        added_item[data_key] = '-'
                            except:
                                added_item[data_key] = '-'
            for column in added_columns:
                if column and added_item.has_key(column):
                    i_name = column.encode('utf-8')
                    row.append(added_item[i_name])
            aaData.append(row)
        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_exceed_times_2():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station_for_envisoft()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    areas_default = db(db.areas.id > 0).select(db.areas.area_name,
                                               db.areas.id)

    careers = db(db.manager_careers.id > 0).select(db.manager_careers.id, db.manager_careers.career_name)

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces, areas_default=areas_default, careers=careers)


@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_exceed_a_station():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station_for_envisoft()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    areas_default = db(db.areas.id > 0).select(db.areas.area_name,
                                               db.areas.id)
    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces, areas_default=areas_default)


@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_data_exceed_by_province():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];

    province_dict = common.get_province_dict()

    provinces = common.get_province_have_station_for_envisoft()

    conditions = (db.stations.id > 0)

    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    areas_default = db(db.areas.id > 0).select(db.areas.area_name,
                                               db.areas.id)
    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type,
                default_provinces=default_provinces, areas_default=areas_default)


@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_data_exceed_by_year():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station_for_envisoft()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    areas_default = db(db.areas.id > 0).select(db.areas.area_name,
                                               db.areas.id)
    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces, areas_default=areas_default)


################################################################################
@service.json
def get_list_report_exceed_ratio_0(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_id = request.vars.station_id
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        province_id = request.vars.province_id
        data_type = request.vars.data_type
        exceed_type_data = request.vars.exceed_type_data
        table = exceed_type_data
        if data_type:
            if int(data_type) == 2:
                if exceed_type_data == 'data_min':
                    table = 'data_adjust'
                else:
                    table = exceed_type_data + "_adjust"

        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')

        conditions = {'station_id': station_id}
        time_range = request.vars.time_range
        if time_range:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=int(time_range))
        aaData = []  # Du lieu json se tra ve
        iTotalRecords = 0  # Tong so ban ghi
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)

        if from_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            conditions['get_time']['$gte'] = date_util.string_to_datetime(from_date)
        if to_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            conditions['get_time']['$lte'] = date_util.string_to_datetime(to_date) + timedelta(days=1)
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions['get_time'] = {'$gte': from_date, '$lte': to_date}
        attrs = {'get_time': 1, 'data': 1}
        rows = pydb[table].find(conditions, attrs).sort('get_time', -1) \
            .skip(iDisplayStart)
        dict_data = {}
        for row in rows:
            key = row['get_time'].strftime('%Y-%m-%d')
            if not dict_data.get(key): dict_data[key] = {'data': []}
            dict_data[key]['data'].append(row['data'])
        iTotalRecords = len(dict_data)

        stt = 0

        iRow = iDisplayStart + 1
        indicators = common.get_indicator_dict()
        conditions = (db.station_indicator.station_id == station_id)
        rows = db(conditions).select(db.station_indicator.ALL)
        qcvn_dict = dict()
        for row in rows:
            indicator_name = indicators[row.indicator_id]
            qcvn_dict[indicator_name] = row.as_dict()

        b = dict_data.keys()
        b.sort(reverse=True)
        if b:
            for i in range(iDisplayStart, iDisplayStart + iDisplayLength):
                if i == iTotalRecords:
                    break;
                else:
                    key = b[i]
                    a = dict_data[key]['data']
                    # Lang update tinh dung dan cua du lieu
                    query = {'station_id': station_id}
                    if not query.has_key('get_time'):
                        query['get_time'] = {}
                    from_date = date_util.string_to_datetime(key)
                    to_date = date_util.string_to_datetime(key) + timedelta(days=1)
                    query['get_time'] = {'$gte': from_date, '$lte': to_date}
                    countTotal = pydb[table].count(query)
                    # countTotal = len(a)
                    count = dict()
                    for i in a:
                        c = i
                        for indicator_name in c:
                            v = str(c[indicator_name])
                            x = v.replace(",", "")
                            if x == 'NULL' or x == 'None' or x == '-':
                                continue
                            if x != '-':
                                z = float(x)
                                name_decode = indicator_name.encode('utf-8')
                                qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                                    name_decode) else None
                                qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                                    name_decode) else None
                                check_qcvn = True
                                if z:
                                    if qcvn_max or qcvn_min:
                                        check_qcvn = True
                                    if qcvn_max and z > qcvn_max:
                                        check_qcvn = False
                                    if qcvn_min and z < qcvn_min:
                                        check_qcvn = False
                                    if not check_qcvn:
                                        if not count.has_key(name_decode):
                                            count[name_decode] = 1
                                        else:
                                            count[name_decode] = int(count[name_decode]) + 1
                    date_time_obj = datetime.strptime(key, '%Y-%m-%d')
                    row = [
                        str(iRow + stt),
                        date_time_obj.strftime('%d/%m/%Y'),
                    ]
                    stt = stt + 1
                    for column in added_columns:
                        if column and count.has_key(column):
                            m = int(count[column])
                            n = (float(m) * 100) / float(countTotal)
                            content = '%s (%s)' % (common.convert_data(n), '%')
                            row.append(content)
                        else:
                            content = '-'
                            row.append(content)
                    aaData.append(row)

        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


####################################################################################3
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_max_min_4():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station_for_envisoft()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)

    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces)


################################################################################
@service.json
def get_list_report_max_min_4(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_type = request.vars.station_type
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        province_id = request.vars.province_id
        data_type = request.vars.data_type
        exceed_type_data = request.vars.exceed_type_data

        conditions = (db.stations.id > 0)
        added_stations = request.vars.added_stations or ''
        if added_stations:
            added_stations = added_stations.split(',')
            conditions &= db.stations.id.belongs(added_stations)
        added_columns = request.vars.custom_added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        if station_type:
            conditions &= (db.stations.station_type == station_type)
        if province_id:
            conditions &= (db.stations.province_id == province_id)
        if not station_type:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station type for viewing data'), success=True)
        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        table = exceed_type_data
        if data_type:
            if int(data_type) == 2:
                if table == 'data_min':
                    table = 'data_adjust'
                else:
                    table = exceed_type_data + '_adjust'

        stations = db(conditions).select(db.stations.id,
                                         db.stations.station_name,
                                         db.stations.station_type,
                                         db.stations.province_id,
                                         limitby=limitby)
        query = db.stations.station_type == station_type
        if province_id:
            query &= db.stations.province_id == province_id
        if added_stations:
            query &= db.stations.id.belongs(added_stations)
        iTotalRecords = db(query).count()
        iRow = iDisplayStart + 1
        indicators = common.get_indicator_dict()
        for c, station in enumerate(stations):
            station_id = station.id
            conditions = (db[table]['id'] > 0)
            conditions &= (db[table]['station_id'] == station_id)
            if from_date:
                conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
            if to_date:
                conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
            if not from_date and not to_date:
                to_date = datetime.now()
                from_date = to_date - timedelta(days=365)
                conditions &= (db[table]['get_time'] >= from_date)
                conditions &= (db[table]['get_time'] < to_date)
            # conditions &= (db[table]['is_exceed'] == True)

            list_data = db(conditions).select(db[table].id,
                                              db[table].get_time,
                                              db[table].station_id,
                                              db[table].data,
                                              )
            # Tong so ban ghi khong thuc hien phan trang
            conditions = (db.station_indicator.station_id == station_id)
            conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
            rows = db(conditions).select(db.station_indicator.ALL)
            qcvn_dict = dict()
            for row in rows:
                indicator_name = indicators[row.indicator_id]
                qcvn_dict[indicator_name] = row.as_dict()
            max = dict()
            min = dict()
            total = dict()
            count = dict()
            # for in
            for i, item in enumerate(list_data):
                if item.data:
                    for indicator_name in item.data:
                        z = str(item.data[indicator_name])
                        if z == 'NULL' or z == '-' or z == 'None':
                            break
                        else:
                            x = z.replace(",", "")
                            v = float(x)
                        name_decode = indicator_name.encode('utf-8')
                        if v or v == 0:
                            # max
                            if not max.has_key(name_decode):
                                max[name_decode] = v
                            else:
                                if v > max[name_decode]:
                                    max[name_decode] = v
                            # min
                            if not min.has_key(name_decode):
                                min[name_decode] = v
                            else:
                                if v < min[name_decode]:
                                    min[name_decode] = v
                            # trung binh
                            if not total.has_key(name_decode):
                                total[name_decode] = v
                                count[name_decode] = 1
                            else:
                                total[name_decode] += v
                                count[name_decode] += 1;
            row = [
                str(iRow + c),
                station.station_name,
            ]
            for column in added_columns:
                if column and count.has_key(column):
                    row.append("{0:.2f}".format(max[column]))
                    row.append("{0:.2f}".format(min[column]))
                    row.append("{0:.2f}".format(total[column] / count[column]))
                else:
                    row.append('-')
                    row.append('-')
                    row.append('-')
            aaData.append(row)
        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


###############################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_exceed_ratio_0():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station_for_envisoft()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces)


##############################################################################
@service.json
def get_list_report_exceed_times_2_new(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_id = request.vars.station_id
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        data_type = request.vars.data_type

        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')

        table = 'data_min'
        if data_type:
            if int(data_type) == 2:
                table = 'data_adjust'
        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        iRow = iDisplayStart + 1
        indicators = common.get_indicator_dict()
        stations = db(db.stations.id == station_id).select(db.stations.station_name)

        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
        if to_date:
            conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)
        # conditions &= (db[table]['is_exceed'] == True)

        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].station_id,
                                          db[table].data,
                                          )
        # Tong so ban ghi khong thuc hien phan trang
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db(conditions).select(db.station_indicator.ALL)
        qcvn_dict = dict()
        for row in rows:
            indicator_name = indicators[row.indicator_id]
            qcvn_dict[indicator_name] = row.as_dict()
        count = dict()
        # for in
        for i, item in enumerate(list_data):
            if item.data:
                for indicator_name in item.data:
                    x = str(item.data[indicator_name])
                    if x == 'None' or x == 'NULL' or x == '-' or x.find(',') != -1:
                        continue
                    else:
                        z = float(x)
                        name_decode = indicator_name.encode('utf-8')
                        qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        check_qcvn = True
                        if z or z == 0:
                            if qcvn_max or qcvn_min:
                                check_qcvn = True
                            if qcvn_max and z > qcvn_max:
                                check_qcvn = False
                            if qcvn_min and z < qcvn_min:
                                check_qcvn = False
                            if not check_qcvn:
                                if not count.has_key(name_decode):
                                    count[name_decode] = 1
                                else:
                                    count[name_decode] = int(count[name_decode]) + 1
        row = [
            1,
            stations[0].station_name,
        ]
        for column in added_columns:
            if column and count.has_key(column):
                row.append(SPAN(count[column], _style="background:#ff0000; color:white", _class="badge",
                                _onclick="show_info('%s')" % (column)))
            else:
                row.append(SPAN('0', _style="background:#00e400; color:white", _class="badge"))

        aaData.append(row)
        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


###############################################################################
@service.json
def get_list_report_exceed_times_2(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_type = request.vars.station_type
        if not station_type:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station type for viewing data'), success=True)
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        province_id = request.vars.province_id
        data_type = request.vars.data_type
        exeed_type_data = request.vars.exeed_type_data
        area_ids = request.vars.area_ids
        careers = request.vars.careers

        conditions = (db.stations.id > 0)
        added_stations = request.vars.added_stations or ''
        if added_stations:
            added_stations = added_stations.split(',')
            conditions &= db.stations.id.belongs(added_stations)
        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        if station_type:
            conditions &= (db.stations.station_type == station_type)
        if province_id:
            conditions &= (db.stations.province_id == province_id)
        if area_ids:
            area_ids = area_ids.split(',')
            conditions &= (db.stations.area_ids.belongs(area_ids))
        if careers:
            careers = careers.split(',')
            conditions &= (db.stations.career.belongs(careers))

        time_range = request.vars.time_range
        if time_range:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=int(time_range))

        table = exeed_type_data
        if data_type:
            if int(data_type) == 2:
                if exeed_type_data == 'data_min':
                    table = 'data_adjust'
                else:
                    table = exeed_type_data + "_adjust"

        print("table", table, added_columns)
        aaData = []  # Du lieu json se tra ve
        arrSorted = []
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)

        stations = db(conditions).select(db.stations.id,
                                         db.stations.station_name,
                                         db.stations.station_type,
                                         db.stations.province_id,
                                         limitby=limitby)

        query = db.stations.station_type == station_type
        if province_id:
            query &= db.stations.province_id == province_id
        if added_stations:
            query &= db.stations.id.belongs(added_stations)
        iTotalRecords = db(query).count()
        iRow = iDisplayStart + 1
        indicators = common.get_indicator_dict()

        for c, station in enumerate(stations):
            count_total_dict = dict()
            station_id = str(station.id)
            conditions = {'station_id': station_id}
            # if table == "data_min" or table == "data_adjust":
            #     conditions['is_exceed'] = True
            if from_date:
                if not conditions.has_key('get_time'):
                    conditions['get_time'] = {}
                conditions['get_time']['$gte'] = date_util.string_to_datetime(from_date)
            if to_date:
                if not conditions.has_key('get_time'):
                    conditions['get_time'] = {}
                conditions['get_time']['$lte'] = date_util.string_to_datetime(to_date) + timedelta(days=1)
            if not from_date and not to_date:
                to_date = datetime.now()
                from_date = to_date - timedelta(days=365)
                conditions['get_time'] = {'$gte': from_date, '$lte': to_date}

            attrs = {'get_time': 1, 'data': 1}
            list_data = pydb[table].find(conditions, attrs).sort('get_time', -1)
            # Tong so ban ghi khong thuc hien phan trang
            conditions = (db.station_indicator.station_id == station_id)
            conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
            rows = db(conditions).select(db.station_indicator.ALL)
            qcvn_dict = dict()
            for row in rows:
                indicator_name = indicators[row.indicator_id]
                qcvn_dict[indicator_name] = row.as_dict()
            count = dict()
            # for in
            for item in list_data:
                if item['data']:

                    for indicator_name in item['data']:
                        x = str(item['data'][indicator_name])
                        if x == 'None' or x == 'NULL' or x == '-' or x.find(',') != -1:
                            continue
                        else:
                            z = float(x)
                            name_decode = indicator_name.encode('utf-8')
                            qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                                name_decode) else None
                            qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                                name_decode) else None
                            check_qcvn = True
                            if z or z == 0:
                                if qcvn_max or qcvn_min:
                                    check_qcvn = True
                                if qcvn_max and z > qcvn_max:
                                    check_qcvn = False
                                if qcvn_min and z < qcvn_min:
                                    check_qcvn = False
                                if not check_qcvn:
                                    if not count.has_key(name_decode):
                                        count[name_decode] = 1
                                    else:
                                        count[name_decode] = int(count[name_decode]) + 1
            row = [
                station.station_name,
            ]

            count_total_dict['total'] = 0
            for column in added_columns:

                if column and count.has_key(column):
                    count_total_dict['total'] += count[column]
                    row.append(SPAN(count[column], _style="background:#ff0000; color:white", _class="badge"))
                else:
                    row.append(SPAN('0', _style="background:#00e400; color:white", _class="badge"))
            count_total_dict['data'] = row
            arrSorted.append(count_total_dict)

        sorted(arrSorted, key=lambda x: x['total'])
        i = 0
        for item in arrSorted:
            if item:
                item['data'].insert(0, iRow + i)
                aaData.append(item['data'])
            i += 1

        return dict(iTotalRecords=len(aaData), iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


###############################################################################
@service.json
def get_report_detail_exceed_times_2_a_station(*args, **kwargs):
    try:
        import time
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_type = request.vars.station_type
        if not station_type:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station type for viewing data'), success=True)
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        province_id = request.vars.province_id
        data_type = request.vars.data_type
        exeed_type_data = request.vars.exeed_type_data
        area_ids = request.vars.area_ids

        conditions = (db.stations.id > 0)
        added_stations = request.vars.added_stations or ''
        if added_stations:
            conditions &= (db.stations.id == added_stations)

        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')

        if area_ids:
            conditions &= (db.stations.area_ids.belongs([area_ids]))

        time_range = request.vars.time_range
        if time_range:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=int(time_range))

        table = exeed_type_data
        if data_type:
            if int(data_type) == 2:
                if exeed_type_data == 'data_min':
                    table = 'data_adjust'
                else:
                    table = exeed_type_data + "_adjust"

        aaData = []  # Du lieu json se tra ve
        arrSorted = []
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)

        stations = db(conditions).select(db.stations.id,
                                         db.stations.station_name,
                                         db.stations.station_type,
                                         db.stations.province_id,
                                         limitby=limitby)

        query = db.stations.station_type == station_type
        if province_id:
            query &= db.stations.province_id == province_id
        if added_stations:
            query &= (db.stations.id == added_stations)
        iTotalRecords = db(query).count()
        iRow = iDisplayStart + 1
        indicators = common.get_indicator_dict()

        count_all_from_station_to_file = {}

        for c, station in enumerate(stations):
            count_total_dict = dict()
            station_id = str(station.id)
            conditions = {'station_id': station_id}
            # if table == "data_min" or table == "data_adjust":
            #     conditions['is_exceed'] = True
            if from_date:
                if not conditions.has_key('get_time'):
                    conditions['get_time'] = {}
                conditions['get_time']['$gte'] = date_util.string_to_datetime(from_date)
            if to_date:
                if not conditions.has_key('get_time'):
                    conditions['get_time'] = {}
                conditions['get_time']['$lte'] = date_util.string_to_datetime(to_date) + timedelta(days=1)
            if not from_date and not to_date:
                to_date = datetime.now()
                from_date = to_date - timedelta(days=365)
                conditions['get_time'] = {'$gte': from_date, '$lte': to_date}

            attrs = {'get_time': 1, 'data': 1}
            list_data = pydb[table].find(conditions, attrs).sort('get_time', -1)

            # Tong so ban ghi khong thuc hien phan trang
            conditions = (db.station_indicator.station_id == station_id)
            conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
            rows = db(conditions).select(db.station_indicator.ALL)
            qcvn_dict = dict()
            for row in rows:
                indicator_name = indicators[row.indicator_id]
                qcvn_dict[indicator_name] = row.as_dict()
            count = dict()
            # for in
            for item in list_data:
                if item['data']:
                    for indicator_name in item['data']:
                        if not count_all_from_station_to_file.has_key(indicator_name.encode('utf-8')):
                            count_all_from_station_to_file[indicator_name.encode('utf-8')] = {
                                "day_count_exceed": [],
                            }
                        else:
                            if not count_all_from_station_to_file[indicator_name.encode('utf-8')].has_key(
                                    "day_count_exceed"):
                                count_all_from_station_to_file[indicator_name.encode('utf-8')]["day_count_exceed"] = []

                        x = str(item['data'][indicator_name])
                        if x == 'None' or x == 'NULL' or x == '-' or x.find(',') != -1:
                            continue
                        else:
                            if not count_all_from_station_to_file[indicator_name.encode('utf-8')].has_key("total_item"):
                                count_all_from_station_to_file[indicator_name.encode('utf-8')]["total_item"] = 1
                            else:
                                count_all_from_station_to_file[indicator_name.encode('utf-8')]["total_item"] = int(
                                    count_all_from_station_to_file[indicator_name.encode('utf-8')]["total_item"]) + 1

                            z = float(x)
                            name_decode = indicator_name.encode('utf-8')
                            qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                                name_decode) else None
                            qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                                name_decode) else None
                            check_qcvn = True
                            if not count_all_from_station_to_file[name_decode].has_key("qcvn_max"):
                                count_all_from_station_to_file[name_decode]["qcvn_max"] = qcvn_max
                            else:
                                count_all_from_station_to_file[name_decode]["qcvn_max"] = qcvn_max
                            if z or z == 0:
                                if qcvn_max or qcvn_min:
                                    check_qcvn = True
                                if qcvn_max and z > qcvn_max:
                                    check_qcvn = False
                                if qcvn_min and z < qcvn_min:
                                    check_qcvn = False
                                if not check_qcvn:
                                    # value max
                                    if not count_all_from_station_to_file[name_decode].has_key("max_exced_value"):
                                        count_all_from_station_to_file[name_decode]["max_exced_value"] = z
                                    elif z > float(count_all_from_station_to_file[name_decode]["max_exced_value"]):
                                        count_all_from_station_to_file[name_decode]["max_exced_value"] = z

                                    # value total exceed
                                    if not count_all_from_station_to_file[name_decode].has_key("total_exceed"):
                                        count_all_from_station_to_file[name_decode]["total_exceed"] = 1
                                    else:
                                        count_all_from_station_to_file[name_decode]["total_exceed"] = int(
                                            count_all_from_station_to_file[name_decode]["total_exceed"]) + 1

                                    # day count
                                    if len(count_all_from_station_to_file[name_decode]["day_count_exceed"]) == 0:
                                        count_all_from_station_to_file[name_decode]["day_count_exceed"].append(
                                            item["get_time"].strftime("%Y-%m-%d"))
                                    elif item["get_time"].strftime("%Y-%m-%d") not in \
                                            count_all_from_station_to_file[name_decode]["day_count_exceed"]:
                                        count_all_from_station_to_file[name_decode]["day_count_exceed"].append(
                                            item["get_time"].strftime("%Y-%m-%d"))

                                    if not count.has_key(name_decode):
                                        count[name_decode] = 1
                                    else:
                                        count[name_decode] = int(count[name_decode]) + 1
        row = [
            [1, "Số ngày có giá trị vượt QCVN".encode('utf-8')],
            [2, "Số lượng giá trị vượt QCVN".encode('utf-8')],
            [3, "Tỷ lệ giá trị vượt QCVN (%)".encode('utf-8')],
            [4, "Giá trị vượt cao nhất".encode('utf-8')],
            [5, "Số lần vượt QCVN cao nhất".encode('utf-8')]
        ]
        for column in added_columns:
            column = column.encode("utf-8")
            if column and count_all_from_station_to_file.has_key(column):
                if count_all_from_station_to_file[column].has_key("day_count_exceed") and len(
                        count_all_from_station_to_file[column]["day_count_exceed"]) > 0:
                    row[0].append(len(count_all_from_station_to_file[column]["day_count_exceed"]))
                else:
                    row[0].append(0)

                if count_all_from_station_to_file[column].has_key("total_exceed") and int(
                        count_all_from_station_to_file[column]["total_exceed"]) > 0:
                    row[1].append(int(count_all_from_station_to_file[column]["total_exceed"]))
                else:
                    row[1].append(0)

                if count_all_from_station_to_file[column].has_key("total_exceed") and count_all_from_station_to_file[
                    column].has_key("total_item") and float(count_all_from_station_to_file[column]["total_item"]) > 0:
                    row[2].append("{:.2f}".format(float(
                        float(count_all_from_station_to_file[column]["total_exceed"]) / float(
                            count_all_from_station_to_file[column]["total_item"])) * 100))
                else:
                    row[2].append(0)

                if count_all_from_station_to_file[column].has_key("max_exced_value") and float(
                        count_all_from_station_to_file[column]["max_exced_value"]) > 0:
                    row[3].append(count_all_from_station_to_file[column]["max_exced_value"])
                else:
                    row[3].append(0)

                if count_all_from_station_to_file[column].has_key("max_exced_value") and count_all_from_station_to_file[
                    name_decode].has_key("qcvn_max") \
                        and (count_all_from_station_to_file[column]["qcvn_max"] is not None and float(
                    count_all_from_station_to_file[column]["qcvn_max"]) > 0) \
                        and (count_all_from_station_to_file[column]["max_exced_value"] is not None and float(
                    count_all_from_station_to_file[column]["max_exced_value"])) > 0:
                    row[4].append("{:.2f}".format(
                        float(count_all_from_station_to_file[column]["max_exced_value"]) / float(
                            count_all_from_station_to_file[column]["qcvn_max"])))
                else:
                    row[4].append(0)
            else:
                row[0].append(0)
                row[1].append(0)
                row[2].append(0)
                row[3].append(0)
                row[4].append(0)

        for r in row:
            aaData.append(r)

        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


###############################################################################
@service.json
def get_list_report_exceed_by_province(*args, **kwargs):
    try:

        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        exeed_type_data = request.vars.exeed_type_data
        from_date = request.vars.from_date
        province_id = request.vars.province_id
        to_date = request.vars.to_date
        table = exeed_type_data

        conditions = (db.stations.id > 0)

        if province_id:
            conditions &= (db.stations.province_id == province_id)
        area_ids = request.vars.area_ids

        if area_ids:
            conditions &= (db.stations.area_ids.belongs([area_ids]))

        time_range = request.vars.time_range
        if time_range:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=int(time_range))

        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi

        stations = db(conditions).select(db.stations.id,
                                         db.stations.station_name,
                                         db.stations.station_type,
                                         db.stations.province_id,
                                         )

        province_dict = common.get_province_dict()

        indicators = common.get_indicator_dict()
        group_by_province_dict = dict()
        for c, station in enumerate(stations):
            count_total_dict = dict()
            station_id = str(station.id)
            conditions = {'station_id': station_id}
            if from_date:
                if not conditions.has_key('get_time'):
                    conditions['get_time'] = {}
                conditions['get_time']['$gte'] = date_util.string_to_datetime(from_date)
            if to_date:
                if not conditions.has_key('get_time'):
                    conditions['get_time'] = {}
                conditions['get_time']['$lte'] = date_util.string_to_datetime(to_date) + timedelta(days=1)
            if not from_date and not to_date:
                to_date = datetime.now()
                from_date = to_date - timedelta(days=365)
                conditions['get_time'] = {'$gte': from_date, '$lte': to_date}

            attrs = {'get_time': 1, 'data': 1}
            list_data = pydb[table].find(conditions, attrs).sort('get_time', -1)
            aaaa = pydb[table].count(conditions)
            # Tong so ban ghi khong thuc hien phan trang
            conditions = (db.station_indicator.station_id == station_id)
            conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
            rows = db(conditions).select(db.station_indicator.ALL)
            qcvn_dict = dict()
            for row in rows:
                indicator_name = indicators[row.indicator_id]
                qcvn_dict[indicator_name] = row.as_dict()
            count = dict()
            # for in
            for item in list_data:
                if item['data']:
                    for indicator_name in item['data']:
                        x = str(item['data'][indicator_name])
                        if x == 'None' or x == 'NULL' or x == '-' or x.find(',') != -1:
                            continue
                        else:
                            z = float(x)
                            name_decode = indicator_name.encode('utf-8')
                            qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                                name_decode) else None
                            qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                                name_decode) else None
                            check_qcvn = True
                            if z or z == 0:
                                if qcvn_max or qcvn_min:
                                    check_qcvn = True
                                if qcvn_max and z > qcvn_max:
                                    check_qcvn = False
                                if qcvn_min and z < qcvn_min:
                                    check_qcvn = False
                                if not check_qcvn:
                                    if group_by_province_dict.has_key(station.province_id):
                                        if station.id not in group_by_province_dict[station.province_id]:
                                            group_by_province_dict[station.province_id].append(station.id)
                                    else:
                                        group_by_province_dict[station.province_id] = [station.id]
        i = 0
        total_stations = 0
        if province_id == "":
            for province in province_dict:
                if province and province_id == "":
                    province_name = province_dict[province]
                    total = 0
                    if group_by_province_dict.has_key(province):
                        total = len(group_by_province_dict[province])
                        total_stations = len(group_by_province_dict[province])
                    else:
                        total = 0
                    aaData.append([i + 1, province_name.encode('utf-8'), total])
                i += 1

        if province_id:
            province_name = province_dict[province_id]
            total = 0
            if group_by_province_dict.has_key(province_id):
                total = len(group_by_province_dict[province_id])
                total_stations = len(group_by_province_dict[province_id])
            else:
                total = 0
            aaData.append([1, province_name.encode('utf-8'), total])

        aaData.insert(0, ["", "Tổng số trạm", total_stations])
        data = aaData[iDisplayStart:iDisplayLength + iDisplayStart]

        return dict(iTotalRecords=iTotalRecords, totalStations=total_stations, iTotalDisplayRecords=iTotalRecords,
                    aaData=data, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


@service.json
def get_list_report_exceed_by_year(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_type = request.vars.station_type
        if not station_type:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station type for viewing data'), success=True)
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        province_id = request.vars.province_id
        data_type = request.vars.data_type
        exeed_type_data = request.vars.exeed_type_data
        area_ids = request.vars.area_ids

        conditions = (db.stations.id > 0)
        added_stations = request.vars.added_stations or ''
        if added_stations:
            added_stations = added_stations.split(',')
            conditions &= db.stations.id.belongs(added_stations)
        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        if station_type:
            conditions &= (db.stations.station_type == station_type)
        if province_id:
            conditions &= (db.stations.province_id == province_id)
        if area_ids:
            area_ids = area_ids.split(',')
            conditions &= (db.stations.area_ids.belongs(area_ids))

        time_range = request.vars.time_range
        if time_range:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=int(time_range))

        table = exeed_type_data
        if data_type:
            if int(data_type) == 2:
                if exeed_type_data == 'data_min':
                    table = 'data_adjust'
                else:
                    table = exeed_type_data + "_adjust"

        aaData = []  # Du lieu json se tra ve
        arrSorted = []

        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)

        stations = db(conditions).select(db.stations.id,
                                         db.stations.station_name,
                                         db.stations.station_type,
                                         db.stations.province_id,
                                         limitby=limitby)

        query = db.stations.station_type == station_type
        if province_id:
            query &= db.stations.province_id == province_id
        if added_stations:
            query &= db.stations.id.belongs(added_stations)
        iTotalRecords = db(query).count()
        iRow = iDisplayStart + 1
        indicators = common.get_indicator_dict()

        for c, station in enumerate(stations):
            count_total_dict = dict()
            station_id = str(station.id)
            conditions = {'station_id': station_id}
            if from_date:
                if not conditions.has_key('get_time'):
                    conditions['get_time'] = {}
                conditions['get_time']['$gte'] = date_util.string_to_datetime(from_date)
            if to_date:
                if not conditions.has_key('get_time'):
                    conditions['get_time'] = {}
                conditions['get_time']['$lte'] = date_util.string_to_datetime(to_date) + timedelta(days=1)
            if not from_date and not to_date:
                to_date = datetime.now()
                from_date = to_date - timedelta(days=365)
                conditions['get_time'] = {'$gte': from_date, '$lte': to_date}

            attrs = {'get_time': 1, 'data': 1}
            list_data = pydb[table].find(conditions, attrs).sort('get_time', -1)
            aaaa = pydb[table].count(conditions)
            # Tong so ban ghi khong thuc hien phan trang
            conditions = (db.station_indicator.station_id == station_id)
            conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
            rows = db(conditions).select(db.station_indicator.ALL)
            qcvn_dict = dict()
            for row in rows:
                indicator_name = indicators[row.indicator_id]
                qcvn_dict[indicator_name] = row.as_dict()
            count = dict()
            # for in
            for item in list_data:
                if item['data']:
                    for indicator_name in item['data']:
                        x = str(item['data'][indicator_name])
                        if x == 'None' or x == 'NULL' or x == '-' or x.find(',') != -1:
                            continue
                        else:
                            z = float(x)
                            name_decode = indicator_name.encode('utf-8')
                            qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                                name_decode) else None
                            qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                                name_decode) else None
                            check_qcvn = True
                            if z or z == 0:
                                if qcvn_max or qcvn_min:
                                    check_qcvn = True
                                if qcvn_max and z > qcvn_max:
                                    check_qcvn = False
                                if qcvn_min and z < qcvn_min:
                                    check_qcvn = False
                                if not check_qcvn:
                                    key = item['get_time'].strftime('%d/%m/%Y')
                                    if not count.has_key(name_decode):
                                        count[name_decode] = [key]
                                    else:
                                        if key not in count[name_decode]:
                                            count[name_decode].append(key)
            row = [
                "{} - {}".format(station.station_name, '(Ngày)'),
            ]

            count_total_dict['total'] = 0
            for column in added_columns:
                if column and count.has_key(column):
                    count_total_dict['total'] += len(count[column])
                    row.append(len(count[column]))
                else:
                    row.append(0)
            count_total_dict['data'] = row
            arrSorted.append(count_total_dict)

        sorted(arrSorted, key=lambda x: x['total'])
        i = 0
        for item in arrSorted:
            if item:
                item['data'].insert(0, iRow + i)
                aaData.append(item['data'])
            i += 1

        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


###############################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_data_time_3():
    now = datetime.now()
    station_id = request.vars.station_id
    provinces = common.get_province_have_station_for_envisoft()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    areas = db(db.areas.id > 0).select()
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
            row = db(db.stations.id.belongs(station_ids)).select(db.stations.area_id, distinct=True)
            areas_ids = [str(it.area_id) for it in row]
            areas = db(db.areas.id.belongs(areas_ids)).select(db.areas.area_name,
                                                              db.areas.id)
    stations = db(conditions).select()
    careers = db(db.manager_careers.id > 0).select(db.manager_careers.id, db.manager_careers.career_name)

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, station_id=station_id, default_provinces=default_provinces,
                areas=areas,
                year=now.year, month=now.month, careers=careers)


###############################################################################
@service.json
def get_list_report_data_time_3_new(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        iDisplayEnd = iDisplayStart + iDisplayLength
        from_date = request.vars.from_date
        report_type = request.vars.report_type
        station_type = request.vars.station_type
        area_id = request.vars.area_id
        data_type = request.vars.data_type
        custom_ratio = request.vars.custom_ratio

        time_range = request.vars.time_range
        if time_range:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=int(time_range))

        if from_date:
            from_date = date_util.string_to_datetime(from_date)
        to_date = request.vars.to_date
        if to_date:
            to_date = date_util.string_to_datetime(to_date)
        if report_type == 'province':
            added_stations = request.vars.added_stations or ''
            if not added_stations:
                return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                            message=T('Chose Station'), success=True)
            if added_stations:
                added_stations = added_stations.split(',')

            aaData = []  # Du lieu json se tra ve
            list_data = None  # Du lieu truy van duoc
            iTotalRecords = 0  # Tong so ban ghi
            limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
            table = 'data_min'
            if data_type:
                if int(data_type) == 2:
                    table = 'data_adjust'
            conditions = db.stations.id.belongs(added_stations)

            stations = db(conditions).select(
                db.stations.id,
                db.stations.station_name,
                db.stations.frequency_receiving_data,
                limitby=limitby
            )
            iTotalRecords = db(conditions).count()
            iRow = iDisplayStart + 1
            for c, station in enumerate(stations):
                station_id = str(station.id)
                conditions = {'station_id': station_id}
                if from_date:
                    if not conditions.has_key('get_time'):
                        conditions['get_time'] = {}
                    conditions['get_time']['$gte'] = from_date
                if to_date:
                    if not conditions.has_key('get_time'):
                        conditions['get_time'] = {}
                    conditions['get_time']['$lte'] = to_date + timedelta(days=1)
                if not from_date and not to_date:
                    to_date = datetime.now()
                    from_date = to_date - timedelta(days=365)
                    conditions['get_time'] = {'$gte': from_date, '$lte': to_date}
                delta = to_date - from_date
                if not station.frequency_receiving_data or station.frequency_receiving_data == 5:
                    total_data = int((delta.days + 1) * 287)
                else:
                    freq = station.frequency_receiving_data
                    total_data = int((delta.days + 1) * 24 * 60 / freq)
                list_data = pydb[table].count(conditions)
                v = float(list_data) / float(total_data) * 100
                v = common.convert_data(v)
                content2 = '%s' % (v)
                if custom_ratio != None and custom_ratio != "":
                    if float(v) >= float(custom_ratio):
                        row = [
                            str(iRow + c),
                            station.station_name,
                        ]
                        content1 = '%s' % common.convert_data(list_data)
                        row.append(content1)

                        row.append(content2)
                        aaData.append(row)
                else:
                    row = [
                        str(iRow + c),
                        station.station_name,
                    ]
                    content1 = '%s' % common.convert_data(list_data)
                    row.append(content1)

                    row.append(content2)
                    aaData.append(row)

            return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
        else:
            if custom_ratio == "" or custom_ratio == None:
                custom_ratio = "80"
            added_provinces = request.vars.added_provinces or ''
            if not added_provinces:
                return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=T('No Province choosen!'),
                            success=True)
            added_provinces = added_provinces.split(',')
            added_provinces_for_page = added_provinces[iDisplayStart: iDisplayEnd]

            result_o = dict()
            inx = 0
            for it in added_provinces:
                result_o[it] = inx
                inx += 1

            aaData = []  # Du lieu json se tra ve
            table = 'data_min'
            if data_type:
                if int(data_type) == 2:
                    table = 'data_adjust'
            conditions = (db.stations.id > 0)
            if station_type:
                conditions &= (db.stations.station_type == station_type)
            if area_id:
                conditions &= (db.stations.area_ids.belong(area_id) | db.stations.area_id == area_id)
            if added_provinces_for_page:
                conditions &= (db.stations.province_id.belongs(added_provinces_for_page))

            stations = db(conditions).select(
                db.stations.id,
                db.stations.station_name,
                db.stations.province_id,
                db.stations.station_type,
                db.stations.frequency_receiving_data
            )
            provinces = dict()
            # Xu ly du lieu tinh
            if added_provinces:
                province_tmp = db(db.provinces.id.belongs(added_provinces)).select(db.provinces.province_name,
                                                                                   db.provinces.id)
                for p in province_tmp:
                    provinces[str(p.id)] = p.province_name
            result_dict = {}

            ids = [str(it.id) for it in stations]
            conditions = (db[table]['id'] > 0)
            conditions &= (db[table]['station_id'].belongs(ids))
            if from_date:
                conditions &= (db[table]['get_time'] >= from_date)
            if to_date:
                conditions &= (db[table]['get_time'] < to_date + timedelta(days=1))
            if not from_date and not to_date:
                to_date = datetime.now()
                from_date = to_date - timedelta(days=30)
                conditions &= (db[table]['get_time'] >= from_date)
                conditions &= (db[table]['get_time'] < to_date)
            count_func = db[table]['station_id'].count()

            try:
                data = db(conditions).select(db[table]['station_id'],
                                             db[table]['is_exceed'],
                                             count_func,
                                             groupby=db[table]['station_id'])
            except:
                return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=T('No data'), success=True)
            data_rs = {}

            for item in data:
                if not data_rs.has_key(item[table]['station_id']):
                    data_rs[item[table]['station_id']] = {'t': 0, 'e': 0}
                for k in item['_extra']:
                    if item[table]['is_exceed']:
                        data_rs[item[table]['station_id']]['e'] = item['_extra'][k]
                    data_rs[item[table]['station_id']]['t'] += item['_extra'][k]
            # duyet tren danh sach tram tra ve ket qua cuoi cung
            total_item_dict = {}
            for station in stations:
                province_id = str(station.province_id)
                station_id = str(station.id)
                delta = to_date - from_date
                if not station.frequency_receiving_data or station.frequency_receiving_data == 5:
                    total_data = int((delta.days + 1) * 287)
                else:
                    freq = station.frequency_receiving_data
                    total_data = int((delta.days + 1) * 24 * 60 / freq)

                if not result_dict.has_key(province_id):
                    province_name = provinces[province_id] if provinces.has_key(province_id) else ''
                    result_dict[province_id] = {
                        'order': result_o[province_id],
                        'name': province_name,
                        'total': 0,
                        'stations': [],
                        'info': {
                            'total': 0,  # Du lieu ly tuong
                            'received': 0,  # Du lieu nhan duoc
                            'exceed': 0,  # du lieu vuot chuan
                        },
                    }
                r_station = {'name': station.station_name, 'total': total_data, 'received': 0, 'exceed': 0}
                if data_rs.has_key(station_id):
                    r_station['received'] = data_rs[station_id]['t']
                    r_station['exceed'] = data_rs[station_id]['e']
                    result_dict[province_id]['info']['received'] += data_rs[station_id]['t']
                    result_dict[province_id]['info']['exceed'] += data_rs[station_id]['e']
                result_dict[province_id]['info']['total'] += total_data  # Du lieu ly tuong
                result_dict[province_id]['stations'].append(r_station)
                result_dict[province_id]['total'] += 1
            iTotalRecords = len(added_provinces)
            inx = 0
            ls = result_dict.values()

            def func_s(e):
                return e['order']

            ls.sort(key=func_s)

            for item in ls:
                inx += 1
                row = [
                    str(iDisplayStart + inx),
                    item['name'],
                    len(item['stations']),
                ]
                rs_dict = {}
                gtkey = '>=' + custom_ratio
                ekey = '<' + custom_ratio
                e0 = '=0'
                if len(item['stations']) > 0:
                    for station in item['stations']:
                        rs_ratio = float(station['received']) / float(station['total']) * 100
                        rs_ratio = common.convert_data(rs_ratio)
                        if float(rs_ratio) > int(custom_ratio):
                            if not rs_dict.has_key(gtkey):
                                rs_dict[gtkey] = 1
                            else:
                                rs_dict[gtkey] = int(rs_dict[gtkey]) + 1

                            if not total_item_dict.has_key(gtkey):
                                total_item_dict[gtkey] = 1
                            else:
                                total_item_dict[gtkey] = int(total_item_dict[gtkey]) + 1
                        elif float(rs_ratio) < int(custom_ratio) and float(rs_ratio) > 0:
                            if not rs_dict.has_key(ekey):
                                rs_dict[ekey] = 1
                            else:
                                rs_dict[ekey] = int(rs_dict[ekey]) + 1

                            if not total_item_dict.has_key(ekey):
                                total_item_dict[ekey] = 1
                            else:
                                total_item_dict[ekey] = int(total_item_dict[ekey]) + 1
                        else:
                            if not rs_dict.has_key(e0):
                                rs_dict[e0] = 1
                            else:
                                rs_dict[e0] = int(rs_dict[e0]) + 1

                            if not total_item_dict.has_key(e0):
                                total_item_dict[e0] = 1
                            else:
                                total_item_dict[e0] = int(total_item_dict[e0]) + 1

                if rs_dict.has_key(gtkey):
                    row.append(rs_dict[gtkey])
                else:
                    row.append(0)
                if rs_dict.has_key(ekey):
                    row.append(rs_dict[ekey])
                else:
                    row.append(0)
                if rs_dict.has_key(e0):
                    row.append(rs_dict[e0])
                else:
                    row.append(0)
                aaData.append(row)

            gt80 = 0
            e80 = 0
            etotal0 = 0
            if total_item_dict.has_key(gtkey):
                gt80 = int(total_item_dict[gtkey])
            if total_item_dict.has_key(ekey):
                e80 = int(total_item_dict[ekey])
            if total_item_dict.has_key(e0):
                etotal0 = int(total_item_dict[e0])
            aaData.insert(0, ["", "Trạm quốc gia", len(stations), gt80, e80, etotal0])
            aaData.insert(1, ["", "Trạm địa phương", len(stations), "", "", ""])

            return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


###############################################################################
@service.json
def get_list_report_data_time_3(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        from_date = request.vars.from_date
        data_type = request.vars.data_type

        time_range = request.vars.time_range
        if time_range:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=int(time_range))

        if from_date:
            from_date = date_util.string_to_datetime(from_date)
        to_date = request.vars.to_date
        if to_date:
            to_date = date_util.string_to_datetime(to_date)
        added_stations = request.vars.added_stations or ''
        if not added_stations:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Chose Station'), success=True)
        if added_stations:
            added_stations = added_stations.split(',')

        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        table = 'data_min'
        if data_type:
            if int(data_type) == 2:
                table = 'data_adjust'
        conditions = db.stations.id.belongs(added_stations)
        stations = db(conditions).select(
            db.stations.id,
            db.stations.station_name,
            db.stations.frequency_receiving_data,
            limitby=limitby
        )
        iTotalRecords = db(conditions).count()
        iRow = iDisplayStart + 1
        for c, station in enumerate(stations):
            station_id = str(station.id)
            conditions = {'station_id': station_id}
            if from_date:
                if not conditions.has_key('get_time'):
                    conditions['get_time'] = {}
                conditions['get_time']['$gte'] = from_date
            if to_date:
                if not conditions.has_key('get_time'):
                    conditions['get_time'] = {}
                conditions['get_time']['$lte'] = to_date + timedelta(days=1)
            if not from_date and not to_date:
                to_date = datetime.now()
                from_date = to_date - timedelta(days=365)
                conditions['get_time'] = {'$gte': from_date, '$lte': to_date}
            delta = to_date - from_date
            if not station.frequency_receiving_data or station.frequency_receiving_data == 5:
                total_data = int((delta.days + 1) * 287)
            else:
                freq = station.frequency_receiving_data
                total_data = int((delta.days + 1) * 24 * 60 / freq)
            list_data = pydb[table].count(conditions)
            row = [
                str(iRow + c),
                station.station_name,
            ]
            content = '%s' % (list_data)
            row.append(content)
            v = float(list_data) / float(total_data) * 100
            v = "{0:.2f}".format(v)
            content = '%s' % (v)
            row.append(content)

            aaData.append(row)
        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_data_info_5():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station_for_envisoft()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
            row = db(db.stations.id.belongs(station_ids)).select(db.stations.area_id, distinct=True)
            areas_ids = [str(it.area_id) for it in row]
            areas = db(db.areas.id.belongs(areas_ids)).select(db.areas.area_name,
                                                              db.areas.id)
    areas_default = db(db.areas.id > 0).select(db.areas.area_name,
                                               db.areas.id)
    stations = db(conditions).select()
    careers = db(db.manager_careers.id > 0).select(db.manager_careers.id, db.manager_careers.career_name)

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces, areas_default=areas_default, careers=careers)


@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_data_info_5_by_station():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station_for_envisoft()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
            row = db(db.stations.id.belongs(station_ids)).select(db.stations.area_id, distinct=True)
            areas_ids = [str(it.area_id) for it in row]
            areas = db(db.areas.id.belongs(areas_ids)).select(db.areas.area_name,
                                                              db.areas.id)

    stations = db(conditions).select()
    areas_default = db(db.areas.id > 0).select(db.areas.area_name,
                                               db.areas.id)
    careers = db(db.manager_careers.id > 0).select(db.manager_careers.id, db.manager_careers.career_name)

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces, areas_default=areas_default, careers=careers)


def push_event(ticket_id, station_type, from_date, to_date, view_type, added_stations, added_columns, page):
    from gluon.scheduler import Scheduler
    primary_id = scheduler.queue_task('task_get_list_report_data_info_5_all', timeout=500,
                                      pvars=dict(ticket_id=ticket_id,
                                                 station_type=station_type,
                                                 from_date=from_date,
                                                 to_date=to_date,
                                                 view_type=view_type,
                                                 added_stations=added_stations,
                                                 page=page,
                                                 added_columns=added_columns))
    logging.info('success create queue task for task_get_list_report_data_info_5_all')
    return [primary_id.id]


################################################################################
@service.json
def get_list_report_data_info_5_in_ticket_report(*args, **kwargs):
    try:
        ticketID = request.vars.ticketid
        total_stations = int(request.vars.total_stations)
        if ticketID and total_stations:
            data = helper.getAADataByTicketReport(ticketID, db)
            if len(data) == total_stations:
                return dict(ticketID=ticketID, aaData=data, status="DONE", success=True)
            else:
                return dict(ticketID=ticketID, aaData=[],
                            message=str("wait more time"), success=True)
    except Exception as ex:
        return dict(ticketID=ticketID, aaData=[], message=str(ex), success=False)


@service.json
def push_event_create_ticket_list_report_data_info_5(*args, **kwargs):
    try:
        station_type = request.vars.station_type
        from_date = request.vars.from_date
        if from_date:
            from_date = date_util.string_to_datetime(from_date)
        to_date = request.vars.to_date
        if to_date:
            to_date = date_util.string_to_datetime(to_date)

        time_range = request.vars.time_range
        if time_range:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=int(time_range))
        view_type = request.vars.view_type
        added_stations = request.vars.added_stations or ''
        if added_stations:
            added_stations = added_stations.split(',')
        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        if not station_type:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station type for viewing data'), success=True)

        import uuid
        import math
        ticket_id = str(uuid.uuid4())
        total_page = math.ceil(len(added_stations) / 10)
        db.ticket_report.insert(ticket_id=ticket_id, total_page=total_page)
        page = 0
        for i in range(0, len(added_stations), 10):
            page += 1
            stations = added_stations[i:i + 10]
            push_event(ticket_id, station_type, request.vars.from_date, request.vars.to_date, view_type,
                       stations,
                       added_columns, page)
        return dict(ticketID=ticket_id, aaData=[], total_stations=len(added_stations), success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, ticketID="", aaData=[], total_stations=0, message=str(ex),
                    success=False)


################################################################################
@service.json
def get_list_report_data_info_5(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_type = request.vars.station_type
        from_date = request.vars.from_date
        if from_date:
            from_date = date_util.string_to_datetime(from_date)
        to_date = request.vars.to_date
        if to_date:
            to_date = date_util.string_to_datetime(to_date)
        time_range = request.vars.time_range
        if time_range:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=int(time_range))
        view_type = request.vars.view_type
        added_stations = request.vars.added_stations or ''
        if added_stations:
            added_stations = added_stations.split(',')
        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        if not station_type:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station type for viewing data'), success=True)
        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        table = 'data_min'
        conditions = db.stations.station_type == station_type
        if added_stations:
            conditions &= db.stations.id.belongs(added_stations)
        stations = db(conditions).select(
            db.stations.id,
            db.stations.station_name,
            db.stations.frequency_receiving_data,
            limitby=limitby
        )
        table_adjust = 'data_adjust'
        iTotalRecords = len(added_stations)
        iRow = iDisplayStart + 1
        for c, station in enumerate(stations):
            station_id = station.id
            delta = to_date - from_date
            if not station.frequency_receiving_data or station.frequency_receiving_data == 5:
                total_data = int((delta.days + 1) * 288)
            else:
                freq = station.frequency_receiving_data
                total_data = int((delta.days + 1) * 24 * 60 / freq)
            # station_ids = [str(item.id) for item in stations
            conditions = (db[table]['id'] > 0)
            conditions &= (db[table]['station_id'] == station_id)
            conditions_adjust = (db[table_adjust]['id'] > 0)
            conditions_adjust &= (db[table_adjust]['station_id'] == station_id)
            if from_date:
                conditions &= (db[table]['get_time'] >= from_date)
                conditions_adjust &= (db[table_adjust]['get_time'] >= from_date)
            if to_date:
                conditions &= (db[table]['get_time'] < to_date + timedelta(days=1))
                conditions_adjust &= (db[table_adjust]['get_time'] < to_date + timedelta(days=1))
            if not from_date and not to_date:
                to_date = datetime.now()
                from_date = to_date - timedelta(days=365)
                conditions &= (db[table]['get_time'] >= from_date)
                conditions &= (db[table]['get_time'] < to_date)
                conditions_adjust &= (db[table_adjust]['get_time'] >= from_date)
                conditions_adjust &= (db[table_adjust]['get_time'] < to_date)
            list_data = []
            list_data_adjust = []
            if view_type == "1":
                list_data = db(conditions).select(db[table].id,
                                                  db[table].get_time,
                                                  db[table].station_id,
                                                  db[table].data,
                                                  )
            elif view_type == "0":
                list_data_adjust = db(conditions_adjust).select(db[table_adjust].id,
                                                                db[table_adjust].get_time,
                                                                db[table_adjust].station_id,
                                                                db[table_adjust].data,
                                                                )
            # Du lieu nhan duoc
            if view_type == '1' or view_type == '0':
                count = dict()
                for i, item in enumerate(list_data):
                    if item.data:
                        for indicator_name in item.data:
                            try:
                                z = str(item.data[indicator_name])
                                if z == 'NULL' or z == 'None' or z == '-':
                                    continue
                                name_decode = indicator_name.encode('utf-8')
                                if not count.has_key(name_decode):
                                    count[name_decode] = 1
                                else:
                                    count[name_decode] = int(count[name_decode]) + 1
                            except:
                                pass
                row = [
                    str(iRow + c),
                    station.station_name,
                ]
                content = '%s %s' % (common.convert_data(total_data), T('data'))
                row.append(content)
                content = '%s (%s)' % (T('Percent Data Recieved'), '%')
                row.append(content)
                for column in added_columns:
                    if column and count.has_key(column):
                        try:
                            v = float(count[column]) / float(total_data) * 100
                            v = common.convert_data(v)
                            row.append(v)
                        except:
                            row.append(0)
                    else:
                        row.append(0)
                aaData.append(row)
            # Du lieu hop le
            if view_type == '0':
                count = dict()
                for i, item in enumerate(list_data_adjust):
                    if item.data:
                        for indicator_name in item.data:
                            try:
                                z = str(item.data[indicator_name])
                                if z == 'NULL' or z == 'None' or z == '-':
                                    continue
                                else:
                                    x = z.replace(",", "")
                                name_decode = indicator_name
                                if not count.has_key(name_decode):
                                    count[name_decode] = 1
                                else:
                                    count[name_decode] = int(count[name_decode]) + 1
                            except:
                                pass
                row = [
                    '',
                    '',
                ]
                content = '%s %s' % (common.convert_data(total_data), T('data'))
                row.append(content)
                content = '%s (%s)' % (T('Percent Data Adjusted'), '%')
                row.append(content)
                for column in added_columns:
                    if column and count.has_key(column):
                        try:
                            v = float(count[column]) / float(total_data) * 100
                            v = common.convert_data(v)
                            row.append(v)
                        except:
                            row.append(0)
                    else:
                        row.append(0)
                aaData.append(row)
            if view_type == '2':
                count = dict()
                for i, item in enumerate(list_data_adjust):
                    if item.data:
                        for indicator_name in item.data:
                            try:
                                z = str(item.data[indicator_name])
                                if z == 'NULL' or z == 'None' or z == '-':
                                    continue
                                else:
                                    x = z.replace(",", "")
                                name_decode = indicator_name
                                if not count.has_key(name_decode):
                                    count[name_decode] = 1
                                else:
                                    count[name_decode] = int(count[name_decode]) + 1
                            except:
                                pass
                row = [
                    str(iRow + c),
                    station.station_name,
                ]
                content = '%s %s' % (common.convert_data(total_data), T('data'))
                row.append(content)
                content = '%s (%s)' % (T('Percent Data Adjusted'), '%')
                row.append(content)
                for column in added_columns:
                    if column and count.has_key(column):
                        try:
                            v = float(count[column]) / float(total_data) * 100
                            v = common.convert_data(v)
                            row.append(v)
                        except:
                            row.append(0)
                    else:
                        row.append(0)
                aaData.append(row)
        return dict(ticketID=ticket_id, aaData=[], success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, ticketID=ticket_id, aaData=[], message=str(ex),
                    success=False)


@service.json
def get_list_report_data_info_5_by_station(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_type = request.vars.station_type
        from_date = request.vars.from_date
        if from_date:
            from_date = date_util.string_to_datetime(from_date)
        to_date = request.vars.to_date
        if to_date:
            to_date = date_util.string_to_datetime(to_date)

        time_range = request.vars.time_range
        if time_range:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=int(time_range))

        view_type = request.vars.view_type

        view_type = request.vars.view_type
        added_stations = request.vars.added_stations or ''
        if added_stations:
            added_stations = added_stations.split(',')
        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        if not station_type:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station type for viewing data'), success=True)
        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        table = 'data_min'
        conditions = db.stations.station_type == station_type
        if added_stations:
            conditions &= db.stations.id.belongs(added_stations)
        stations = db(conditions).select(
            db.stations.id,
            db.stations.station_name,
            db.stations.frequency_receiving_data,
            limitby=limitby
        )

        table_adjust = 'data_adjust'
        iTotalRecords = len(added_stations)
        iRow = iDisplayStart + 1
        station = stations[0]
        station_id = station.id
        delta = to_date - from_date
        if not station.frequency_receiving_data or station.frequency_receiving_data == 5:
            total_data = int((delta.days + 1) * 288)
        else:
            freq = station.frequency_receiving_data
            total_data = int((delta.days + 1) * 24 * 60 / freq)
        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        conditions_adjust = (db[table_adjust]['id'] > 0)
        conditions_adjust &= (db[table_adjust]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= from_date)
            conditions_adjust &= (db[table_adjust]['get_time'] >= from_date)
        if to_date:
            conditions &= (db[table]['get_time'] < to_date + timedelta(days=1))
            conditions_adjust &= (db[table_adjust]['get_time'] < to_date + timedelta(days=1))
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)
            conditions_adjust &= (db[table_adjust]['get_time'] >= from_date)
            conditions_adjust &= (db[table_adjust]['get_time'] < to_date)
        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].station_id,
                                          db[table].data,
                                          )

        list_data_adjust = db(conditions_adjust).select(db[table_adjust].id,
                                                        db[table_adjust].get_time,
                                                        db[table_adjust].station_id,
                                                        db[table_adjust].data,
                                                        )

        row1 = [
            "1",
            "Số giá trị quan trắc theo thiết kế",
        ]
        count = dict()
        for i, item in enumerate(list_data):
            if item.data:
                for indicator_name in item.data:
                    try:
                        z = str(item.data[indicator_name])
                        if z == 'NULL' or z == 'None' or z == '-':
                            continue
                        else:
                            x = z.replace(",", "")
                        name_decode = indicator_name.encode('utf-8')
                        if not count.has_key(name_decode):
                            count[name_decode] = 1
                        else:
                            count[name_decode] = int(count[name_decode]) + 1
                    except:
                        pass
        for column in added_columns:
            if column and count.has_key(column):
                row1.append(total_data)
            else:
                row1.append(0)
        aaData.append(row1)

        row2 = [
            "2",
            "Số giá trị quan trắc nhận được",
        ]
        for column in added_columns:
            if column and count.has_key(column):
                try:
                    v = count[column]
                    row2.append(v)
                except:
                    row2.append(0)
            else:
                row2.append(0)
        aaData.append(row2)

        count_adjust = dict()
        for i, item in enumerate(list_data_adjust):
            if item.data:
                for indicator_name in item.data:
                    try:
                        z = str(item.data[indicator_name])
                        if z == 'NULL' or z == 'None' or z == '-':
                            continue
                        else:
                            x = z.replace(",", "")
                        name_decode = indicator_name.encode('utf-8')
                        if not count_adjust.has_key(name_decode):
                            count_adjust[name_decode] = 1
                        else:
                            count_adjust[name_decode] = int(count_adjust[name_decode]) + 1
                    except:
                        pass
        row3 = [
            "3",
            "Số giá trị quan trắc lỗi",
        ]

        for column in added_columns:
            total_raw = 0
            total_adjust = 0
            if column and count.has_key(column):
                try:
                    total_raw = count[column]
                except:
                    total_raw = 0
            else:
                total_raw = 0
            if column and count_adjust.has_key(column):
                try:
                    total_adjust = count_adjust[column]
                except:
                    total_adjust = 0
            else:
                total_adjust = 0
            row3.append(str(total_raw - total_adjust))
        aaData.append(row3)

        row4 = [
            "4",
            "Số giá trị quan trắc hợp lệ"
        ]

        for column in added_columns:
            if column and count_adjust.has_key(column):
                row4.append(count_adjust[column])
            else:
                row4.append(0)
        aaData.append(row4)

        row5 = [
            "5",
            "Tỉ lệ dữ liệu nhận được so với số giá trị theo thiết kế",
        ]
        for column in added_columns:
            if column and count.has_key(column):
                try:
                    v = float(count[column]) / float(total_data) * 100
                    v = common.convert_data(v)
                    row5.append(str(v) + "%")
                except:
                    row5.append("0" + "%")
            else:
                row5.append("0" + "%")
        aaData.append(row5)

        row6 = [
            "6",
            "Tỉ lệ dữ liệu lỗi so với số giá trị theo thiết kế",
        ]

        for column in added_columns:
            total_raw = 0
            total_adjust = 0
            if column and count.has_key(column):
                try:
                    total_raw = count[column]
                except:
                    total_raw = 0
            else:
                total_raw = 0
            if column and count_adjust.has_key(column):
                try:
                    total_adjust = count_adjust[column]
                except:
                    total_adjust = 0
            else:
                total_adjust = 0
            try:
                v = float(total_raw - total_adjust) / float(total_data) * 100
                v = common.convert_data(v)
                row6.append(str(v) + "%")
            except:
                row6.append("0" + "%")
        aaData.append(row6)

        row7 = [
            "7",
            "Tỉ lệ dữ liệu hợp lệ so với số giá trị theo thiết kế",
        ]

        for column in added_columns:
            total_adjust = 0
            if column and count_adjust.has_key(column):
                try:
                    total_adjust = count_adjust[column]
                except:
                    total_adjust = 0
            else:
                total_adjust = 0
            try:
                v = float(total_adjust) / float(total_data) * 100
                v = common.convert_data(v)
                row7.append(str(v) + "%")
            except:
                row7.append("0" + "%")
        aaData.append(row7)

        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_data_hour_6():
    now = datetime.now()
    station_id = request.vars.station_id
    provinces = common.get_province_have_station_for_envisoft()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
            row = db(db.stations.id.belongs(station_ids)).select(db.stations.area_id, distinct=True)
            areas_ids = [str(it.area_id) for it in row]
            areas = db(db.areas.id.belongs(areas_ids)).select(db.areas.area_name,
                                                              db.areas.id)
    stations = db(conditions).select()
    areas_default = db(db.areas.id > 0).select(db.areas.area_name,
                                               db.areas.id)

    careers = db(db.manager_careers.id > 0).select(db.manager_careers.id, db.manager_careers.career_name)

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, station_id=station_id, default_provinces=default_provinces,
                year=now.year, month=now.month, areas_default=areas_default, careers=careers)


################################################################################
@service.json
def get_list_report_data_hour_6(*args, **kwargs):
    from applications.eos.modules.w2pex import date_util

    iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
    iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    data_type = request.vars.data_type
    year = request.vars.Year
    month = request.vars.Month
    if data_type:
        data_type = int(data_type)
    # if year:
    #     year = int(year)
    #     from_date = datetime(year=year, month=1, day=1)
    #     to_date = datetime(year=year + 1, month=1, day=1)
    # Lang update chon Thang
    if month and year:
        month = int(month)
        year = int(year)
        from_date = datetime(year=year, month=month, day=1)
        to_date = date_util.get_first_day_next_month(from_date)
        to_date = datetime.combine(to_date, datetime.min.time())
    elif year:
        year = int(year)
        from_date = datetime(year=year, month=1, day=1)
        to_date = datetime(year=year + 1, month=1, day=1)

    time_range = request.vars.time_range
    if time_range:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=int(time_range))

    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if not station_id:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station for viewing data'), success=True)
    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc
    iTotalRecords = 0  # Tong so ban ghi
    limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
    if data_type == 2:
        table = 'data_hour_adjust'
    else:
        table = 'data_hour'
    conditions = {'station_id': station_id}
    if from_date:
        if not conditions.has_key('get_time'):
            conditions['get_time'] = {}
        # conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
        conditions['get_time']['$gte'] = from_date
    if to_date:
        if not conditions.has_key('get_time'):
            conditions['get_time'] = {}
        # conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
        conditions['get_time']['$lte'] = to_date
    if not from_date and not to_date:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=365)
        conditions['get_time'] = {'$gte': from_date, '$lte': to_date}
    attrs = {'get_time': 1, 'data': 1}
    list_data = pydb[table].find(conditions, attrs).sort('get_time', -1)
    # Tong so ban ghi khong thuc hien phan trang
    iTotalRecords = 24
    # Thu tu ban ghi
    iRow = iDisplayStart + 1

    si_dict = dict()
    conditions = (db.station_indicator.station_id == station_id)
    conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
    rows = db(conditions).select(db.station_indicator.ALL)
    indicator_ids = []
    for row in rows:
        indicator_ids.append(row.indicator_id)
        si_dict[str(row.indicator_id)] = row.as_dict()
    indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                    db.indicators.unit)
    data_dict = dict()
    for i in range(0, 24):
        if not data_dict.has_key(i):
            data_dict[i] = dict()
            for indi in indicators:
                if not data_dict[i].has_key(indi.indicator):
                    data_dict[i][indi.indicator] = dict(total=0, day=0)

    for item in list_data:
        get_time = item['get_time']
        key = get_time.hour
        for indicator in indicators:
            i_name = str(indicator.indicator)
            i_name_decode = i_name.encode('utf-8')
            v = item['data'][i_name_decode] if item['data'].has_key(i_name_decode) else ''
            if v == '' or v is None:
                pass
            else:
                data_dict[key][i_name]['total'] += float(v)
                data_dict[key][i_name]['day'] += 1
    for it in data_dict:
        row = [
            it
        ]
        for i in added_columns:
            if data_dict[it][i]['day'] >0:
                row.append(round(data_dict[it][i]['total'] / data_dict[it][i]['day'], 2))
            else:
                row.append(round(0))
        aaData.append(row)
    data = []
    for i in range(iDisplayStart, iDisplayStart + iDisplayLength):
        if i < 24:
            data.append(aaData[i])
    return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=data, success=True)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_data_hour_max_7():
    import datetime
    now = datetime.datetime.now()
    # print now.year
    station_id = request.vars.station_id
    provinces = common.get_province_have_station_for_envisoft()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
            row = db(db.stations.id.belongs(station_ids)).select(db.stations.area_id, distinct=True)
            areas_ids = [str(it.area_id) for it in row]
            areas = db(db.areas.id.belongs(areas_ids)).select(db.areas.area_name,
                                                              db.areas.id)
    stations = db(conditions).select()
    areas_default = db(db.areas.id > 0).select(db.areas.area_name,
                                               db.areas.id)

    careers = db(db.manager_careers.id > 0).select(db.manager_careers.id, db.manager_careers.career_name)

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, station_id=station_id, default_provinces=default_provinces,
                year=now.year, month=now.month, areas_default=areas_default, careers=careers)


################################################################################
@service.json
def get_list_report_data_hour_max_7(*args, **kwargs):
    from applications.eos.modules.w2pex import date_util
    from calendar import monthrange
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_id = request.vars.station_id
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        data_type = request.vars.data_type
        if data_type:
            data_type = int(data_type)
        month = request.vars.Month
        year = request.vars.Year
        if month and year:
            month = int(month)
            year = int(year)
            from_date = datetime(year=year, month=month, day=1)
            to_date = date_util.get_first_day_next_month(from_date)
            to_date = datetime.combine(to_date, datetime.min.time())
        elif year:
            year = int(year)
            from_date = datetime(year=year, month=1, day=1)
            to_date = datetime(year=year + 1, month=1, day=1)

        time_range = request.vars.time_range
        if time_range:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=int(time_range))

        added_columns = request.vars.added_columns or ''

        if added_columns:
            added_columns = added_columns.split(',')
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        if data_type == 2:
            table = 'data_hour_adjust'
        else:
            table = 'data_hour'
        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= from_date)
        if to_date:
            conditions &= (db[table]['get_time'] < to_date)

        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].data,
                                          orderby=~db[table].get_time)
        # Thu tu ban ghi
        iRow = iDisplayStart + 1
        if month:
            iTotalRecords = monthrange(year, month)[1]
        else:
            iTotalRecords = (to_date - from_date).days
        si_dict = dict()
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db(conditions).select(db.station_indicator.ALL)
        indicator_ids = []
        for row in rows:
            indicator_ids.append(row.indicator_id)
            si_dict[str(row.indicator_id)] = row.as_dict()
        indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                        db.indicators.unit)

        data_final = dict()

        for item in list_data:
            key = datetime.strftime(item.get_time, '%d/%m/%Y')
            if not data_final.has_key(key):
                data_final[key] = dict()
                for indicator in indicators:
                    i_name = str(indicator.indicator)
                    i_name_decode = i_name.encode('utf-8')
                    v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                    if v == '' or v is None:
                        pass
                    else:
                        try:
                            v = float(v)
                            if not data_final[key].has_key(i_name):
                                data_final["has_max"] = true
                                data_final[key][i_name] = "{0:.2f}".format(v)
                            else:
                                if v > float(data_final[key][i_name]):
                                    data_final["has_max"] = true
                                    data_final[key][i_name] = "{0:.2f}".format(v)
                        except:
                            pass
            else:
                for indicator in indicators:
                    i_name = str(indicator.indicator)
                    i_name_decode = i_name.encode('utf-8')
                    v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                    if v == '' or v is None:
                        pass
                    else:
                        try:
                            v = float(v)
                            if not data_final[key].has_key(i_name):
                                data_final[key][i_name] = "{0:.2f}".format(v)
                            else:
                                if v > float(data_final[key][i_name]):
                                    data_final[key][i_name] = "{0:.2f}".format(v)
                        except:
                            pass
        delta = timedelta(days=1)
        while from_date < to_date:
            key = datetime.strftime(from_date, "%d/%m/%Y")
            row = [
                key
            ]
            if not data_final.has_key(key):
                data_final[key] = dict()
                for column in added_columns:
                    row.append('-')
            else:
                for column in added_columns:
                    if column and data_final[key].has_key(column):
                        row.append(data_final[key][column])
                    else:
                        row.append('-')
            aaData.append(row)
            from_date += delta
        data = []
        for i_count in range(iDisplayStart, iDisplayStart + iDisplayLength):
            if i_count < iTotalRecords:
                data.append(aaData[i_count])
        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=data, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_data_day_8():
    import datetime
    now = datetime.datetime.now()
    station_id = request.vars.station_id
    provinces = common.get_province_have_station_for_envisoft()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()
    areas_default = db(db.areas.id > 0).select(db.areas.area_name,
                                               db.areas.id)

    careers = db(db.manager_careers.id > 0).select(db.manager_careers.id, db.manager_careers.career_name)

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, station_id=station_id, default_provinces=default_provinces,
                year=now.year, month=now.month, areas_default=areas_default, careers=careers,
                )


################################################################################
@service.json
def get_list_report_data_day_8(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_id = request.vars.station_id
        from_date = request.vars.from_date
        data_type = request.vars.data_type

        if data_type:
            data_type = int(data_type)
        to_date = request.vars.to_date
        month = request.vars.Month
        year = request.vars.Year
        if month and year:
            month = int(month)
            year = int(year)
            from_date = datetime(year=year, month=month, day=1)
            to_date = date_util.get_first_day_next_month(from_date)
            to_date = datetime.combine(to_date, datetime.min.time())
        elif year:
            year = int(year)
            from_date = datetime(year=year, month=1, day=1)
            to_date = datetime(year=year + 1, month=1, day=1)

        time_range = request.vars.time_range
        if time_range:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=int(time_range))

        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        if data_type == 2:
            table = 'data_day_adjust'
        else:
            table = 'data_day'

        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= from_date)
        if to_date:
            conditions &= (db[table]['get_time'] < to_date)
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)

        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].data,
                                          orderby=db[table].get_time,
                                          limitby=limitby)
        # Tong so ban ghi khong thuc hien phan trang
        iTotalRecords = db(conditions).count(db[table].id)
        # Thu tu ban ghi
        iRow = iDisplayStart + 1

        si_dict = dict()
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db().select(db.station_indicator.ALL)
        indicator_ids = []
        for row in rows:
            indicator_ids.append(row.indicator_id)
            si_dict[str(row.indicator_id)] = row.as_dict()
        indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                        db.indicators.unit)

        # Duyet tung phan tu trong mang du lieu vua truy van duoc
        for i, item in enumerate(list_data):
            added_item = dict()
            if added_columns:
                for indicator in indicators:
                    i_name = str(indicator.indicator)
                    i_name_decode = i_name.encode('utf-8')
                    v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                    if v == '' or v is None:
                        added_item[i_name] = '-'
                    else:
                        try:
                            v = float(v)
                            added_item[i_name] = "{0:.2f}".format(v)
                        except:
                            added_item[i_name] = '-'
            row = [
                str(iRow + i),
                item.get_time.strftime('%d/%m/%Y'),
            ]
            for column in added_columns:
                if column and added_item.has_key(column):
                    row.append(added_item[column])
            aaData.append(row)
        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_data_hour_8h_max_9():
    import datetime
    now = datetime.datetime.now()
    # print now.year
    station_id = request.vars.station_id
    provinces = common.get_province_have_station_for_envisoft()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()
    areas_default = db(db.areas.id > 0).select(db.areas.area_name,
                                               db.areas.id)
    careers = db(db.manager_careers.id > 0).select(db.manager_careers.id, db.manager_careers.career_name)

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, station_id=station_id, default_provinces=default_provinces,
                year=now.year, month=now.month, areas_default=areas_default, careers=careers)


################################################################################
@service.json
def get_list_report_data_hour_8h_max_9(*args, **kwargs):
    from applications.eos.modules.w2pex import date_util
    from calendar import monthrange
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_id = request.vars.station_id
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        data_type = request.vars.data_type
        if data_type:
            data_type = int(data_type)
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        month = request.vars.Month
        year = request.vars.Year
        if month and year:
            month = int(month)
            year = int(year)
            from_date = datetime(year=year, month=month, day=1)
            to_date = date_util.get_first_day_next_month(from_date)
            to_date = datetime.combine(to_date, datetime.min.time())

        time_range = request.vars.time_range
        if time_range:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=int(time_range))

        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')

        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        if data_type == 2:
            table = 'data_hour_8h_adjust'
        else:
            table = 'data_hour_8h'

        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= from_date)
        if to_date:
            conditions &= (db[table]['get_time'] < to_date)

        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].data,
                                          orderby=~db[table].get_time)
        # Tong so ban ghi khong thuc hien phan trang
        iTotalRecords = monthrange(year, month)[1]
        # Thu tu ban ghi
        iRow = iDisplayStart + 1

        si_dict = dict()
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db().select(db.station_indicator.ALL)
        indicator_ids = []
        for row in rows:
            indicator_ids.append(row.indicator_id)
            si_dict[str(row.indicator_id)] = row.as_dict()
        indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                        db.indicators.unit)

        for i_count in range(iDisplayStart + 1, iDisplayLength + 1):
            if i_count <= iTotalRecords:
                added_item = dict()
                for i, item in enumerate(list_data):
                    if item.get_time.day == i_count:
                        if added_columns:
                            for indicator in indicators:
                                i_name = str(indicator.indicator)
                                i_name_decode = i_name.encode('utf-8')
                                v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                                if v == '' or v is None:
                                    pass
                                else:
                                    try:
                                        v = float(v)
                                        if not added_item.has_key(i_name):
                                            added_item[i_name] = "{0:.2f}".format(v)
                                        else:
                                            if v > float(added_item[i_name]):
                                                added_item[i_name] = "{0:.2f}".format(v)
                                    except:
                                        pass
                row = [
                    str(i_count) + '/' + str(month),
                ]
                for column in added_columns:
                    if column and added_item.has_key(column):
                        row.append(added_item[column])
                    else:
                        row.append('-')
                aaData.append(row)
        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_aqi_10():
    now = datetime.now()
    station_id = request.vars.station_id
    provinces = common.get_province_have_station_for_envisoft()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()

    areas_default = db(db.areas.id > 0).select(db.areas.area_name,
                                               db.areas.id)
    careers = db(db.manager_careers.id > 0).select(db.manager_careers.id, db.manager_careers.career_name)
    return dict(provinces=provinces, stations=stations, station_id=station_id, default_provinces=default_provinces,
                year=now.year, areas_default=areas_default, careers=careers, month=now.month)


################################################################################
@service.json
def get_list_report_aqi_10(*args, **kwargs):
    from applications.eos.modules.w2pex import date_util
    from calendar import monthrange
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_id = request.vars.station_id
        data_type = request.vars.data_type
        if data_type:
            data_type = int(data_type)
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        month = request.vars.Month
        year = request.vars.Year
        if month and year:
            month = int(month)
            year = int(year)
            from_date = datetime(year=year, month=month, day=1)
            to_date = date_util.get_first_day_next_month(from_date)
            to_date = datetime.combine(to_date, datetime.min.time())

        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        aaData = []  # Du lieu json se tra ve
        # lay du lieu aqi hour
        if data_type == 2:
            table = 'aqi_data_adjust_hour'
        else:
            table = 'aqi_data_hour'
        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        time_range = request.vars.time_range
        if time_range:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=int(time_range))

        if from_date:
            conditions &= (db[table]['get_time'] >= from_date)
        if to_date:
            conditions &= (db[table]['get_time'] < to_date)
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)

        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].data,
                                          orderby=~db[table].get_time)
        # Tong so ban ghi khong thuc hien phan trang

        aqi_hour = dict()
        # Duyet tung phan tu trong mang du lieu vua truy van duoc
        for i, item in enumerate(list_data):
            day = item.get_time.day
            hour = item.get_time.hour
            key = (day, hour)
            v = item.data['aqi'] if item.data.has_key('aqi') else ''
            if v == '' or v is None:
                aqi_hour[key] = '-'
            else:
                try:
                    v = float(v)
                    aqi_hour[key] = "{0:.2f}".format(v)
                except:
                    aqi_hour[key] = '-'
        #####################################################
        # lay du lieu aqi day
        if data_type == 2:
            table = 'aqi_data_adjust_24h'
        else:
            table = 'aqi_data_24h'
        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= from_date)
        if to_date:
            conditions &= (db[table]['get_time'] < to_date)
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)

        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].data_24h,
                                          orderby=~db[table].get_time)
        # Tong so ban ghi khong thuc hien phan trang

        aqi_day = dict()
        # Duyet tung phan tu trong mang du lieu vua truy van duoc
        for i, item in enumerate(list_data):
            day = item.get_time.day
            v = item.data_24h['aqi'] if item.data_24h.has_key('aqi') else ''
            # v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
            if v == '' or v is None:
                aqi_day[day] = '-'
            else:
                try:
                    v = float(v)
                    aqi_day[day] = "{0:.2f}".format(v)
                except:
                    aqi_day[day] = '-'

        #####################################################
        iTotalRecords = iTotalRecords = monthrange(year, month)[1]
        # Thu tu ban ghi
        iRow = iDisplayStart + 1
        for i_count in range(iDisplayStart + 1, iDisplayLength + 1):
            if i_count <= iTotalRecords:
                row = [
                    str(i_count),
                ]
                for j in range(0, 24):
                    key = (i_count, j)
                    if aqi_hour.has_key(key):
                        row.append(aqi_hour[key])
                    else:
                        row.append('-')
                if aqi_day.has_key(i_count):
                    row.append(aqi_day[i_count])
                else:
                    row.append('-')
                aaData.append(row)

        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


@service.json
def get_list_report_aqi_by_time(*args, **kwargs):
    station_id = request.vars.station_id
    by_time = request.vars.by_time
    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)
    year = request.vars.Year
    from datetime import datetime as ddd
    start_date = ddd(year=int(year), month=1, day=1)
    end_date = ddd(year=int(year), month=12, day=31)

    quarter = request.vars.quarter
    if quarter:
        if quarter == "1":
            start_date = ddd(year=int(year), month=1, day=1)
            end_date = ddd(year=int(year), month=3, day=31)
        elif quarter == "2":
            start_date = ddd(year=int(year), month=4, day=1)
            end_date = ddd(year=int(year), month=6, day=30)
        elif quarter == "3":
            start_date = ddd(year=int(year), month=7, day=1)
            end_date = ddd(year=int(year), month=9, day=30)
        elif quarter == "4":
            start_date = ddd(year=int(year), month=10, day=1)
            end_date = ddd(year=int(year), month=12, day=31)

    from applications.eos.services import aqi_service
    result = aqi_service.AqiService(pydb, T, db).get_report_aqi_day_in_time(str(station_id), int(data_type), start_date,
                                                                            end_date)
    return dict(aaData=result, success=True)


@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_aqi_day_in_time():
    now = datetime.now()
    station_id = request.vars.station_id
    provinces = common.get_province_have_station_for_envisoft()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()

    areas_default = db(db.areas.id > 0).select(db.areas.area_name,
                                               db.areas.id)

    careers = db(db.manager_careers.id > 0).select(db.manager_careers.id, db.manager_careers.career_name)
    return dict(provinces=provinces, stations=stations, station_id=station_id, default_provinces=default_provinces,
                year=now.year, areas_default=areas_default, careers=careers, month=now.month)


@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_public_aqi_station():
    return dict(aaData=[])


@service.json
def call_report_public_aqi_station(*args, **kwargs):
    from_date = request.vars.from_date
    end_date = request.vars.end_date
    from applications.eos.services import aqi_service
    result = []
    time_range = request.vars.time_range
    if time_range:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=int(time_range))

    if from_date is not None and end_date is not None:
        result = aqi_service.AqiService(pydb, T, db).get_report_public_aqi_station(
            date_util.string_to_datetime(from_date), date_util.string_to_datetime(end_date) + timedelta(days=1))

    return dict(iTotalRecords=1, iTotalDisplayRecords=1, aaData=result, success=True)


@service.json
def export_excel_report_public_aqi_station():
    from openpyxl.styles import Alignment
    import os.path, openpyxl

    from applications.eos.services import aqi_service
    aaData = aqi_service.AqiService(pydb, T, db).get_report_public_aqi_station()

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    # headers_1 = []
    # headers_2 = []

    # headers = []
    # headers.append('#')
    # headers_1.append(headers)
    #
    # # headers = []
    # # headers.append('#')
    # headers_2.append(headers)
    #
    # for header in headers_1:
    #     ws2.append(header)
    # for header in headers_2:
    #     ws2.append(header)
    for row in aaData:
        ws2.append(row)
    # ws2.merge_cells('A1:A2')
    # ws2.merge_cells('Z1:Z2')
    # ws2.merge_cells('B1:Y1')
    for col in ws2.columns:
        for cell in col:
            alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
            cell.alignment = alignment_obj

    file_name = request.now.strftime('report_public_aqi_station' + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_data_month_11():
    import datetime
    now = datetime.datetime.now()
    station_id = request.vars.station_id
    provinces = common.get_province_have_station_for_envisoft()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()
    areas_default = db(db.areas.id > 0).select(db.areas.area_name,
                                               db.areas.id)

    careers = db(db.manager_careers.id > 0).select(db.manager_careers.id, db.manager_careers.career_name)

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, station_id=station_id, default_provinces=default_provinces,
                year=now.year, month=now.month, areas_default=areas_default, careers=careers)


################################################################################
@service.json
def get_list_report_data_month_11(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_id = request.vars.station_id
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        data_type = request.vars.data_type
        if data_type:
            data_type = int(data_type)
        # month = request.vars.Month
        year = request.vars.Year
        # if month and year:
        if year:
            # month = int(month)
            year = int(year)
            from_date = datetime(year=year, month=1, day=1)
            to_date = datetime(year=year + 1, month=1, day=1)

        time_range = request.vars.time_range
        if time_range:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=int(time_range))

        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        # iTotalRecords = 0  # Tong so ban ghi
        # limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        if data_type == 2:
            table = 'data_mon_adjust'
        else:
            table = 'data_mon'

        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= from_date)
        if to_date:
            conditions &= (db[table]['get_time'] < to_date)
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)

        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].data,
                                          orderby=db[table].get_time)
        # Tong so ban ghi khong thuc hien phan trang
        iTotalRecords = 1
        # Thu tu ban ghi
        iRow = iDisplayStart + 1

        si_dict = dict()
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db().select(db.station_indicator.ALL)
        indicator_ids = []
        for row in rows:
            indicator_ids.append(row.indicator_id)
            si_dict[str(row.indicator_id)] = row.as_dict()
        indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                        db.indicators.unit)
        result_dict = dict()
        total_result = dict()
        count_result = dict()
        # Duyet tung phan tu trong mang du lieu vua truy van duoc
        for i, item in enumerate(list_data):
            added_item = dict()
            if added_columns:
                for indicator in indicators:
                    i_name = str(indicator.indicator)
                    i_name_decode = i_name.encode('utf-8')
                    v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                    if v == '' or v is None:
                        added_item[i_name] = '-'
                    else:
                        try:
                            v = float(v)
                            added_item[i_name] = "{0:.2f}".format(v)
                            if total_result.has_key(i_name):
                                total_result[i_name] += v
                                count_result[i_name] += 1
                            else:
                                total_result[i_name] = v
                                count_result[i_name] = 1
                        except:
                            added_item[i_name] = '-'
            row = [
                item.get_time.month
            ]
            for column in added_columns:
                if column and added_item.has_key(column):
                    row.append(added_item[column])
            result_dict[item.get_time.month] = row
        for i in range(1, 13):
            if not result_dict.has_key(i):
                row = [
                    i,
                ]
                for column in added_columns:
                    row.append('-')
                aaData.append(row)
            else:
                aaData.append(result_dict[i])
        row = [
            T('AVG Year'),
        ]
        for column in added_columns:
            if total_result.has_key(column):
                avg = float(total_result[column]) / count_result[column]
                row.append("{0:.2f}".format(avg))
            else:
                row.append('-')
        aaData.append(row)
        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_lost_indicator_12():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station_for_envisoft()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces)


################################################################################
################################################################################
@service.json
def get_list_report_lost_indicator_12(*args, **kwargs):
    try:
        station_id = request.vars.station_id
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        indicator_array = request.vars.indicators
        connection_loss = request.vars.connection_loss

        if indicator_array:
            indicator_array = indicator_array.split(',')
        discontinuity_type = request.vars.discontinuity_type
        aaData = []
        iDisplayStart = int(request.vars.iDisplayStart)
        iDisplayLength = int(request.vars.iDisplayLength)

        condition_fetch_indi = db.station_indicator.station_id == station_id
        condition_fetch_indi &= db.station_indicator.status == 1
        indicator_station = db(condition_fetch_indi).select(db.station_indicator.ALL)
        indicator_ids = [str(it.indicator_id) for it in indicator_station]
        condition_fetch_indi_1 = db.indicators.id.belongs(indicator_ids)
        condition_fetch_indi_1 &= db.indicators.indicator.belongs(indicator_array)
        indicator_fetch = db(condition_fetch_indi_1).select(db.indicators.indicator, db.indicators.id)
        indicator_ids = [str(i.id) for i in indicator_fetch]
        data_final = list()
        # disconnect
        if discontinuity_type in ['disconnect_type', 'all']:
            condition_search = get_condition_report_12(station_id, from_date, to_date,
                                                       'station_off_log', 'disconnect_type', 0)
            if connection_loss:
                condition_search &= db.station_off_log.duration >= float(connection_loss) * 3600
            data = db(condition_search).select(db.station_off_log.start_off,
                                               db.station_off_log.end_off,
                                               db.station_off_log.duration,
                                               orderby=~db.station_off_log.start_off)
            if data:
                for c, item in enumerate(data):
                    if connection_loss:
                        if item.duration >= float(connection_loss) * 3600:
                            row = [
                                c + 1,
                                'Mất kết nối',
                                indicator_array,
                                item.start_off,
                                item.end_off,
                                common.format_passed_time(item.duration),
                            ]
                            aaData.append(row)
                    else:
                        row = [
                            c + 1,
                            'Mất kết nối',
                            indicator_array,
                            item.start_off,
                            item.end_off,
                            common.format_passed_time(item.duration),
                        ]
                        aaData.append(row)
        # sensor error
        if discontinuity_type in ['device_failure', 'all'] and (connection_loss == ''):
            condition_search = get_condition_report_12(station_id, from_date, to_date,
                                                       'sensor_trouble_history', 'device_failure', 2)
            condition_search &= db.sensor_trouble_history.indicator_id.belongs(indicator_ids)
            data = db(condition_search).select(db.sensor_trouble_history.indicator_name,
                                               db.sensor_trouble_history.indicator_id,
                                               db.sensor_trouble_history.get_time,
                                               orderby=~db.sensor_trouble_history.get_time)
            max_off_logs = len(aaData)
            data_sensor_trouble_history = dict()
            if data:
                for item in data:
                    time = item.get_time - timedelta(minutes=5)
                    if not data_sensor_trouble_history.has_key(item.indicator_id):
                        data_sensor_trouble_history[item.indicator_id] = {
                            item.get_time: dict(indicator_name=item.indicator_name,
                                                end_off=item.get_time + timedelta(minutes=5),
                                                start_off=item.get_time,
                                                duration=300,
                                                type='Lỗi thiết bị')}
                    else:
                        if data_sensor_trouble_history[item.indicator_id].has_key(time):
                            data_sensor_trouble_history[item.indicator_id][time]['end_off'] = item.get_time + timedelta(
                                minutes=5)
                            data_sensor_trouble_history[item.indicator_id][time]['duration'] += 300
                            data_sensor_trouble_history[item.indicator_id][item.get_time] = data_sensor_trouble_history[
                                item.indicator_id].pop(time)

                        else:
                            data_sensor_trouble_history[item.indicator_id][item.get_time] = dict(
                                indicator_name=item.indicator_name,
                                end_off=item.get_time + timedelta(minutes=5), start_off=item.get_time, duration=300,
                                type='Lỗi thiết bị')
            inx = 0
            for indicator in data_sensor_trouble_history:
                data = data_sensor_trouble_history[indicator]
                for i, key in enumerate(sorted(data)):
                    row = [
                        i + max_off_logs + 1 + inx,
                        data[key]['type'],
                        data[key]['indicator_name'],
                        data[key]['start_off'],
                        data[key]['end_off'],
                        data[key]['duration'],
                    ]
                    aaData.append(row)
                inx = len(aaData) - 1
        # Calibration
        if discontinuity_type in ['calibration_type', 'all']:
            condition_search = get_condition_report_12(station_id, from_date, to_date,
                                                       'sensor_trouble_history', 'calibration_type', 1)
            condition_search &= db.sensor_trouble_history.indicator_id.belongs(indicator_ids)
            data = db(condition_search).select(db.sensor_trouble_history.indicator_name,
                                               db.sensor_trouble_history.indicator_id,
                                               db.sensor_trouble_history.get_time,
                                               orderby=~db.sensor_trouble_history.get_time)
            max_sensor_error = len(aaData)
            data_calib = dict()
            if data:
                for item in data:
                    time = item.get_time - timedelta(minutes=5)
                    if not data_calib.has_key(item.indicator_id):
                        data_calib[item.indicator_id] = {
                            item.get_time: dict(indicator_name=item.indicator_name,
                                                end_off=item.get_time + timedelta(minutes=5),
                                                start_off=item.get_time,
                                                duration=300,
                                                type='Hiệu chuẩn')}
                    else:
                        if data_calib[item.indicator_id].has_key(time):
                            data_calib[item.indicator_id][time]['end_off'] = item.get_time + timedelta(minutes=5),
                            data_calib[item.indicator_id][time]['duration'] += 300
                            data_calib[item.indicator_id][item.get_time] = data_calib[
                                item.indicator_id].pop(time)

                        else:
                            data_calib[item.indicator_id][item.get_time] = dict(
                                indicator_name=item.indicator_name,
                                end_off=item.get_time + timedelta(minutes=5), start_off=item.get_time, duration=300,
                                type='Hiệu chuẩn')
            inx = 0
            for indicator in data_calib:
                data = data_calib[indicator]
                for i, key in enumerate(sorted(data)):
                    row = [
                        i + max_sensor_error + 1 + inx,
                        data[key]['type'],
                        data[key]['indicator_name'],
                        data[key]['start_off'],
                        data[key]['end_off'],
                        data[key]['duration'],
                    ]
                    aaData.append(row)
                inx = len(aaData)
        iTotalRecords = len(aaData)
        aaData = aaData[iDisplayStart:iDisplayStart + iDisplayLength]
        return dict(success=True, iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_data_received_ratio_13():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station_for_envisoft()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    areas = db(db.areas.id > 0).select()
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
            row = db(db.stations.id.belongs(station_ids)).select(db.stations.area_id, distinct=True)
            areas_ids = [str(it.area_id) for it in row]
            areas = db(db.areas.id.belongs(areas_ids)).select(db.areas.area_name,
                                                              db.areas.id)
    # stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, view_type=view_type, default_provinces=default_provinces, areas=areas)


###############################################################################
@service.json
def get_list_report_data_received_ratio_13(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        iDisplayEnd = iDisplayStart + iDisplayLength
        station_type = request.vars.station_type
        area_id = request.vars.area_id
        from_date = request.vars.from_date
        data_type = request.vars.data_type
        if from_date:
            from_date = date_util.string_to_datetime(from_date)
        to_date = request.vars.to_date
        if to_date:
            to_date = date_util.string_to_datetime(to_date)
        added_provinces = request.vars.added_stations or ''
        if not added_provinces:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=T('No Province choosen!'),
                        success=True)
        added_provinces = added_provinces.split(',')
        added_provinces_for_page = added_provinces[iDisplayStart: iDisplayEnd]

        result_o = dict()
        inx = 0
        for it in added_provinces:
            result_o[it] = inx
            inx += 1

        aaData = []  # Du lieu json se tra ve
        table = 'data_min'
        if data_type:
            if int(data_type) == 2:
                table = 'data_adjust'
        conditions = (db.stations.id > 0)
        if station_type:
            conditions &= (db.stations.station_type == station_type)
        if area_id:
            conditions &= (db.stations.area_id == area_id)
        if added_provinces_for_page:
            conditions &= (db.stations.province_id.belongs(added_provinces_for_page))
        stations = db(conditions).select(
            db.stations.id,
            db.stations.station_name,
            db.stations.province_id,
            db.stations.station_type,
            db.stations.frequency_receiving_data
        )
        provinces = dict()
        # Xu ly du lieu tinh
        if added_provinces:
            province_tmp = db(db.provinces.id.belongs(added_provinces)).select(db.provinces.province_name,
                                                                               db.provinces.id)
            for p in province_tmp:
                provinces[str(p.id)] = p.province_name
        result_dict = {}

        ids = [str(it.id) for it in stations]
        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'].belongs(ids))
        if from_date:
            conditions &= (db[table]['get_time'] >= from_date)
        if to_date:
            conditions &= (db[table]['get_time'] < to_date + timedelta(days=1))
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=30)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)
        count_func = db[table]['station_id'].count()
        try:
            data = db(conditions).select(db[table]['station_id'],
                                         db[table]['is_exceed'],
                                         count_func,
                                         groupby=db[table]['station_id'])
        except:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=T('No data'), success=True)
        data_rs = {}
        for item in data:
            if not data_rs.has_key(item[table]['station_id']):
                data_rs[item[table]['station_id']] = {'t': 0, 'e': 0}
            for k in item['_extra']:
                if item[table]['is_exceed']:
                    data_rs[item[table]['station_id']]['e'] = item['_extra'][k]
                data_rs[item[table]['station_id']]['t'] += item['_extra'][k]
        # duyet tren danh sach tram tra ve ket qua cuoi cung
        for station in stations:
            province_id = str(station.province_id)
            station_id = str(station.id)
            delta = to_date - from_date
            if not station.frequency_receiving_data or station.frequency_receiving_data == 5:
                total_data = int((delta.days + 1) * 287)
            else:
                freq = station.frequency_receiving_data
                total_data = int((delta.days + 1) * 24 * 60 / freq)

            if not result_dict.has_key(province_id):
                province_name = provinces[province_id] if provinces.has_key(province_id) else ''
                result_dict[province_id] = {
                    'order': result_o[province_id],
                    'name': province_name,
                    'total': 0,
                    'stations': [],
                    'info': {
                        'total': 0,  # Du lieu ly tuong
                        'received': 0,  # Du lieu nhan duoc
                        'exceed': 0,  # du lieu vuot chuan
                    },
                }
            r_station = {'name': station.station_name, 'total': total_data, 'received': 0, 'exceed': 0}
            if data_rs.has_key(station_id):
                r_station['received'] = data_rs[station_id]['t']
                r_station['exceed'] = data_rs[station_id]['e']
                result_dict[province_id]['info']['received'] += data_rs[station_id]['t']
                result_dict[province_id]['info']['exceed'] += data_rs[station_id]['e']
            result_dict[province_id]['info']['total'] += total_data  # Du lieu ly tuong
            result_dict[province_id]['stations'].append(r_station)
            result_dict[province_id]['total'] += 1
        iTotalRecords = len(added_provinces)
        inx = 0
        ls = result_dict.values()

        def func_s(e):
            return e['order']

        ls.sort(key=func_s)
        for item in ls:
            inx += 1
            row = [
                str(iDisplayStart + inx),
                item['name'],
            ]
            row.append(item['info']['exceed'])
            content = '%s' % (item['info']['received'])
            row.append(content)
            ratio = float(item['info']['received']) / float(item['info']['total']) * 100
            ratio = "{0:.2f}".format(ratio)
            content = '%s' % (ratio)
            row.append(content)
            aaData.append(row)

        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_data_status_history_14():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station_for_envisoft()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    areas = db(db.areas.id > 0).select()
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
            row = db(db.stations.id.belongs(station_ids)).select(db.stations.area_id, distinct=True)
            areas_ids = [str(it.area_id) for it in row]
            areas = db(db.areas.id.belongs(areas_ids)).select(db.areas.area_name,
                                                              db.areas.id)
    # stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, view_type=view_type, default_provinces=default_provinces, areas=areas)


###############################################################################
@service.json
def get_list_report_data_status_history_14(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        iDisplayEnd = iDisplayStart + iDisplayLength
        station_type = request.vars.station_type
        area_id = request.vars.area_id
        to_date = request.vars.to_date
        added_provinces = request.vars.added_stations or ''
        if not to_date:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=T('Date'), success=True)
        if not added_provinces:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=T('No Province choosen!'),
                        success=True)
        if to_date:
            if to_date.find('/') == -1:
                to_date = datetime.strptime(to_date, '%Y-%m-%d')
            else:
                to_date = to_date.split(':')[0]
                to_date = datetime.strptime(to_date, '%Y/%m/%d %H')
        if added_provinces:
            added_provinces = added_provinces.split(',')
        result_o = dict()
        inx = 0
        for it in added_provinces:
            result_o[it] = inx
            inx += 1
        aaData = []  # Du lieu json se tra ve
        table = 'data_hour_status_history'
        # conditions = (db[table]['id'] > 0)
        if to_date:
            conditions = (db[table]['time'] == to_date)
        stations = {}
        data = db(conditions).select(db[table].data, orderby=db[table]._id)
        if data:
            stations = data[0]['data']
        provinces = dict()
        # Xu ly du lieu tinh
        if added_provinces:
            province_tmp = db(db.provinces.id.belongs(added_provinces)).select(db.provinces.province_name,
                                                                               db.provinces.id)
            for p in province_tmp:
                provinces[str(p.id)] = p.province_name

        result_dict = {}
        for province_id in provinces:
            province_name = provinces[province_id] if provinces.has_key(province_id) else ''
            result_dict[province_id] = {
                'order': result_o[province_id],
                'name': province_name,
                'total': 0,
                'stations': [],
                'info': {
                    'total': 0,  # Tong tram
                    'connect': 0,  # Tong tram ket noi
                    'adjusting': 0,  # Tong tram hieu chuan
                    'exceed': 0,  # Tong tram vuot nguong
                    'disconnect': 0,  # Tong tram mat ket noi
                    'sensor_error': 0,
                },
            }
        # duyet tren danh sach tram tra ve ket qua cuoi cung
        for station in stations:
            station_id = str(station)
            is_get_station = True
            if station_type:
                if str(stations[station]['station_type']) != station_type:
                    is_get_station = False
            if area_id:
                if str(stations[station]['area_id']) != area_id:
                    is_get_station = False
            if added_provinces:
                if not str(stations[station]['province_id']) in added_provinces:
                    is_get_station = False
            if is_get_station:
                province_id = str(stations[station]['province_id'])
                status = stations[station]['status']
                r_station = {'id': station_id, 'status': status}
                if status == 4:
                    result_dict[province_id]['info']['disconnect'] += 1
                else:
                    if status == 3:
                        result_dict[province_id]['info']['exceed'] += 1
                    elif status == 5:
                        result_dict[province_id]['info']['adjusting'] += 1
                    elif status == 6:
                        result_dict[province_id]['info']['sensor_error'] += 1
                    result_dict[province_id]['info']['connect'] += 1
                result_dict[province_id]['stations'].append(r_station)
                result_dict[province_id]['info']['total'] += 1
        iTotalRecords = len(provinces)
        inx = 0
        ls = result_dict.values()

        def func_s(e):
            return e['order']

        ls.sort(key=func_s)
        ls = ls[iDisplayStart:iDisplayEnd]

        for item in ls:
            inx += 1
            row = [str(iDisplayStart + inx), item['name']]
            row.append(item['info']['total'])
            row.append(item['info']['connect'])
            row.append(item['info']['adjusting'])
            row.append(item['info']['sensor_error'])
            row.append(item['info']['exceed'])
            row.append(item['info']['disconnect'])
            if item['info']['total'] != 0:
                ratio = float(item['info']['connect']) / float(item['info']['total']) * 100
            else:
                ratio = 0.00
            content = '%s (%s)' % (common.convert_data(ratio), '%')
            row.append(content)
            aaData.append(row)

        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


################################################################################

@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_widget_data_collect_15():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    provinces = common.get_province_have_station_for_envisoft()
    areas = db(db.areas.id > 0).select()
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
            row = db(db.stations.id.belongs(station_ids)).select(db.stations.area_id, distinct=True)
            areas_ids = [str(it.area_id) for it in row]
            areas = db(db.areas.id.belongs(areas_ids)).select(db.areas.area_name,
                                                              db.areas.id)

    return dict(provinces=provinces, view_type=view_type, areas=areas)


################################################################################
@service.json
def get_list_report_widget_data_collect_15(*args, **kwargs):
    from w2pex import date_util
    from datetime import datetime, date
    try:
        area_id = request.vars.area_id
        province_id = request.vars.province_id
        date_type = request.vars.date_type
        if not province_id and not area_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station type or an area for viewing data'), success=True)
        n_qty_this_month = 0
        expected_this_month = 0
        n_qty_last_month = 0
        expected_last_month = 0
        conditions = (db.stations.id > 0)
        if province_id and province_id != 'all_provinces':
            conditions &= (db.stations.province_id == province_id)
        if area_id:
            conditions &= (db.stations.area_id == area_id)
        # hungdx phan quyen quan ly trạm theo user issue 44
        if current_user:
            if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
                list_station_manager = db(db.manager_stations.user_id == current_user.id).select(
                    db.manager_stations.station_id)
                station_ids = [str(item.station_id) for item in list_station_manager]
                conditions &= (db.stations.id.belongs(station_ids))

        # station_ids = []
        # if area_id or province_id:
        ids = db(conditions).select(db.stations.id)
        station_ids = [str(item.id) for item in ids]

        if date_type == "" or date_type == "month":
            first_date_in_this_month = date_util.get_first_day_current_month(date.today())
            first_date_in_this_month = datetime.combine(first_date_in_this_month, datetime.min.time())
            first_date_in_last_month = date_util.get_first_day_last_month(date.today())
            first_date_in_last_month = datetime.combine(first_date_in_last_month, datetime.min.time())
            first_date_in_next_month = date_util.get_first_day_next_month(date.today())
            first_date_in_next_month = datetime.combine(first_date_in_next_month, datetime.min.time())
        elif date_type == "quarter":
            first_date_in_this_month = date_util.get_first_day_current_quarter(date.today())
            first_date_in_this_month = datetime.combine(first_date_in_this_month, datetime.min.time())
            first_date_in_last_month = date_util.get_first_day_last_quarter(date.today())
            first_date_in_last_month = datetime.combine(first_date_in_last_month, datetime.min.time())
            first_date_in_next_month = date_util.get_first_day_next_quarter(date.today())
            first_date_in_next_month = datetime.combine(first_date_in_next_month, datetime.min.time())
        elif date_type == "year":
            first_date_in_this_month = date_util.get_first_day_current_year(date.today())
            first_date_in_this_month = datetime.combine(first_date_in_this_month, datetime.min.time())
            first_date_in_last_month = date_util.get_first_day_last_year(date.today())
            first_date_in_last_month = datetime.combine(first_date_in_last_month, datetime.min.time())
            first_date_in_next_month = date_util.get_first_day_next_year(date.today())
            first_date_in_next_month = datetime.combine(first_date_in_next_month, datetime.min.time())
        else:
            first_date_in_this_month = date_util.get_first_day_current_month(date.today())
            first_date_in_this_month = datetime.combine(first_date_in_this_month, datetime.min.time())
            first_date_in_last_month = date_util.get_first_day_last_month(date.today())
            first_date_in_last_month = datetime.combine(first_date_in_last_month, datetime.min.time())
            first_date_in_next_month = date_util.get_first_day_next_month(date.today())
            first_date_in_next_month = datetime.combine(first_date_in_next_month, datetime.min.time())

        t1 = 0
        t2 = 0
        n_this_month = 0
        n_last_month = 0
        # alarm_level
        if True:
            alarm_level_this_month = dict()
            alarm_level_last_month = dict()
            for item in const.STATION_STATUS:
                alarm_level_this_month[item] = dict()
                alarm_level_this_month[item]['value'] = const.STATION_STATUS[item]['value']
                alarm_level_this_month[item]['name'] = str(T(const.STATION_STATUS[item]['name']))
                alarm_level_this_month[item]['color'] = const.STATION_STATUS[item]['color']
                alarm_level_this_month[item]['icon'] = const.STATION_STATUS[item]['icon']
                alarm_level_this_month[item]['qty'] = 0
                alarm_level_this_month[item]['percent'] = 0

                alarm_level_last_month[item] = dict()
                alarm_level_last_month[item]['value'] = const.STATION_STATUS[item]['value']
                alarm_level_last_month[item]['name'] = str(T(const.STATION_STATUS[item]['name']))
                alarm_level_last_month[item]['color'] = const.STATION_STATUS[item]['color']
                alarm_level_last_month[item]['icon'] = const.STATION_STATUS[item]['icon']
                alarm_level_last_month[item]['qty'] = 0
                alarm_level_last_month[item]['percent'] = 0
        # station_type
        if True:
            station_type_this_month = dict()
            station_type_last_month = dict()
            # for item in const.STATION_TYPE:
            for item in common.get_station_types():
                _key = str(item['value'])
                # This month
                station_type_this_month[_key] = dict()
                station_type_this_month[_key]['value'] = item['value']  # const.STATION_TYPE[item]['value']
                station_type_this_month[_key]['name'] = T(item['name'])  # str(T(const.STATION_TYPE[item]['name']))
                station_type_this_month[_key]['qty'] = 0
                # Last month
                station_type_last_month[_key] = dict()
                station_type_last_month[_key]['value'] = item['value']  # const.STATION_TYPE[item]['value']
                station_type_last_month[_key]['name'] = T(item['name'])  # str(T(const.STATION_TYPE[item]['name']))
                station_type_last_month[_key]['qty'] = 0

        alarm_level_this_month.pop('PREPARING')
        alarm_level_last_month.pop('PREPARING')
        alarm_level_this_month.pop('TENDENCY')
        alarm_level_last_month.pop('TENDENCY')

        stations = db(db.stations.id.belongs(station_ids)).select()
        ## tinh lượng datamin phải nhận trong tháng này
        conditions = (db.data_min_month_collect.station_id.belongs(station_ids))
        conditions &= (db.data_min_month_collect.date_month >= first_date_in_this_month)
        conditions &= (db.data_min_month_collect.date_month < first_date_in_next_month)
        data_min_month_this_month = db(conditions).select(db.data_min_month_collect.station_id,
                                                          db.data_min_month_collect.actual_datamin)
        data_min_month_this_month_dict = dict()
        for item in data_min_month_this_month:
            data_min_month_this_month_dict[item.station_id] = item
        expected_this_month = 0.0
        days_this_month = datetime.now().day - 1
        days_this_month += 1.0 / 24.0 * datetime.now().hour
        for row in stations:
            # tần suất nhận dữ liệu
            freq = row['frequency_receiving_data']
            # freq = 5
            indicator_number = db(db.station_indicator.station_id == row.id).count(db.station_indicator.indicator_id)
            if not indicator_number:
                indicator_number = 0
            if (not freq) or (freq == 0):
                freq = 5
            expected_this_month_each = (indicator_number * days_this_month * 24 * 60 / freq)
            if data_min_month_this_month_dict.has_key(str(row.id)):
                actual_this_month = data_min_month_this_month_dict[str(row.id)].actual_datamin

                if actual_this_month:
                    if expected_this_month_each < actual_this_month:
                        expected_this_month_each = actual_this_month
            expected_this_month += expected_this_month_each
        # Count number records of data_min in this month
        if data_min_month_this_month:
            data = db(conditions).select(
                db.data_min_month_collect.actual_datamin.sum().with_alias('actual_datamin'),
                db.data_min_month_collect.qty_good.sum().with_alias('qty_good'),
                db.data_min_month_collect.qty_exceed.sum().with_alias('qty_exceed'),
                db.data_min_month_collect.qty_adjusting.sum().with_alias('qty_adjusting'),
                db.data_min_month_collect.qty_error.sum().with_alias('qty_error'),
            ).first()
            n_this_month = data['actual_datamin'] if data['actual_datamin'] else 0
            t1 = n_this_month
            n_qty_this_month = data['actual_datamin'] if data['actual_datamin'] else 0
            # Tinh theo %
            if expected_this_month and expected_this_month > 0:
                n_this_month = round(100 * data['actual_datamin'] / expected_this_month, 2) if data[
                    'actual_datamin'] else 0
            else:
                n_this_month = 0
            # n_this_month = "{:,}".format(n_this_month)
            alarm_level_this_month['GOOD']['qty'] = common.convert_data(data['qty_good'])
            alarm_level_this_month['GOOD']['percent'] = common.convert_data(
                100.0 * data['qty_good'] / expected_this_month if expected_this_month > 0 and data['qty_good'] else 0)

            alarm_level_this_month['EXCEED']['qty'] = common.convert_data(data['qty_exceed'])
            alarm_level_this_month['EXCEED']['percent'] = common.convert_data(
                100.0 * data['qty_exceed'] / expected_this_month if expected_this_month > 0 and data[
                    'qty_exceed'] else 0)

            alarm_level_this_month['ADJUSTING']['qty'] = common.convert_data(data['qty_adjusting'])
            alarm_level_this_month['ADJUSTING']['percent'] = common.convert_data(
                100.0 * data['qty_adjusting'] / expected_this_month if expected_this_month > 0 and data[
                    'qty_adjusting'] else 0)
            alarm_level_this_month['ERROR']['qty'] = common.convert_data(data['qty_error'])
            alarm_level_this_month['ERROR']['percent'] = common.convert_data(
                100.0 * data['qty_error'] / expected_this_month if expected_this_month > 0 and data['qty_error'] else 0)

            offline_count = expected_this_month - data['qty_good'] - data['qty_exceed'] - data['qty_adjusting'] - data[
                'qty_error']
            alarm_level_this_month['OFFLINE']['qty'] = common.convert_data(offline_count)
            alarm_level_this_month['OFFLINE']['percent'] = common.convert_data(
                100.0 * offline_count / expected_this_month if expected_this_month > 0 and offline_count else 0)

        conditions = (db.data_min_month_collect.station_id.belongs(station_ids))
        conditions &= (db.data_min_month_collect.date_month >= first_date_in_last_month)
        conditions &= (db.data_min_month_collect.date_month < first_date_in_this_month)
        data_min_month_last_month = db(conditions).select(db.data_min_month_collect.station_id,
                                                          db.data_min_month_collect.actual_datamin)
        data_min_month_last_month_dict = dict()
        for item in data_min_month_last_month:
            data_min_month_last_month_dict[item.station_id] = item
        expected_last_month = 0.0
        # số ngày trong tháng
        days_last_month = (
                first_date_in_last_month.replace(month=first_date_in_last_month.month % 12 + 1, day=1) - timedelta(
            days=1)).day
        for row in stations:
            # tần suất nhận dữ liệu
            freq = row['frequency_receiving_data']
            # freq = 5
            indicator_number = db(db.station_indicator.station_id == row.id).count(db.station_indicator.indicator_id)
            if not indicator_number:
                indicator_number = 0
            if (not freq) or (freq == 0):
                freq = 5
            expected_last_month_each = (indicator_number * days_last_month * 24 * 60 / freq)
            # Trường hợp data nhận được nhiều hơn dự kiến (expected)
            if data_min_month_last_month_dict.has_key(str(row.id)):
                actual_last_month = data_min_month_last_month_dict[str(row.id)].actual_datamin
                if actual_last_month:
                    if expected_last_month_each < actual_last_month:
                        expected_last_month_each = actual_last_month
                expected_last_month += expected_last_month_each

        # Count number records of data_min in last month
        if data_min_month_last_month:
            # conditions = (db.data_min_month_collect.station_id.belongs(station_ids))
            # conditions &= (db.data_min_month_collect.date_month >= first_date_in_last_month)
            # conditions &= (db.data_min_month_collect.date_month < first_date_in_this_month)
            data = db(conditions).select(
                db.data_min_month_collect.actual_datamin.sum().with_alias('actual_datamin'),
                db.data_min_month_collect.qty_good.sum().with_alias('qty_good'),
                db.data_min_month_collect.qty_exceed.sum().with_alias('qty_exceed'),
                db.data_min_month_collect.qty_adjusting.sum().with_alias('qty_adjusting'),
                db.data_min_month_collect.qty_error.sum().with_alias('qty_error'),
            ).first()
            n_last_month = data['actual_datamin'] if data['actual_datamin'] else 0
            t2 = n_last_month
            n_qty_last_month = data['actual_datamin'] if data['actual_datamin'] else 0
            if expected_last_month and expected_last_month > 0:
                n_last_month = round(100 * data['actual_datamin'] / expected_last_month, 2) if data[
                    'actual_datamin'] else 0
            else:
                n_last_month = 0
            alarm_level_last_month['GOOD']['qty'] = common.convert_data(data['qty_good'])
            alarm_level_last_month['GOOD']['percent'] = common.convert_data(
                100.0 * data['qty_good'] / expected_last_month if expected_last_month > 0 and data['qty_good'] else 0)

            alarm_level_last_month['EXCEED']['qty'] = common.convert_data(data['qty_exceed'])
            alarm_level_last_month['EXCEED']['percent'] = common.convert_data(
                100.0 * data['qty_exceed'] / expected_last_month if expected_last_month > 0 and data[
                    'qty_exceed'] else 0)

            alarm_level_last_month['ADJUSTING']['qty'] = common.convert_data(data['qty_adjusting'])
            alarm_level_last_month['ADJUSTING']['percent'] = common.convert_data(
                100.0 * data['qty_adjusting'] / expected_last_month if expected_last_month > 0 and data[
                    'qty_adjusting'] else 0)
            alarm_level_last_month['ERROR']['qty'] = common.convert_data(data['qty_error'])
            alarm_level_last_month['ERROR']['percent'] = common.convert_data(
                100.0 * data['qty_error'] / expected_last_month if expected_last_month > 0 and data['qty_error'] else 0)

            offline_count = expected_last_month - data['qty_good'] - data['qty_exceed'] - data['qty_adjusting'] - data[
                'qty_error']
            alarm_level_last_month['OFFLINE']['qty'] = common.convert_data(offline_count)
            alarm_level_last_month['OFFLINE']['percent'] = common.convert_data(
                100.0 * offline_count / expected_last_month if expected_last_month > 0 and offline_count else 0)
        # So sanh gtri collect tong cong
        collect_icon = 'fa fa-arrow-up text-info'
        if t1 < t2:
            collect_icon = 'fa fa-arrow-down text-info'
        elif t1 == t2:
            collect_icon = 'fa fa-pause text-warning'

        # So sanh gtri 3 nguong thang nay va thang truoc de display Icon len/xuong cho dung
        for item in alarm_level_last_month:
            if alarm_level_this_month[item]['percent'] > alarm_level_last_month[item]['percent']:
                alarm_level_this_month[item]['icon'] = 'fa fa-arrow-up text-danger'
            if alarm_level_this_month[item]['percent'] < alarm_level_last_month[item]['percent']:
                alarm_level_this_month[item]['icon'] = 'fa fa-arrow-down text-info'
            else:
                alarm_level_this_month[item]['icon'] = 'fa fa-pause text-warning'

        # Count number records of station_off_log in this month
        if True:
            conditions = (db.data_min_month_collect.station_id.belongs(station_ids))
            conditions &= (db.data_min_month_collect.date_month >= first_date_in_this_month)
            conditions &= (db.data_min_month_collect.date_month < first_date_in_next_month)
            conditions &= (db.data_min_month_collect.number_off_log > 0)
            rows = db(conditions).select(
                db.data_min_month_collect.id,
                db.data_min_month_collect.station_type,
                db.data_min_month_collect.number_off_log
            )
            n2_this_month = 0
            for row in rows:
                n2_this_month += row.number_off_log
                for item in station_type_this_month:
                    if row.station_type == station_type_this_month[item]['value']:
                        station_type_this_month[item]['qty'] += row.number_off_log
                        break
            for item in station_type_this_month:
                if station_type_this_month[item]['qty']:
                    station_type_this_month[item]['qty'] = common.convert_data(station_type_this_month[item]['qty'])
            t1 = n2_this_month
            n2_this_month = common.convert_data(n2_this_month)

        # Count number records of station_off_log in last month
        if True:
            conditions = (db.data_min_month_collect.station_id.belongs(station_ids))
            conditions &= (db.data_min_month_collect.date_month >= first_date_in_last_month)
            conditions &= (db.data_min_month_collect.date_month < first_date_in_this_month)
            conditions &= (db.data_min_month_collect.number_off_log > 0)
            rows = db(conditions).select(
                db.data_min_month_collect.id,
                db.data_min_month_collect.station_type,
                db.data_min_month_collect.number_off_log
            )
            n2_last_month = 0
            for row in rows:
                n2_last_month += row.number_off_log
                for item in station_type_last_month:
                    if row.station_type == station_type_last_month[item]['value']:
                        station_type_last_month[item]['qty'] += row.number_off_log
                        break
            for item in station_type_last_month:
                if station_type_last_month[item]['qty']:
                    station_type_last_month[item]['qty'] = common.convert_data(station_type_last_month[item]['qty'])
            t2 = n2_last_month
            n2_last_month = common.convert_data(n2_last_month)

        # So sanh gtri offline cua station
        offline_icon = 'fa fa-arrow-up text-danger'
        if t1 < t2:
            collect_icon = 'fa fa-arrow-down text-info'
        elif t1 == t2:
            collect_icon = 'fa fa-pause text-warning'
        aaData = []
        aaData.append(['<strong>I</strong>', '<strong>Dữ liệu thu thập</strong>',
                       '<strong>%s%s</strong><br/><small>%s dữ liệu đã thu thập trên %s dữ liệu cần thu thập</small>'
                       % (common.convert_data(n_this_month), '%', common.convert_data(n_qty_this_month),
                          common.convert_data(expected_this_month)),
                       '<strong>%s%s</strong><br/><small>%s dữ liệu đã thu thập trên %s dữ liệu cần thu thập</small>'
                       % (common.convert_data(n_last_month), '%', common.convert_data(n_qty_last_month),
                          common.convert_data(expected_last_month))])
        aaData.append(['1', 'Hoạt động tốt',
                       str(alarm_level_this_month['GOOD']['percent']) + '% ' + '<small>(%s dữ liệu)</small>' % (
                           alarm_level_this_month['GOOD']['qty']),
                       str(alarm_level_last_month['GOOD']['percent']) + '% ' + '<small>(%s dữ liệu)</small>' % (
                           alarm_level_last_month['GOOD']['qty'])])
        aaData.append(['2', 'Vượt quy chuẩn',
                       str(alarm_level_this_month['EXCEED']['percent']) + '% ' + '<small>(%s dữ liệu)</small>' % (
                           alarm_level_this_month['EXCEED']['qty']),
                       str(alarm_level_last_month['EXCEED']['percent']) + '% ' + '<small>(%s dữ liệu)</small>' % (
                           alarm_level_last_month['EXCEED']['qty'])])
        aaData.append(['3', 'Hiệu chuẩn',
                       str(alarm_level_this_month['ADJUSTING']['percent']) + '% ' + '<small>(%s dữ liệu)</small>' % (
                           alarm_level_this_month['ADJUSTING']['qty']),
                       str(alarm_level_last_month['ADJUSTING']['percent']) + '% ' + '<small>(%s dữ liệu)</small>' % (
                           alarm_level_last_month['ADJUSTING']['qty'])])
        aaData.append(['4', 'Lỗi thiết bị',
                       str(alarm_level_this_month['ERROR']['percent']) + '% ' + '<small>(%s dữ liệu)</small>' % (
                           alarm_level_this_month['ERROR']['qty']),
                       str(alarm_level_last_month['ERROR']['percent']) + '% ' + '<small>(%s dữ liệu)</small>' % (
                           alarm_level_last_month['ERROR']['qty'])])
        aaData.append(['5', 'Mất kết nối',
                       str(alarm_level_this_month['OFFLINE']['percent']) + '% ' + '<small>(%s dữ liệu)</small>' % (
                           alarm_level_this_month['OFFLINE']['qty']),
                       str(alarm_level_last_month['OFFLINE']['percent']) + '% ' + '<small>(%s dữ liệu)</small>' % (
                           alarm_level_last_month['OFFLINE']['qty'])])
        aaData.append(['<strong>II</strong>', '<strong>Sự kiện mất dữ liệu<strong>',
                       '<strong>%s</strong><br/><small> Các sự kiện mất dữ liệu<small/>' % (n2_this_month),
                       '<strong>%s</strong><br/><small> Các sự kiện mất dữ liệu<small/' % (n2_last_month)])
        aaData.append(['1', 'Nước mặt', station_type_this_month['1']['qty'], station_type_last_month['1']['qty']])
        aaData.append(['2', 'Nước thải', station_type_this_month['0']['qty'], station_type_last_month['0']['qty']])
        aaData.append(['3', 'Khí thải', station_type_this_month['3']['qty'], station_type_last_month['3']['qty']])
        # Lang update loi du lieu thi sai
        aaData.append(['4', 'Không khí', station_type_this_month['4']['qty'], station_type_last_month['4']['qty']])
        aaData.append(['5', 'Nước ngầm', station_type_this_month['2']['qty'], station_type_last_month['2']['qty']])

        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_ratio_hour_16():
    now = datetime.now()
    station_id = request.vars.station_id
    provinces = common.get_province_have_station()
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, station_id=station_id, default_provinces=default_provinces,
                year=now.year, month=now.month)


################################################################################
@service.json
def get_list_report_ratio_hour_16(*args, **kwargs):
    try:
        station_id = request.vars.station_id
        data_type = request.vars.data_type
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        added_columns = request.vars.indicator or ''
        if added_columns:
            added_columns = added_columns.split(',')
        table = 'data_hour'
        if data_type:
            if int(data_type) == 2:
                table = 'data_hour_adjust'
        conditions = {'station_id': station_id}
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        month = request.vars.Month
        year = request.vars.Year
        if month and year:
            year = int(year)
            if "quarter" in month:
                quarter = month.replace("quarter", "")
                quarter = int(quarter)
                if quarter == 1:
                    from_date = datetime(year=year, month=1, day=1)
                    to_date = datetime(year=year, month=4, day=1)
                elif quarter == 2:
                    from_date = datetime(year=year, month=4, day=1)
                    to_date = datetime(year=year, month=7, day=1)
                elif quarter == 3:
                    from_date = datetime(year=year, month=7, day=1)
                    to_date = datetime(year=year, month=10, day=1)
                elif quarter == 4:
                    from_date = datetime(year=year, month=10, day=1)
                    to_date = datetime(year=year + 1, month=1, day=1)
            else:
                month = int(month)
                from_date = datetime(year=year, month=month, day=1)
                to_date = date_util.get_first_day_next_month(from_date)
                to_date = datetime.combine(to_date, datetime.min.time())
        else:
            if from_date:
                from_date = date_util.string_to_datetime(from_date)
            if to_date:
                to_date = date_util.string_to_datetime(to_date) + timedelta(days=1)
        if from_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            conditions['get_time']['$gte'] = from_date
        if to_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            conditions['get_time']['$lte'] = to_date
        duration = (to_date - from_date).days
        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = len(added_columns)  # Tong so ban ghi
        rows = db(db.station_indicator.station_id == station_id).select(db.station_indicator.ALL,
                                                                        orderby=db.station_indicator.station_id)

        dic_station_indicator = dict()
        indicators = common.get_indicator_dict()

        for row in rows:
            if not dic_station_indicator.has_key(row['station_id']):
                dic_station_indicator[str(row['station_id'])] = dict()
            if indicators.has_key(row['indicator_id']):
                indicator_name = indicators[row['indicator_id']]
                dic_station_indicator[str(row['station_id'])][indicator_name] = row.as_dict()

        qcvn_dict = dic_station_indicator[station_id]

        dic_data_c = {}
        attrs = {'get_time': 1, 'data': 1}
        list_data = pydb[table].find(conditions, attrs).sort('get_time', 1)
        for item in list_data:
            timeconvert = item['get_time'].strftime("%m/%d/%Y")
            for i in added_columns:
                if item['data'].has_key(i):
                    x = str(item['data'][i])
                    if x == 'None' or x == 'NULL' or x == '-' or x.find(',') != -1:
                        continue
                    else:
                        z = float(x)
                        name_decode = i.encode('utf-8')
                        qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        check_qcvn = True
                        if z or z == 0:
                            if qcvn_max or qcvn_min:
                                check_qcvn = True
                            if qcvn_max and z > qcvn_max:
                                check_qcvn = False
                            if qcvn_min and z < qcvn_min:
                                check_qcvn = False
                            if not check_qcvn:
                                if not dic_data_c.has_key(timeconvert):
                                    dic_data_c[timeconvert] = {i: {'total': 1, 'status': True}}
                                else:
                                    if dic_data_c[timeconvert].has_key(i):
                                        dic_data_c[timeconvert][i]['total'] += 1
                                    else:
                                        dic_data_c[timeconvert][i] = {'total': 1, 'status': True}
        for c, it in enumerate(added_columns):
            row = [
                it,
                0,
                0,
                0
            ]
            for i in dic_data_c:
                if dic_data_c[i].has_key(it):
                    row[2] += dic_data_c[i][it]['total']
                    row[1] += 1
            row[3] = round(float(row[2]) / float(duration * 24), 2)
            aaData.append(row)
        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


################################################################################
@service.json
def export_excel_0():
    import os.path, openpyxl, pyexcelerate
    from datetime import datetime, timedelta
    from w2pex import date_util
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    data_type = request.vars.data_type

    table = 'data_min'
    if data_type:
        if int(data_type) == 2:
            table = 'data_adjust'
    conditions = {'station_id': station_id}
    # conditions['is_exceed'] = True
    aaData = []  # Du lieu json se tra ve
    iTotalRecords = 0  # Tong so ban ghi
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')

    time_range = request.vars.time_range
    if time_range:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=int(time_range))

    if from_date:
        if not conditions.has_key('get_time'):
            conditions['get_time'] = {}
        conditions['get_time']['$gte'] = date_util.string_to_datetime(from_date)
    if to_date:
        if not conditions.has_key('get_time'):
            conditions['get_time'] = {}
        conditions['get_time']['$lte'] = date_util.string_to_datetime(to_date) + timedelta(days=1)
    if not from_date and not to_date:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=365)
        conditions['get_time'] = {'$gte': from_date, '$lte': to_date}
    attrs = {'get_time': 1, 'data': 1}
    rows = pydb[table].find(conditions, attrs).sort('get_time', -1)
    dict_data = {}
    for row in rows:
        key = row['get_time'].strftime('%Y-%m-%d')
        if not dict_data.get(key): dict_data[key] = {'data': []}
        dict_data[key]['data'].append(row['data'])

    iTotalRecords = len(dict_data)
    indicators = common.get_indicator_dict()
    conditions = (db.station_indicator.station_id == station_id)
    conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
    rows = db(conditions).select(db.station_indicator.ALL)
    qcvn_dict = dict()
    for row in rows:
        indicator_name = indicators[row.indicator_id]
        qcvn_dict[indicator_name] = row.as_dict()

    temp_headers = []
    headers = []
    temp_headers.append('Ngày giờ')
    for item in added_columns:
        temp_headers.append(str(item))
    headers.append(temp_headers)
    for header in headers:
        aaData.append(header)

    for key in sorted(dict_data.keys(), reverse=True):
        count = dict()
        a = dict_data[key]['data']
        # Lang update tinh dung dan cua du lieu
        query = {'station_id': station_id}
        if not query.has_key('get_time'):
            query['get_time'] = {}
        from_date = date_util.string_to_datetime(key)
        to_date = date_util.string_to_datetime(key) + timedelta(days=1)
        query['get_time'] = {'$gte': from_date, '$lte': to_date}
        countTotal = pydb[table].count(query)
        # countTotal = len(a)
        for i in a:
            for indicator_name in i:
                v = str(i[indicator_name])
                x = v.replace(",", "")
                if x == 'NULL' or x == 'None' or x == '-':
                    continue
                if x != '-':
                    z = float(x)
                    name_decode = indicator_name.encode('utf-8')
                    qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                        name_decode) else None
                    qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                        name_decode) else None
                    check_qcvn = True
                    if z:
                        if qcvn_max or qcvn_min:
                            check_qcvn = True
                        if qcvn_max and z > qcvn_max:
                            check_qcvn = False
                        if qcvn_min and z < qcvn_min:
                            check_qcvn = False
                        if not check_qcvn:
                            if not count.has_key(name_decode):
                                count[name_decode] = 1
                            else:
                                count[name_decode] = int(count[name_decode]) + 1

        date_time_obj = datetime.strptime(key, '%Y-%m-%d')
        row = [
            date_time_obj.strftime('%d/%m/%Y'),
        ]
        for column in added_columns:
            if column and count.has_key(column):
                m = int(count[column])
                n = (100 * float(m)) / float(countTotal)
                n = round(n, 2)
                # n = "{0:.2f}".format(n)
                content = '%s (%s)' % (n, '%')
                row.append(content)
            else:
                content = '-'
                row.append(content)
        aaData.append(row)

    wb = pyexcelerate.Workbook()
    wb.new_sheet('Data', aaData)

    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace(",", " ")

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


###############################################################################
@service.json
def export_excel_1():
    from datetime import datetime, timedelta
    import os.path, openpyxl, pyexcelerate
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    data_type = request.vars.data_type
    exeed_type_data = request.vars.exeed_type_data
    if data_type:
        data_type = int(data_type)

    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    aaData = []
    table = exeed_type_data
    if data_type == 2:
        if exeed_type_data == 'data_min':
            table = 'data_adjust'
        else:
            table = exeed_type_data + "_adjust"

    conditions = {'station_id': station_id}
    time_range = request.vars.time_range
    if time_range:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=int(time_range))

    if from_date:
        if not conditions.has_key('get_time'):
            conditions['get_time'] = {}
        conditions['get_time']['$gte'] = date_util.string_to_datetime(from_date)
    if to_date:
        if not conditions.has_key('get_time'):
            conditions['get_time'] = {}
        # conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
        conditions['get_time']['$lte'] = date_util.string_to_datetime(to_date) + timedelta(days=1)
    if not from_date and not to_date:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=365)
        conditions['get_time'] = {'$gte': from_date, '$lte': to_date}

    attrs = {'get_time': 1, 'data': 1}

    # Tong so ban ghi khong thuc hien phan trang
    rows = db(db.station_indicator.station_id == station_id).select(db.station_indicator.ALL,
                                                                    orderby=db.station_indicator.station_id)

    dic_station_indicator = dict()
    indicators = common.get_indicator_dict()

    for row in rows:
        if not dic_station_indicator.has_key(row['station_id']):
            dic_station_indicator[str(row['station_id'])] = dict()
        if indicators.has_key(row['indicator_id']):
            indicator_name = indicators[row['indicator_id']]
            dic_station_indicator[str(row['station_id'])][indicator_name] = row.as_dict()

    qcvn_dict = dic_station_indicator[station_id]
    query = common.where_is_exceed(added_columns, qcvn_dict)
    if len(query) > 0:
        conditions["$or"] = query
    list_data = pydb[table].find(conditions, attrs).sort('get_time', -1)

    if list_data is None:
        raise HTTP(400, "Không đủ dữ liệu xuất excel, vui lòng thử lại!")
    else:
        temp_headers = []
        headers = []
        temp_headers.append('STT')
        temp_headers.append('Datetime')
        for i, indicator in enumerate(added_columns):
            temp_headers.append(indicator)
        headers.append(temp_headers)
        for header in headers:
            aaData.append(header)

    for i, item in enumerate(list_data):
        row_data = [
            str(1 + i),
            item['get_time'].strftime("%Y-%m-%d %H:%M:%S")
        ]
        added_item = dict()
        if item['data']:
            if added_columns:
                for data_key in added_columns:
                    i_name = str(data_key).encode('utf-8')
                    current_indicator = item['data'][i_name] if item['data'].has_key(i_name) else ''
                    if not current_indicator:
                        added_item[i_name] = '-'
                    else:
                        try:
                            z = float(current_indicator)
                            # qcvn_detail_min_value = current_indicator['station_qcvn_min']
                            # qcvn_detail_max_value = current_indicator['station_qcvn_max']
                            qcvn_min = qcvn_dict[data_key]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                                data_key) else None
                            qcvn_max = qcvn_dict[data_key]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                                data_key) else None
                            check_qcvn = True
                            if z or z == 0:
                                if qcvn_min or qcvn_max:
                                    check_qcvn = True
                                if qcvn_max and z > qcvn_max:
                                    check_qcvn = False
                                    title = 'QCVN Max: %s' % qcvn_max
                                if qcvn_min and z < qcvn_min:
                                    check_qcvn = False
                                    title = 'QCVN Min: %s' % qcvn_min
                                if not check_qcvn:
                                    added_item[data_key] = common.convert_data(z)
                                else:
                                    added_item[data_key] = '-'
                        except:
                            row_data.append('-')
        for column in added_columns:
            if column and added_item.has_key(column):
                row_data.append(added_item[column])
        aaData.append(row_data)

    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace(",", " ")
    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb = pyexcelerate.Workbook()
    wb.new_sheet('Data', aaData)
    wb.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


##############################################################################
@service.json
def export_excel_2():
    import os.path, openpyxl, pyexcelerate
    # get search parameters
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    province_id = request.vars.province_id
    data_type = request.vars.data_type
    station_type = request.vars.station_type
    exeed_type_data = request.vars.exeed_type_data
    area_ids = request.vars.area_ids

    conditions = (db.stations.id > 0)
    added_stations = request.vars.added_stations or ''
    if added_stations:
        added_stations = added_stations.split(',')
        conditions &= db.stations.id.belongs(added_stations)
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if station_type:
        conditions &= (db.stations.station_type == station_type)
    if province_id:
        conditions &= (db.stations.province_id == province_id)
    if area_ids:
        area_ids = area_ids.split(',')
        conditions &= (db.stations.area_ids.belongs(area_ids))

    time_range = request.vars.time_range
    if time_range:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=int(time_range))

    table = exeed_type_data
    if data_type:
        if int(data_type) == 2:
            if exeed_type_data == 'data_min':
                table = 'data_adjust'
            else:
                table = exeed_type_data + '_adjust'

    aaData = []  # Du lieu json se tra ve
    arrSorted = []
    list_data = None  # Du lieu truy van duoc
    iTotalRecords = 0  # Tong so ban ghi

    stations = db(conditions).select(db.stations.id,
                                     db.stations.station_name,
                                     db.stations.station_type,
                                     db.stations.province_id)

    query = db.stations.station_type == station_type
    if province_id:
        query &= db.stations.province_id == province_id
    if added_stations:
        query &= db.stations.id.belongs(added_stations)
    indicators = common.get_indicator_dict()
    if stations is None:
        raise HTTP(400, "Không đủ dữ liệu xuất excel, vui lòng thử lại!")
    else:
        temp_headers = []
        headers = []
        temp_headers.append('Tên trạm')
        for i, indicator in enumerate(added_columns):
            temp_headers.append(indicator.upper())
        headers.append(temp_headers)
        for header in headers:
            aaData.append(header)
    iRow = 1
    for c, station in enumerate(stations):
        count_total_dict = dict()
        station_id = str(station.id)
        conditions = {'station_id': station_id}
        # conditions['is_exceed'] = True
        if from_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            conditions['get_time']['$gte'] = date_util.string_to_datetime(from_date)
        if to_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            conditions['get_time']['$lte'] = date_util.string_to_datetime(to_date) + timedelta(days=1)
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions['get_time'] = {'$gte': from_date, '$lte': to_date}

        attrs = {'get_time': 1, 'data': 1}
        list_data = pydb[table].find(conditions, attrs).sort('get_time', -1)

        # Tong so ban ghi khong thuc hien phan trang
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db(conditions).select(db.station_indicator.ALL)
        qcvn_dict = dict()
        for row in rows:
            indicator_name = indicators[row.indicator_id]
            qcvn_dict[indicator_name] = row.as_dict()
        count = dict()
        # for in
        for item in list_data:
            if item['data']:
                for indicator_name in item['data']:
                    x = str(item['data'][indicator_name])
                    if x == 'None' or x == 'NULL' or x == '-' or x.find(',') != -1:
                        continue
                    else:
                        z = float(x)
                        name_decode = indicator_name.encode('utf-8')
                        qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        check_qcvn = True
                        if z or z == 0:
                            if qcvn_max or qcvn_min:
                                check_qcvn = True
                            if qcvn_max and z > qcvn_max:
                                check_qcvn = False
                            if qcvn_min and z < qcvn_min:
                                check_qcvn = False
                            if not check_qcvn:
                                if not count.has_key(name_decode):
                                    count[name_decode] = 1
                                else:
                                    count[name_decode] = int(count[name_decode]) + 1
        row = [
            station.station_name,
        ]
        count_total_dict['total'] = 0
        for column in added_columns:
            if column and count.has_key(column):
                count_total_dict['total'] += count[column]
                row.append(count[column])
            else:
                row.append(0)

        count_total_dict['data'] = row
        arrSorted.append(count_total_dict)

    sorted(arrSorted, key=lambda x: x['total'])
    i = 0
    for item in arrSorted:
        if item:
            aaData.append(item['data'])
        i += 1

    # Get station name

    file_name = request.now.strftime('_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb = pyexcelerate.Workbook()
    wb.new_sheet('Data', aaData)
    wb.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


@service.json
def export_excel_2_common():
    import os.path, openpyxl, pyexcelerate
    # get search parameters
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    province_id = request.vars.province_id
    data_type = request.vars.data_type
    station_type = request.vars.station_type
    exeed_type_data = request.vars.exeed_type_data
    area_ids = request.vars.area_ids

    conditions = (db.stations.id > 0)
    added_stations = request.vars.added_stations or ''
    if added_stations:
        conditions &= (db.stations.id == added_stations)
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')

    table = exeed_type_data
    if data_type:
        if int(data_type) == 2:
            if exeed_type_data == 'data_min':
                table = 'data_adjust'
            else:
                table = exeed_type_data + '_adjust'

    aaData = []  # Du lieu json se tra ve
    arrSorted = []
    list_data = None  # Du lieu truy van duoc
    iTotalRecords = 0  # Tong so ban ghi

    stations = db(conditions).select(db.stations.id,
                                     db.stations.station_name,
                                     db.stations.station_type,
                                     db.stations.province_id)

    query = db.stations.station_type == station_type
    if province_id:
        query &= db.stations.province_id == province_id
    if added_stations:
        query &= db.stations.id.belongs(added_stations)
    indicators = common.get_indicator_dict()
    if stations is None:
        raise HTTP(400, "Không đủ dữ liệu xuất excel, vui lòng thử lại!")
    else:
        temp_headers = []
        headers = []
        temp_headers.append('Tên trạm')
        for i, indicator in enumerate(added_columns):
            temp_headers.append(indicator.upper())
        headers.append(temp_headers)
        for header in headers:
            aaData.append(header)
    for c, station in enumerate(stations):
        count_total_dict = dict()
        station_id = str(station.id)
        conditions = {'station_id': station_id}
        # conditions['is_exceed'] = True
        if from_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            conditions['get_time']['$gte'] = date_util.string_to_datetime(from_date)
        if to_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            conditions['get_time']['$lte'] = date_util.string_to_datetime(to_date) + timedelta(days=1)
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions['get_time'] = {'$gte': from_date, '$lte': to_date}

        attrs = {'get_time': 1, 'data': 1}
        list_data = pydb[table].find(conditions, attrs).sort('get_time', -1)

        # Tong so ban ghi khong thuc hien phan trang
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db(conditions).select(db.station_indicator.ALL)
        qcvn_dict = dict()
        for row in rows:
            indicator_name = indicators[row.indicator_id]
            qcvn_dict[indicator_name] = row.as_dict()
        count = dict()
        # for in
        for item in list_data:
            if item['data']:
                for indicator_name in item['data']:
                    x = str(item['data'][indicator_name])
                    if x == 'None' or x == 'NULL' or x == '-' or x.find(',') != -1:
                        continue
                    else:
                        z = float(x)
                        name_decode = indicator_name.encode('utf-8')
                        qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        check_qcvn = True
                        if z or z == 0:
                            if qcvn_max or qcvn_min:
                                check_qcvn = True
                            if qcvn_max and z > qcvn_max:
                                check_qcvn = False
                            if qcvn_min and z < qcvn_min:
                                check_qcvn = False
                            if not check_qcvn:
                                if not count.has_key(name_decode):
                                    count[name_decode] = 1
                                else:
                                    count[name_decode] = int(count[name_decode]) + 1
        row = [
            station.station_name,
        ]
        count_total_dict['total'] = 0
        for column in added_columns:
            if column and count.has_key(column):
                count_total_dict['total'] += count[column]
                row.append(count[column])
            else:
                row.append(0)
        count_total_dict['data'] = row
        arrSorted.append(count_total_dict)

        sorted(arrSorted, key=lambda x: x['total'])
        for item in arrSorted:
            if item:
                aaData.append(item['data'])

    # Get station name

    file_name = request.now.strftime('_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb = pyexcelerate.Workbook()
    wb.new_sheet('Data', aaData)
    wb.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


@service.json
def export_excel_2_a_station():
    import os.path, openpyxl, pyexcelerate
    # get search parameters
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    province_id = request.vars.province_id
    data_type = request.vars.data_type
    station_type = request.vars.station_type
    exeed_type_data = request.vars.exeed_type_data
    area_ids = request.vars.area_ids

    conditions = (db.stations.id > 0)
    added_stations = request.vars.added_stations or ''
    if added_stations:
        conditions &= (db.stations.id == added_stations)
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if station_type:
        conditions &= (db.stations.station_type == station_type)
    if province_id:
        conditions &= (db.stations.province_id == province_id)
    if area_ids:
        area_ids = area_ids.split(',')
        conditions &= (db.stations.area_ids.belongs(area_ids))

    time_range = request.vars.time_range
    if time_range:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=int(time_range))

    table = exeed_type_data
    if data_type:
        if int(data_type) == 2:
            if exeed_type_data == 'data_min':
                table = 'data_adjust'
            else:
                table = exeed_type_data + '_adjust'

    aaData = []  # Du lieu json se tra ve
    arrSorted = []
    list_data = None  # Du lieu truy van duoc
    iTotalRecords = 0  # Tong so ban ghi

    stations = db(conditions).select(db.stations.id,
                                     db.stations.station_name,
                                     db.stations.station_type,
                                     db.stations.province_id)

    query = db.stations.station_type == station_type
    if province_id:
        query &= db.stations.province_id == province_id
    if added_stations:
        query &= db.stations.id.belongs(added_stations)
    indicators = common.get_indicator_dict()
    if stations is None:
        raise HTTP(400, "Không đủ dữ liệu xuất excel, vui lòng thử lại!")
    else:
        temp_headers = []
        headers = []
        temp_headers.append('Tên trạm')
        for i, indicator in enumerate(added_columns):
            temp_headers.append(indicator.upper())
        headers.append(temp_headers)
        for header in headers:
            aaData.append(header)
    for c, station in enumerate(stations):
        count_total_dict = dict()
        station_id = str(station.id)
        conditions = {'station_id': station_id}
        conditions['is_exceed'] = True
        if from_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            conditions['get_time']['$gte'] = date_util.string_to_datetime(from_date)
        if to_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            conditions['get_time']['$lte'] = date_util.string_to_datetime(to_date) + timedelta(days=1)
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions['get_time'] = {'$gte': from_date, '$lte': to_date}

        attrs = {'get_time': 1, 'data': 1}
        list_data = pydb[table].find(conditions, attrs).sort('get_time', -1)

        # Tong so ban ghi khong thuc hien phan trang
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db(conditions).select(db.station_indicator.ALL)
        qcvn_dict = dict()
        for row in rows:
            indicator_name = indicators[row.indicator_id]
            qcvn_dict[indicator_name] = row.as_dict()
        count = dict()
        # for in
        for item in list_data:
            if item['data']:
                for indicator_name in item['data']:
                    x = str(item['data'][indicator_name])
                    if x == 'None' or x == 'NULL' or x == '-' or x.find(',') != -1:
                        continue
                    else:
                        z = float(x)
                        name_decode = indicator_name.encode('utf-8')
                        qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        check_qcvn = True
                        if z or z == 0:
                            if qcvn_max or qcvn_min:
                                check_qcvn = True
                            if qcvn_max and z > qcvn_max:
                                check_qcvn = False
                            if qcvn_min and z < qcvn_min:
                                check_qcvn = False
                            if not check_qcvn:
                                if not count.has_key(name_decode):
                                    count[name_decode] = 1
                                else:
                                    count[name_decode] = int(count[name_decode]) + 1
        row = [
            station.station_name,
        ]
        count_total_dict['total'] = 0
        for column in added_columns:
            if column and count.has_key(column):
                count_total_dict['total'] += count[column]
                row.append(count[column])
            else:
                row.append(0)
        count_total_dict['data'] = row
        arrSorted.append(count_total_dict)

        sorted(arrSorted, key=lambda x: x['total'])

        for item in arrSorted:
            if item:
                aaData.append(item['data'])

    # Get station name

    file_name = request.now.strftime('_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb = pyexcelerate.Workbook()
    wb.new_sheet('Data', aaData)
    wb.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


@service.json
def export_excel_2_a_station():
    import os.path, openpyxl, pyexcelerate
    # get search parameters
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    province_id = request.vars.province_id
    data_type = request.vars.data_type
    station_type = request.vars.station_type
    exeed_type_data = request.vars.exeed_type_data
    area_ids = request.vars.area_ids

    conditions = (db.stations.id > 0)
    added_stations = request.vars.added_stations or ''
    if added_stations:
        conditions &= (db.stations.id == added_stations)
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if station_type:
        conditions &= (db.stations.station_type == station_type)
    if province_id:
        conditions &= (db.stations.province_id == province_id)
    if area_ids:
        area_ids = area_ids.split(',')
        conditions &= (db.stations.area_ids.belongs(area_ids))

    time_range = request.vars.time_range
    if time_range:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=int(time_range))

    table = exeed_type_data
    if data_type:
        if int(data_type) == 2:
            if exeed_type_data == 'data_min':
                table = 'data_adjust'
            else:
                table = exeed_type_data + '_adjust'

    aaData = []  # Du lieu json se tra ve
    arrSorted = []
    list_data = None  # Du lieu truy van duoc
    iTotalRecords = 0  # Tong so ban ghi

    stations = db(conditions).select(db.stations.id,
                                     db.stations.station_name,
                                     db.stations.station_type,
                                     db.stations.province_id)

    query = db.stations.station_type == station_type
    if province_id:
        query &= db.stations.province_id == province_id
    if added_stations:
        query &= db.stations.id.belongs(added_stations)
    indicators = common.get_indicator_dict()
    if stations is None:
        raise HTTP(400, "Không đủ dữ liệu xuất excel, vui lòng thử lại!")
    else:
        temp_headers = []
        headers = []
        temp_headers.append('Tên trạm')
        for i, indicator in enumerate(added_columns):
            temp_headers.append(indicator.upper())
        headers.append(temp_headers)
        for header in headers:
            aaData.append(header)
    for c, station in enumerate(stations):
        count_total_dict = dict()
        station_id = str(station.id)
        conditions = {'station_id': station_id}
        conditions['is_exceed'] = True
        if from_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            conditions['get_time']['$gte'] = date_util.string_to_datetime(from_date)
        if to_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            conditions['get_time']['$lte'] = date_util.string_to_datetime(to_date) + timedelta(days=1)
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions['get_time'] = {'$gte': from_date, '$lte': to_date}

        attrs = {'get_time': 1, 'data': 1}
        list_data = pydb[table].find(conditions, attrs).sort('get_time', -1)

        # Tong so ban ghi khong thuc hien phan trang
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db(conditions).select(db.station_indicator.ALL)
        qcvn_dict = dict()
        for row in rows:
            indicator_name = indicators[row.indicator_id]
            qcvn_dict[indicator_name] = row.as_dict()
        count = dict()
        # for in
        for item in list_data:
            if item['data']:
                for indicator_name in item['data']:
                    x = str(item['data'][indicator_name])
                    if x == 'None' or x == 'NULL' or x == '-' or x.find(',') != -1:
                        continue
                    else:
                        z = float(x)
                        name_decode = indicator_name.encode('utf-8')
                        qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        check_qcvn = True
                        if z or z == 0:
                            if qcvn_max or qcvn_min:
                                check_qcvn = True
                            if qcvn_max and z > qcvn_max:
                                check_qcvn = False
                            if qcvn_min and z < qcvn_min:
                                check_qcvn = False
                            if not check_qcvn:
                                if not count.has_key(name_decode):
                                    count[name_decode] = 1
                                else:
                                    count[name_decode] = int(count[name_decode]) + 1
        row = [
            station.station_name,
        ]
        count_total_dict['total'] = 0
        for column in added_columns:
            if column and count.has_key(column):
                count_total_dict['total'] += count[column]
                row.append(count[column])
            else:
                row.append(0)
        count_total_dict['data'] = row
        arrSorted.append(count_total_dict)

        sorted(arrSorted, key=lambda x: x['total'])

        for item in arrSorted:
            if item:
                aaData.append(item['data'])

    # Get station name

    file_name = request.now.strftime('_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb = pyexcelerate.Workbook()
    wb.new_sheet('Data', aaData)
    wb.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


@service.json
def export_data_exceed_by_province():
    import os.path, openpyxl, pyexcelerate
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    exeed_type_data = request.vars.exeed_type_data
    province_id = request.vars.province_id

    conditions = (db.stations.id > 0)
    if province_id:
        conditions &= (db.stations.province_id == province_id)

    area_ids = request.vars.area_ids
    if area_ids:
        conditions &= (db.stations.area_ids.belongs([area_ids]))

    time_range = request.vars.time_range
    if time_range:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=int(time_range))

    table = exeed_type_data

    aaData = []  # Du lieu json se tra ve
    arrSorted = []
    list_data = None  # Du lieu truy van duoc

    stations = db(conditions).select(db.stations.id,
                                     db.stations.station_name,
                                     db.stations.station_type,
                                     db.stations.province_id)

    province_dict = common.get_province_dict()
    indicators = common.get_indicator_dict()
    if stations is None:
        raise HTTP(400, "Không đủ dữ liệu xuất excel, vui lòng thử lại!")
    else:
        temp_headers = []
        headers = []
        temp_headers.append('LBL_STT')
        temp_headers.append('Tỉnh/TP')
        temp_headers.append('Số lượng')
        headers.append(temp_headers)
        for header in headers:
            aaData.append(header)
        row_data = ["", "Tổng số trạm", len(stations)]
        aaData.append(row_data)

    group_by_province_dict = dict()
    for c, station in enumerate(stations):
        station_id = str(station.id)
        conditions = {'station_id': station_id}
        conditions['is_exceed'] = True
        if from_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            conditions['get_time']['$gte'] = date_util.string_to_datetime(from_date)
        if to_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            conditions['get_time']['$lte'] = date_util.string_to_datetime(to_date) + timedelta(days=1)
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions['get_time'] = {'$gte': from_date, '$lte': to_date}

        attrs = {'get_time': 1, 'data': 1}
        list_data = pydb[table].find(conditions, attrs).sort('get_time', -1)
        aaaa = pydb[table].count(conditions)
        # Tong so ban ghi khong thuc hien phan trang
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db(conditions).select(db.station_indicator.ALL)
        qcvn_dict = dict()
        for row in rows:
            indicator_name = indicators[row.indicator_id]
            qcvn_dict[indicator_name] = row.as_dict()
        count = dict()
        # for in
        for item in list_data:
            if item['data']:
                for indicator_name in item['data']:
                    x = str(item['data'][indicator_name])
                    if x == 'None' or x == 'NULL' or x == '-' or x.find(',') != -1:
                        continue
                    else:
                        z = float(x)
                        name_decode = indicator_name.encode('utf-8')
                        qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        check_qcvn = True
                        if z or z == 0:
                            if qcvn_max or qcvn_min:
                                check_qcvn = True
                            if qcvn_max and z > qcvn_max:
                                check_qcvn = False
                            if qcvn_min and z < qcvn_min:
                                check_qcvn = False
                            if not check_qcvn:
                                if group_by_province_dict.has_key(station.province_id):
                                    if station.id not in group_by_province_dict[station.province_id]:
                                        group_by_province_dict[station.province_id].append(station.id)
                                else:
                                    group_by_province_dict[station.province_id] = [station.id]
    i = 0
    for province in province_dict:
        if province:
            province_name = province_dict[province]
            total = 0
            if group_by_province_dict.has_key(province):
                total = len(group_by_province_dict[province])
            else:
                total = 0
            aaData.append([i + 1, province_name.encode('utf-8'), total])
        i += 1

    file_name = request.now.strftime('_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb = pyexcelerate.Workbook()
    wb.new_sheet('Data', aaData)
    wb.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


@service.json
def export_excel_common_a_station():
    import os.path, openpyxl, pyexcelerate, time
    # get search parameters
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    province_id = request.vars.province_id
    data_type = request.vars.data_type
    station_type = request.vars.station_type
    exeed_type_data = request.vars.exeed_type_data
    area_ids = request.vars.area_ids

    conditions = (db.stations.id > 0)

    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if station_type:
        conditions &= (db.stations.station_type == station_type)
    if province_id:
        conditions &= (db.stations.province_id == province_id)
    if area_ids:
        area_ids = area_ids.split(',')
        conditions &= (db.stations.area_ids.belongs(area_ids))

    time_range = request.vars.time_range
    if time_range:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=int(time_range))

    table = exeed_type_data
    if data_type:
        if int(data_type) == 2:
            if exeed_type_data == 'data_min':
                table = 'data_adjust'
            else:
                table = exeed_type_data + '_adjust'

    aaData = []  # Du lieu json se tra ve
    arrSorted = []
    list_data = None  # Du lieu truy van duoc
    iTotalRecords = 0  # Tong so ban ghi

    stations = db(conditions).select(db.stations.id,
                                     db.stations.station_name,
                                     db.stations.station_type,
                                     db.stations.province_id)

    query = db.stations.station_type == station_type
    if province_id:
        query &= db.stations.province_id == province_id

    indicators = common.get_indicator_dict()
    if stations is None:
        raise HTTP(400, "Không đủ dữ liệu xuất excel, vui lòng thử lại!")
    else:
        temp_headers = []
        headers = []
        temp_headers.append('Thông số')
        for i, indicator in enumerate(added_columns):
            temp_headers.append(indicator.upper())
        headers.append(temp_headers)
        for header in headers:
            aaData.append(header)

    count_all_from_station_to_file = dict()

    for c, station in enumerate(stations):
        count_total_dict = dict()
        station_id = str(station.id)
        conditions = {'station_id': station_id}
        if table == 'data_min' or table == 'data_adjust':
            conditions['is_exceed'] = True

        if from_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            conditions['get_time']['$gte'] = date_util.string_to_datetime(from_date)
        if to_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            conditions['get_time']['$lte'] = date_util.string_to_datetime(to_date) + timedelta(days=1)
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions['get_time'] = {'$gte': from_date, '$lte': to_date}

        attrs = {'get_time': 1, 'data': 1}
        list_data = pydb[table].find(conditions, attrs).sort('get_time', -1)

        # Tong so ban ghi khong thuc hien phan trang
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db(conditions).select(db.station_indicator.ALL)
        qcvn_dict = dict()

        for row in rows:
            indicator_name = indicators[row.indicator_id]
            qcvn_dict[indicator_name] = row.as_dict()
        count = dict()
        # for in
        for item in list_data:
            if item['data']:
                for indicator_name in item['data']:
                    if not count_all_from_station_to_file.has_key(indicator_name.encode('utf-8')):
                        count_all_from_station_to_file[indicator_name.encode('utf-8')] = {
                            "day_count_exceed": [],
                        }
                    else:
                        if not count_all_from_station_to_file[indicator_name.encode('utf-8')].has_key(
                                "day_count_exceed"):
                            count_all_from_station_to_file[indicator_name.encode('utf-8')]["day_count_exceed"] = []

                    x = str(item['data'][indicator_name])
                    if x == 'None' or x == 'NULL' or x == '-' or x.find(',') != -1:
                        continue
                    else:
                        if not count_all_from_station_to_file[indicator_name.encode('utf-8')].has_key("total_item"):
                            count_all_from_station_to_file[indicator_name.encode('utf-8')]["total_item"] = 1
                        else:
                            count_all_from_station_to_file[indicator_name.encode('utf-8')]["total_item"] = int(
                                count_all_from_station_to_file[indicator_name.encode('utf-8')]["total_item"]) + 1

                        z = float(x)
                        name_decode = indicator_name.encode('utf-8')
                        qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        check_qcvn = True
                        if z or z == 0:
                            if qcvn_max or qcvn_min:
                                check_qcvn = True
                            if qcvn_max and z > qcvn_max:
                                check_qcvn = False
                            if qcvn_min and z < qcvn_min:
                                check_qcvn = False
                            if not check_qcvn:
                                # value max
                                if not count_all_from_station_to_file[name_decode].has_key("max_exced_value"):
                                    count_all_from_station_to_file[name_decode]["max_exced_value"] = z
                                elif z > float(count_all_from_station_to_file[name_decode]["max_exced_value"]):
                                    count_all_from_station_to_file[name_decode]["max_exced_value"] = z

                                # value total exceed
                                if not count_all_from_station_to_file[name_decode].has_key("total_exceed"):
                                    count_all_from_station_to_file[name_decode]["total_exceed"] = 1
                                else:
                                    count_all_from_station_to_file[name_decode]["total_exceed"] = int(
                                        count_all_from_station_to_file[name_decode]["total_exceed"]) + 1

                                # day count
                                if len(count_all_from_station_to_file[name_decode]["day_count_exceed"]) == 0:
                                    count_all_from_station_to_file[name_decode]["day_count_exceed"].append(
                                        item["get_time"].strftime("%Y-%m-%d"))
                                elif item["get_time"].strftime("%Y-%m-%d") not in \
                                        count_all_from_station_to_file[name_decode]["day_count_exceed"]:
                                    count_all_from_station_to_file[name_decode]["day_count_exceed"].append(
                                        item["get_time"].strftime("%Y-%m-%d"))

                                if not count.has_key(name_decode):
                                    count[name_decode] = 1
                                else:
                                    count[name_decode] = int(count[name_decode]) + 1
    row = [
        ["Số ngày có giá trị vượt QCVN".encode('utf-8')],
        ["Số lượng giá trị vượt QCVN".encode('utf-8')],
        ["Tỷ lệ giá trị vượt QCVN (%)".encode('utf-8')],
        ["Giá trị vượt cao nhất".encode('utf-8')],
        ["Số lần vượt QCVN cao nhất".encode('utf-8')]
    ]
    for column in added_columns:
        if column and count_all_from_station_to_file.has_key(column):
            if count_all_from_station_to_file[column].has_key("day_count_exceed") and len(
                    count_all_from_station_to_file[column]["day_count_exceed"]) > 0:
                row[0].append(len(count_all_from_station_to_file[column]["day_count_exceed"]))
            else:
                row[0].append(0)

            if count_all_from_station_to_file[column].has_key("total_exceed") and int(
                    count_all_from_station_to_file[column]["total_exceed"]) > 0:
                row[1].append(int(count_all_from_station_to_file[column]["total_exceed"]))
            else:
                row[1].append(0)

            if count_all_from_station_to_file[column].has_key("total_exceed") and count_all_from_station_to_file[
                column].has_key("total_item") and float(count_all_from_station_to_file[column]["total_item"]) > 0:
                row[2].append("{:.2f}".format(float(
                    float(count_all_from_station_to_file[column]["total_exceed"]) / float(
                        count_all_from_station_to_file[column]["total_item"])) * 100))
            else:
                row[2].append(0)

            if count_all_from_station_to_file[column].has_key("max_exced_value") and float(
                    count_all_from_station_to_file[column]["max_exced_value"]) > 0:
                row[3].append(count_all_from_station_to_file[column]["max_exced_value"])
            else:
                row[3].append(0)

            if len(count_all_from_station_to_file[column]["day_count_exceed"]) > 0:
                row[4].append(len(count_all_from_station_to_file[column]["day_count_exceed"]))
            else:
                row[4].append(0)
        else:
            row[0].append(0)
            row[1].append(0)
            row[2].append(0)
            row[3].append(0)
            row[4].append(0)

    for r in row:
        aaData.append(r)

    # Get station name

    file_name = request.now.strftime('_common_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb = pyexcelerate.Workbook()
    wb.new_sheet('Data', aaData)
    wb.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


##################################################################################
@service.json
def export_excel_3():
    import os.path, openpyxl, pyexcelerate
    from_date = request.vars.from_date
    data_type = request.vars.data_type
    station_type = request.vars.station_type
    area_id = request.vars.area_id
    report_type = request.vars.report_type
    custom_ratio = request.vars.custom_ratio

    time_range = request.vars.time_range
    if time_range:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=int(time_range))

    if from_date:
        from_date = date_util.string_to_datetime(from_date)
    to_date = request.vars.to_date
    if to_date:
        to_date = date_util.string_to_datetime(to_date)
    aaData = []  # Du lieu json se tra ve
    if report_type == 'province':
        added_stations = request.vars.added_stations or ''
        if added_stations:
            added_stations = added_stations.split(',')

        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        table = 'data_min'
        if data_type:
            if int(data_type) == 2:
                table = 'data_adjust'
        conditions = db.stations.id.belongs(added_stations)
        stations = db(conditions).select(
            db.stations.id,
            db.stations.station_name,
            db.stations.frequency_receiving_data
        )
        if stations is None:
            raise HTTP(400, "Không đủ dữ liệu xuất excel, vui lòng thử lại!")
        else:
            headers = []
            temp_headers = ['Tên trạm', 'Tổng số dữ liệu nhận được', 'Tỉ lệ dữ liệu nhận được(%)']
            headers.append(temp_headers)
            for header in headers:
                aaData.append(header)
        for c, station in enumerate(stations):
            station_id = str(station.id)
            conditions = {'station_id': station_id}
            if from_date:
                if not conditions.has_key('get_time'):
                    conditions['get_time'] = {}
                conditions['get_time']['$gte'] = from_date
            if to_date:
                if not conditions.has_key('get_time'):
                    conditions['get_time'] = {}
                conditions['get_time']['$lte'] = to_date + timedelta(days=1)
            if not from_date and not to_date:
                to_date = datetime.now()
                from_date = to_date - timedelta(days=365)
                conditions['get_time'] = {'$gte': from_date, '$lte': to_date}
            delta = to_date - from_date
            if not station.frequency_receiving_data or station.frequency_receiving_data == 5:
                total_data = int((delta.days + 1) * 287)
            else:
                freq = station.frequency_receiving_data
                total_data = int((delta.days + 1) * 24 * 60 / freq)
            list_data = pydb[table].count(conditions)
            row = [
                station.station_name,
            ]
            content = '%s' % (list_data)
            row.append(content)
            v = float(list_data) / float(total_data) * 100
            v = "{0:.2f}".format(v)
            content = '%s' % (v)
            row.append(content)

            aaData.append(row)
    else:
        if custom_ratio == "" or custom_ratio == None:
            custom_ratio = "80"
        added_provinces = request.vars.added_provinces or ''
        if not added_provinces:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('No Province choosen'), success=True)
        else:
            headers = []
            temp_headers = ['STT', 'Tỉnh/TP', 'Số trạm', 'Tỉ lệ dữ liệu nhận được(%)']
            headers.append(temp_headers)
            for header in headers:
                aaData.append(header)
        added_provinces = added_provinces.split(',')

        result_o = dict()
        inx = 0
        for it in added_provinces:
            result_o[it] = inx
            inx += 1

        table = 'data_min'
        if data_type:
            if int(data_type) == 2:
                table = 'data_adjust'
        conditions = (db.stations.id > 0)
        if station_type:
            conditions &= (db.stations.station_type == station_type)
        if area_id:
            conditions &= (db.stations.area_id == area_id)
        if added_provinces:
            conditions &= (db.stations.province_id.belongs(added_provinces))
        stations = db(conditions).select(
            db.stations.id,
            db.stations.station_name,
            db.stations.province_id,
            db.stations.station_type,
            db.stations.frequency_receiving_data
        )
        provinces = dict()
        # Xu ly du lieu tinh
        if added_provinces:
            province_tmp = db(db.provinces.id.belongs(added_provinces)).select(db.provinces.province_name,
                                                                               db.provinces.id)
            for p in province_tmp:
                provinces[str(p.id)] = p.province_name
        result_dict = {}

        ids = [str(it.id) for it in stations]
        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'].belongs(ids))
        if from_date:
            conditions &= (db[table]['get_time'] >= from_date)
        if to_date:
            conditions &= (db[table]['get_time'] < to_date + timedelta(days=1))
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=30)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)
        count_func = db[table]['station_id'].count()
        data = db(conditions).select(db[table]['station_id'],
                                     db[table]['is_exceed'],
                                     count_func,
                                     groupby=db[table]['station_id'])
        data_rs = {}
        gtkey = '>=' + custom_ratio
        ekey = '<' + custom_ratio
        e0 = '=0'
        for item in data:
            if not data_rs.has_key(item[table]['station_id']):
                data_rs[item[table]['station_id']] = {'t': 0, 'e': 0}
            for k in item['_extra']:
                if item[table]['is_exceed']:
                    data_rs[item[table]['station_id']]['e'] = item['_extra'][k]
                data_rs[item[table]['station_id']]['t'] += item['_extra'][k]
        # duyet tren danh sach tram tra ve ket qua cuoi cung
        for station in stations:
            province_id = str(station.province_id)
            station_id = str(station.id)
            delta = to_date - from_date
            if not station.frequency_receiving_data or station.frequency_receiving_data == 5:
                total_data = int((delta.days + 1) * 287)
            else:
                freq = station.frequency_receiving_data
                total_data = int((delta.days + 1) * 24 * 60 / freq)

            if not result_dict.has_key(province_id):
                province_name = provinces[province_id] if provinces.has_key(province_id) else ''
                result_dict[province_id] = {
                    'order': result_o[province_id],
                    'name': province_name,
                    'total': 0,
                    'stations': [],
                    'info': {
                        'total': 0,  # Du lieu ly tuong
                        'received': 0,  # Du lieu nhan duoc
                        'exceed': 0,  # du lieu vuot chuan
                    },
                }
            r_station = {'name': station.station_name, 'total': total_data, 'received': 0, 'exceed': 0}
            if data_rs.has_key(station_id):
                r_station['received'] = data_rs[station_id]['t']
                r_station['exceed'] = data_rs[station_id]['e']
                result_dict[province_id]['info']['received'] += data_rs[station_id]['t']
                result_dict[province_id]['info']['exceed'] += data_rs[station_id]['e']
            result_dict[province_id]['info']['total'] += total_data  # Du lieu ly tuong
            result_dict[province_id]['stations'].append(r_station)
            result_dict[province_id]['total'] += 1
        inx = 0
        ls = result_dict.values()

        def func_s(e):
            return e['order']

        ls.sort(key=func_s)

        for item in ls:
            inx += 1
            row = [
                str(inx),
                item['name'],
            ]
            content = '%s' % (item['info']['received'])
            row.append(content)
            ratio = float(item['info']['received']) / float(item['info']['total']) * 100
            ratio = "{0:.2f}".format(ratio)
            content = '%s' % (ratio)
            row.append(content)
            aaData.append(row)
    wb = pyexcelerate.Workbook()
    wb.new_sheet('Data', aaData)

    # Get station name

    file_name = request.now.strftime('_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


# ################################################################################
@service.json
def export_excel_4():
    import os.path, openpyxl
    # get search parameters
    station_type = request.vars.station_type
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    province_id = request.vars.province_id
    data_type = request.vars.data_type
    exceed_type_data = request.vars.exceed_type_data
    table = exceed_type_data
    if data_type:
        if int(data_type) == 2:
            if table == 'data_min':
                table = 'data_adjust'
            else:
                table = exceed_type_data + '_adjust'
    added_columns = request.vars.custom_added_columns or ''
    added_stations = request.vars.added_stations or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if not station_type:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station type for viewing data'), success=True)
    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc
    iTotalRecords = 0  # Tong so ban ghi # Tuple dung de phan trang (vtri bat dau - chieu dai)
    conditions = db.stations.station_type == station_type
    if added_stations:
        added_stations = added_stations.split(',')
        conditions &= db.stations.id.belongs(added_stations)
    if province_id:
        conditions &= (db.stations.province_id == province_id)
    stations = db(conditions).select(
        db.stations.id,
        db.stations.station_name,
    )

    iTotalRecords = db(db.stations.station_type == station_type).count()
    iRow = 1
    indicators = common.get_indicator_dict()
    for c, station in enumerate(stations):
        station_id = station.id
        # station_ids = [str(item.id) for item in stations
        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
        if to_date:
            conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)
        # conditions &= (db[table]['is_exceed'] == True)

        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].station_id,
                                          db[table].data,
                                          )
        # Tong so ban ghi khong thuc hien phan trang
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db(conditions).select(db.station_indicator.ALL)
        qcvn_dict = dict()
        for row in rows:
            indicator_name = indicators[row.indicator_id]
            qcvn_dict[indicator_name] = row.as_dict()
        max = dict()
        min = dict()
        total = dict()
        count = dict()
        # for in
        for i, item in enumerate(list_data):
            if item.data:
                for indicator_name in item.data:
                    z = str(item.data[indicator_name])
                    if z == 'NULL' or z == 'None' or z == '-':
                        break
                    else:
                        x = z.replace(",", "")
                        v = float(x)
                    name_decode = indicator_name.encode('utf-8')
                    if v:
                        # max
                        if not max.has_key(name_decode):
                            max[name_decode] = v
                        else:
                            if v > max[name_decode]:
                                max[name_decode] = v
                        # min
                        if not min.has_key(name_decode):
                            min[name_decode] = v
                        else:
                            if v < min[name_decode]:
                                min[name_decode] = v
                        # trung binh
                        if not total.has_key(name_decode):
                            total[name_decode] = v
                            count[name_decode] = 1
                        else:
                            total[name_decode] += v
                            count[name_decode] += 1;
        row = [
            str(iRow + c),
            station.station_name,
        ]
        for column in added_columns:
            if column and count.has_key(column):
                row.append("{0:.2f}".format(max[column]))
                row.append("{0:.2f}".format(min[column]))
                row.append("{0:.2f}".format(total[column] / count[column]))
            else:
                row.append('-')
                row.append('-')
                row.append('-')
        aaData.append(row)

    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()
    temp_headers = []
    temp_headers_2 = []
    headers = []
    headers_2 = []
    temp_headers.append('#')
    temp_headers.append('Ten Tram')
    temp_headers_2.append(' ')
    temp_headers_2.append(' ')

    for item in added_columns:
        temp_headers.append('')
        temp_headers.append(str(item))
        temp_headers.append('')
    headers.append(temp_headers)
    for header in headers:
        ws2.append(header)

    for item in added_columns:
        temp_headers_2.append('Max')
        temp_headers_2.append('Min')
        temp_headers_2.append('TB')
    headers_2.append(temp_headers_2)
    for header in headers_2:
        ws2.append(header)
    # Write data
    for row in aaData:
        ws2.append(row)

    # Get station name

    file_name = request.now.strftime('_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


# ################################################################################
@service.json
def export_excel_5():
    import os.path, openpyxl
    # get search parameters
    station_type = request.vars.station_type
    province_id = request.vars.province_id
    from_date = request.vars.from_date
    if from_date:
        from_date = date_util.string_to_datetime(from_date)
    to_date = request.vars.to_date
    if to_date:
        to_date = date_util.string_to_datetime(to_date)
    view_type = request.vars.view_type
    added_stations = request.vars.added_stations or ''
    if added_stations:
        added_stations = added_stations.split(',')
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if not station_type:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station type for viewing data'), success=True)

    # get search parameters
    ticket_id = request.vars.ticket_id
    aaData = helper.getAADataByTicketReport(ticket_id, db)  # Du lieu json se tra ve
    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()
    temp_headers = []
    headers = []
    temp_headers.append(' ')
    temp_headers.append('Tên trạm')
    temp_headers.append('Tổng số dữ liệu')
    temp_headers.append('Nội dung')
    for item in added_columns:
        temp_headers.append(str(item))
    headers.append(temp_headers)
    for header in headers:
        ws2.append(header)
    # Write data
    for row in aaData:
        ws2.append(row)

    # Get station name

    file_name = request.now.strftime('_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


@service.json
def export_excel_5_by_station():
    import os.path, openpyxl
    # get search parameters
    station_type = request.vars.station_type
    from_date = request.vars.from_date
    if from_date:
        from_date = date_util.string_to_datetime(from_date)
    to_date = request.vars.to_date
    if to_date:
        to_date = date_util.string_to_datetime(to_date)

    time_range = request.vars.time_range
    if time_range:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=int(time_range))

    view_type = request.vars.view_type
    added_stations = request.vars.added_stations or ''
    if added_stations:
        added_stations = added_stations.split(',')
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if not station_type:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station type for viewing data'), success=True)
    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc
    table = 'data_min'
    conditions = db.stations.station_type == station_type
    if added_stations:
        conditions &= db.stations.id.belongs(added_stations)
    stations = db(conditions).select(
        db.stations.id,
        db.stations.station_name,
        db.stations.frequency_receiving_data,
    )

    table_adjust = 'data_adjust'
    station = stations[0]
    station_id = station.id
    delta = to_date - from_date
    if not station.frequency_receiving_data or station.frequency_receiving_data == 5:
        total_data = int((delta.days + 1) * 288)
    else:
        freq = station.frequency_receiving_data
        total_data = int((delta.days + 1) * 24 * 60 / freq)
    conditions = (db[table]['id'] > 0)
    conditions &= (db[table]['station_id'] == station_id)
    conditions_adjust = (db[table_adjust]['id'] > 0)
    conditions_adjust &= (db[table_adjust]['station_id'] == station_id)
    if from_date:
        conditions &= (db[table]['get_time'] >= from_date)
        conditions_adjust &= (db[table_adjust]['get_time'] >= from_date)
    if to_date:
        conditions &= (db[table]['get_time'] < to_date + timedelta(days=1))
        conditions_adjust &= (db[table_adjust]['get_time'] < to_date + timedelta(days=1))
    if not from_date and not to_date:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=365)
        conditions &= (db[table]['get_time'] >= from_date)
        conditions &= (db[table]['get_time'] < to_date)
        conditions_adjust &= (db[table_adjust]['get_time'] >= from_date)
        conditions_adjust &= (db[table_adjust]['get_time'] < to_date)
    list_data = db(conditions).select(db[table].id,
                                      db[table].get_time,
                                      db[table].station_id,
                                      db[table].data,
                                      )

    list_data_adjust = db(conditions_adjust).select(db[table_adjust].id,
                                                    db[table_adjust].get_time,
                                                    db[table_adjust].station_id,
                                                    db[table_adjust].data,
                                                    )

    row1 = [
        "1",
        "Số giá trị quan trắc theo thiết kế",
    ]
    count = dict()
    for i, item in enumerate(list_data):
        if item.data:
            for indicator_name in item.data:
                try:
                    z = str(item.data[indicator_name])
                    if z == 'NULL' or z == 'None' or z == '-':
                        continue
                    else:
                        x = z.replace(",", "")
                    name_decode = indicator_name.encode('utf-8')
                    if not count.has_key(name_decode):
                        count[name_decode] = 1
                    else:
                        count[name_decode] = int(count[name_decode]) + 1
                except:
                    pass
    for column in added_columns:
        if column and count.has_key(column):
            row1.append(total_data)
        else:
            row1.append(0)
    aaData.append(row1)

    row2 = [
        "2",
        "Số giá trị quan trắc nhận được",
    ]
    for column in added_columns:
        if column and count.has_key(column):
            try:
                v = count[column]
                row2.append(v)
            except:
                row2.append(0)
        else:
            row2.append(0)
    aaData.append(row2)

    count_adjust = dict()
    for i, item in enumerate(list_data_adjust):
        if item.data:
            for indicator_name in item.data:
                try:
                    z = str(item.data[indicator_name])
                    if z == 'NULL' or z == 'None' or z == '-':
                        continue
                    else:
                        x = z.replace(",", "")
                    name_decode = indicator_name.encode('utf-8')
                    if not count_adjust.has_key(name_decode):
                        count_adjust[name_decode] = 1
                    else:
                        count_adjust[name_decode] = int(count_adjust[name_decode]) + 1
                except:
                    pass
    row3 = [
        "3",
        "Số giá trị quan trắc lỗi",
    ]

    for column in added_columns:
        total_raw = 0
        total_adjust = 0
        if column and count.has_key(column):
            try:
                total_raw = count[column]
            except:
                total_raw = 0
        else:
            total_raw = 0
        if column and count_adjust.has_key(column):
            try:
                total_adjust = count_adjust[column]
            except:
                total_adjust = 0
        else:
            total_adjust = 0
        row3.append(str(total_raw - total_adjust))
    aaData.append(row3)

    row4 = [
        "4",
        "Số giá trị quan trắc hợp lệ"
    ]

    for column in added_columns:
        if column and count_adjust.has_key(column):
            row4.append(total_adjust)
        else:
            row4.append(0)
    aaData.append(row4)

    row5 = [
        "5",
        "Tỉ lệ dữ liệu nhận được so với số giá trị theo thiết kế",
    ]
    for column in added_columns:
        if column and count.has_key(column):
            try:
                v = float(count[column]) / float(total_data) * 100
                v = common.convert_data(v)
                row5.append(str(v) + "%")
            except:
                row5.append("0" + "%")
        else:
            row5.append("0" + "%")
    aaData.append(row5)

    row6 = [
        "6",
        "Tỉ lệ dữ liệu lỗi so với số giá trị theo thiết kế",
    ]

    for column in added_columns:
        total_raw = 0
        total_adjust = 0
        if column and count.has_key(column):
            try:
                total_raw = count[column]
            except:
                total_raw = 0
        else:
            total_raw = 0
        if column and count_adjust.has_key(column):
            try:
                total_adjust = count_adjust[column]
            except:
                total_adjust = 0
        else:
            total_adjust = 0
        try:
            v = float(total_raw - total_adjust) / float(total_data) * 100
            v = common.convert_data(v)
            row6.append(str(v) + "%")
        except:
            row6.append("0" + "%")
    aaData.append(row6)

    row7 = [
        "7",
        "Tỉ lệ dữ liệu hợp lệ so với số giá trị theo thiết kế",
    ]

    for column in added_columns:
        total_adjust = 0
        if column and count_adjust.has_key(column):
            try:
                total_adjust = count_adjust[column]
            except:
                total_adjust = 0
        else:
            total_adjust = 0
        try:
            v = float(total_adjust) / float(total_data) * 100
            v = common.convert_data(v)
            row7.append(str(v) + "%")
        except:
            row7.append("0" + "%")
    aaData.append(row7)

    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()
    temp_headers = []
    headers = []
    temp_headers.append(' ')
    temp_headers.append('Tên trạm')
    for item in added_columns:
        temp_headers.append(str(item))
    headers.append(temp_headers)
    for header in headers:
        ws2.append(header)
    # Write data
    for row in aaData:
        ws2.append(row)

    # Get station name

    file_name = request.now.strftime('_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


# ################################################################################
@service.json
def export_excel_6():
    from applications.eos.modules.w2pex import date_util
    import os.path, openpyxl, pyexcelerate
    # get search parameters
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    data_type = request.vars.data_type
    year = request.vars.Year
    month = request.vars.Month
    if data_type:
        data_type = int(data_type)
    # if year:
    #     year = int(year)
    #     from_date = datetime(year=year, month=1, day=1)
    #     to_date = datetime(year=year + 1, month=1, day=1)
    if month and year:
        month = int(month)
        year = int(year)
        from_date = datetime(year=year, month=month, day=1)
        to_date = date_util.get_first_day_next_month(from_date)
        to_date = datetime.combine(to_date, datetime.min.time())
    elif year:
        year = int(year)
        from_date = datetime(year=year, month=1, day=1)
        to_date = datetime(year=year + 1, month=1, day=1)

    time_range = request.vars.time_range
    if time_range:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=int(time_range))

    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc
    if data_type == 2:
        table = 'data_hour_adjust'
    else:
        table = 'data_hour'
    conditions = {'station_id': station_id}
    if from_date:
        if not conditions.has_key('get_time'):
            conditions['get_time'] = {}
        # conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
        conditions['get_time']['$gte'] = from_date
    if to_date:
        if not conditions.has_key('get_time'):
            conditions['get_time'] = {}
        # conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
        conditions['get_time']['$lte'] = to_date
    if not from_date and not to_date:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=365)
        conditions['get_time'] = {'$gte': from_date, '$lte': to_date}
    attrs = {'get_time': 1, 'data': 1}
    list_data = pydb[table].find(conditions, attrs).sort('get_time', -1)
    si_dict = dict()
    conditions = (db.station_indicator.station_id == station_id)
    conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
    rows = db(conditions).select(db.station_indicator.ALL)
    indicator_ids = []
    for row in rows:
        indicator_ids.append(row.indicator_id)
        si_dict[str(row.indicator_id)] = row.as_dict()
    indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                    db.indicators.unit)
    indicator_dict = dict()
    for it in indicators:
        if not indicator_dict.has_key(it['indicator']):
            indicator_dict[it['indicator']] = it.unit

    first_headers = ['Thông số']
    second_headers = ['Đơn vị']
    third_headers = ['Giờ']
    headers = []
    for item in added_columns:
        first_headers.append(str(item))
    for indicator in added_columns:
        if indicator_dict.has_key(indicator):
            second_headers.append(indicator_dict[indicator])
    headers.append(first_headers)
    headers.append(second_headers)
    headers.append(third_headers)
    for header in headers:
        aaData.append(header)
    data_dict = dict()
    for i in range(0, 24):
        if not data_dict.has_key(i):
            data_dict[i] = dict()
            for indi in indicators:
                if not data_dict[i].has_key(indi.indicator):
                    data_dict[i][indi.indicator] = dict(total=0, day=0)
    for item in list_data:
        get_time = item['get_time']
        key = get_time.hour
        for indicator in indicators:
            i_name = str(indicator.indicator)
            i_name_decode = i_name.encode('utf-8')
            v = item['data'][i_name_decode] if item['data'].has_key(i_name_decode) else ''
            if v == '' or v is None:
                pass
            else:
                data_dict[key][i_name]['total'] += float(v)
                data_dict[key][i_name]['day'] += 1
    for it in data_dict:
        row = [
            it
        ]
        for i in added_columns:
            row.append(round(data_dict[it][i]['total'] / data_dict[it][i]['day'], 2))
        aaData.append(row)
    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace(",", " ")

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb = pyexcelerate.Workbook()
    wb.new_sheet('Data', aaData)
    wb.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


# ################################################################################
@service.json
def export_excel_7():
    import os.path, openpyxl, pyexcelerate
    # get search parameters
    from applications.eos.modules.w2pex import date_util
    from calendar import monthrange
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date

    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)
    month = request.vars.Month
    year = request.vars.Year
    if month and year:
        month = int(month)
        year = int(year)
        from_date = datetime(year=year, month=month, day=1)
        to_date = date_util.get_first_day_next_month(from_date)
        to_date = datetime.combine(to_date, datetime.min.time())
    elif year:
        year = int(year)
        from_date = datetime(year=year, month=1, day=1)
        to_date = datetime(year=year + 1, month=1, day=1)

    time_range = request.vars.time_range
    if time_range:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=int(time_range))

    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if not station_id:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station for viewing data'), success=True)
    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc
    if data_type == 2:
        table = 'data_hour_adjust'
    else:
        table = 'data_hour'
    conditions = (db[table]['id'] > 0)
    conditions &= (db[table]['station_id'] == station_id)
    if from_date:
        conditions &= (db[table]['get_time'] >= from_date)
    if to_date:
        conditions &= (db[table]['get_time'] < to_date)

    list_data = db(conditions).select(db[table].id,
                                      db[table].get_time,
                                      db[table].data,
                                      orderby=~db[table].get_time)
    # Thu tu ban ghi
    si_dict = dict()
    conditions = (db.station_indicator.station_id == station_id)
    conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
    rows = db(conditions).select(db.station_indicator.ALL)
    indicator_ids = []
    for row in rows:
        indicator_ids.append(row.indicator_id)
        si_dict[str(row.indicator_id)] = row.as_dict()
    indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                    db.indicators.unit)
    indicator_dict = dict()
    for it in indicators:
        if not indicator_dict.has_key(it['indicator']):
            indicator_dict[it['indicator']] = it.unit

    first_headers = ['Thông số']
    second_headers = ['Đơn vị']
    third_headers = ['Ngày']
    headers = []
    for item in added_columns:
        first_headers.append(str(item))
    for indicator in added_columns:
        if indicator_dict.has_key(indicator):
            second_headers.append(indicator_dict[indicator])
    headers.append(first_headers)
    headers.append(second_headers)
    headers.append(third_headers)
    for header in headers:
        aaData.append(header)

    data_final = dict()
    for item in list_data:
        key = datetime.strftime(item.get_time, '%d/%m/%Y')
        if not data_final.has_key(key):
            data_final[key] = dict()
            for indicator in indicators:
                i_name = str(indicator.indicator)
                i_name_decode = i_name.decode('utf-8')
                v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                if v == '' or v is None:
                    pass
                else:
                    try:
                        v = float(v)
                        if not data_final[key].has_key(i_name):
                            data_final[key][i_name] = "{0:.2f}".format(v)
                        else:
                            if v > float(data_final[key][i_name]):
                                data_final[key][i_name] = "{0:.2f}".format(v)
                    except:
                        pass
    delta = timedelta(days=1)
    while from_date < to_date:
        key = datetime.strftime(from_date, "%d/%m/%Y")
        row = [
            key
        ]
        if not data_final.has_key(key):
            data_final[key] = dict()
            for column in added_columns:
                row.append('-')
        else:
            for column in added_columns:
                if column and data_final[key].has_key(column):
                    row.append(data_final[key][column])
                else:
                    row.append('-')
        aaData.append(row)
        from_date += delta

    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace(",", " ")

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb = pyexcelerate.Workbook()
    wb.new_sheet('Data', aaData)
    wb.save(file_path)
    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


# ################################################################################
@service.json
def export_excel_8():
    import os.path, openpyxl
    # get search parameters
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    month = request.vars.Month
    year = request.vars.Year
    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)
    if month and year:
        month = int(month)
        year = int(year)
        from_date = datetime(year=year, month=month, day=1)
        to_date = date_util.get_first_day_next_month(from_date)
        to_date = datetime.combine(to_date, datetime.min.time())
    elif year:
        year = int(year)
        from_date = datetime(year=year, month=1, day=1)
        to_date = datetime(year=year + 1, month=1, day=1)

    time_range = request.vars.time_range
    if time_range:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=int(time_range))

    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if not station_id:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station for viewing data'), success=True)
    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc
    iTotalRecords = 0  # Tong so ban ghi  # Tuple dung de phan trang (vtri bat dau - chieu dai)
    if data_type == 2:
        table = 'data_day_adjust'
    else:
        table = 'data_day'

    conditions = (db[table]['id'] > 0)
    conditions &= (db[table]['station_id'] == station_id)
    if from_date:
        conditions &= (db[table]['get_time'] >= from_date)
    if to_date:
        conditions &= (db[table]['get_time'] < to_date)
    if not from_date and not to_date:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=365)
        conditions &= (db[table]['get_time'] >= from_date)
        conditions &= (db[table]['get_time'] < to_date)

    list_data = db(conditions).select(db[table].id,
                                      db[table].get_time,
                                      db[table].data,
                                      orderby=db[table].get_time)
    # Tong so ban ghi khong thuc hien phan trang
    iTotalRecords = db(conditions).count(db[table].id)
    # Thu tu ban ghi
    si_dict = dict()
    conditions = (db.station_indicator.station_id == station_id)
    conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
    rows = db().select(db.station_indicator.ALL)
    indicator_ids = []
    for row in rows:
        indicator_ids.append(row.indicator_id)
        si_dict[str(row.indicator_id)] = row.as_dict()
    indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                    db.indicators.unit)
    indicator_dict = dict()
    for it in indicators:
        if not indicator_dict.has_key(it['indicator']):
            indicator_dict[it['indicator']] = it.unit
    # Duyet tung phan tu trong mang du lieu vua truy van duoc
    for i, item in enumerate(list_data):
        added_item = dict()
        if added_columns:
            for indicator in indicators:
                i_name = str(indicator.indicator)
                i_name_decode = i_name.decode('utf-8')
                v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                if v == '' or v is None:
                    added_item[i_name] = '-'
                else:
                    try:
                        v = float(v)
                        added_item[i_name] = "{0:.2f}".format(v)
                    except:
                        added_item[i_name] = '-'
        row = [
            item.get_time.strftime('%d/%m/%Y'),
        ]
        for column in added_columns:
            if column and added_item.has_key(column):
                row.append(added_item[column])
        aaData.append(row)
    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()
    if list_data.first() is None:
        raise HTTP(400, "Không đủ dữ liệu xuất excel, vui lòng thử lại!")
    else:
        # Write header
        first_headers = ['Thông số']
        second_headers = ['Đơn vị']
        third_headers = ['Ngày']
        headers = []
        for item in added_columns:
            first_headers.append(str(item))
        for indicator in added_columns:
            if indicator_dict.has_key(indicator):
                second_headers.append(indicator_dict[indicator])
        headers.append(first_headers)
        headers.append(second_headers)
        headers.append(third_headers)
        for header in headers:
            ws2.append(header)
        # Write data
        for row in aaData:
            ws2.append(row)

    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace(",", " ")

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


# ################################################################################
@service.json
def export_excel_9():
    import os.path, openpyxl
    # get search parameters
    from applications.eos.modules.w2pex import date_util
    from calendar import monthrange
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    month = request.vars.Month
    year = request.vars.Year
    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)

    # month_date = request.vars.month_date
    if month and year:
        month = int(month)
        year = int(year)
        from_date = datetime(year=year, month=month, day=1)
        to_date = date_util.get_first_day_next_month(from_date)
        to_date = datetime.combine(to_date, datetime.min.time())

    time_range = request.vars.time_range
    if time_range:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=int(time_range))

    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if not station_id:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station for viewing data'), success=True)
    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc
    if data_type == 2:
        table = 'data_hour_8h_adjust'
    else:
        table = 'data_hour_8h'

    conditions = (db[table]['id'] > 0)
    conditions &= (db[table]['station_id'] == station_id)
    if from_date:
        conditions &= (db[table]['get_time'] >= from_date)
    if to_date:
        conditions &= (db[table]['get_time'] < to_date)

    list_data = db(conditions).select(db[table].id,
                                      db[table].get_time,
                                      db[table].data,
                                      orderby=db[table].get_time)
    # limitby=limitby
    # Tong so ban ghi khong thuc hien phan trang
    iTotalRecords = monthrange(year, month)[1]
    # Thu tu ban ghi

    si_dict = dict()
    conditions = (db.station_indicator.station_id == station_id)
    conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
    rows = db().select(db.station_indicator.ALL)
    indicator_ids = []
    for row in rows:
        indicator_ids.append(row.indicator_id)
        si_dict[str(row.indicator_id)] = row.as_dict()
    indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                    db.indicators.unit)

    for i_count in range(1, iTotalRecords + 1):
        added_item = dict()
        for i, item in enumerate(list_data):
            if (item.get_time.day == i_count) and (item.get_time.hour >= 8):
                if added_columns:
                    for indicator in indicators:
                        i_name = str(indicator.indicator)
                        i_name_decode = i_name.decode('utf-8')
                        v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                        if v == '' or v is None:
                            pass
                        else:
                            try:
                                v = float(v)
                                if not added_item.has_key(i_name):
                                    added_item[i_name] = "{0:.2f}".format(v)
                                else:
                                    if v > int(added_item[i_name]):
                                        added_item[i_name] = "{0:.2f}".format(v)
                            except:
                                pass
        row = [
            str(i_count) + '/' + str(month),
        ]
        for column in added_columns:
            if column and added_item.has_key(column):
                row.append(added_item[column])
            else:
                row.append('-')
        aaData.append(row)

    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()
    if list_data.first() is None:
        raise HTTP(400, "Không đủ dữ liệu xuất excel,vui lòng thử lại!")
    else:
        # Write header
        temp_headers = []
        headers = []
        date_str = "%s/%s" % (month, year)
        temp_headers.append(date_str)
        for item in added_columns:
            temp_headers.append(str(item))
        headers.append(temp_headers)
        for header in headers:
            ws2.append(header)
        # Write data
        for row in aaData:
            ws2.append(row)

    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace(",", " ")

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


# ################################################################################

@service.json
def export_excel_report_aqi_day_in_time():
    from openpyxl.styles import Alignment
    import os.path, openpyxl
    # get search parameters
    from applications.eos.modules.w2pex import date_util
    from calendar import monthrange
    # So luong ban ghi se lay toi da
    station_id = request.vars.station_id
    by_time = request.vars.by_time
    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)
    year = request.vars.Year
    from datetime import datetime as ddd
    start_date = ddd(year=int(year), month=1, day=1)
    end_date = ddd(year=int(year), month=12, day=31)

    quarter = request.vars.quarter
    if quarter:
        if quarter == "1":
            start_date = ddd(year=int(year), month=1, day=1)
            end_date = ddd(year=int(year), month=3, day=31)
        elif quarter == "2":
            start_date = ddd(year=int(year), month=4, day=1)
            end_date = ddd(year=int(year), month=6, day=30)
        elif quarter == "3":
            start_date = ddd(year=int(year), month=7, day=1)
            end_date = ddd(year=int(year), month=9, day=30)
        elif quarter == "4":
            start_date = ddd(year=int(year), month=10, day=1)
            end_date = ddd(year=int(year), month=12, day=31)

    from applications.eos.services import aqi_service
    aaData = aqi_service.AqiService(pydb, T, db).get_report_aqi_day_in_time(str(station_id), int(data_type), start_date,
                                                                            end_date)

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    headers_1 = []
    headers_2 = []

    headers = []
    headers.append('#')
    headers_1.append(headers)

    headers = []
    headers.append('#')
    headers_2.append(headers)

    for header in headers_1:
        ws2.append(header)
    for header in headers_2:
        ws2.append(header)
    for row in aaData:
        ws2.append(row)
    ws2.merge_cells('A1:A2')
    ws2.merge_cells('Z1:Z2')
    ws2.merge_cells('B1:Y1')
    for col in ws2.columns:
        for cell in col:
            alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
            cell.alignment = alignment_obj

    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace(",", " ")

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


@service.json
def export_excel_10():
    from openpyxl.styles import Alignment
    import os.path, openpyxl
    # get search parameters
    from applications.eos.modules.w2pex import date_util
    from calendar import monthrange
    # So luong ban ghi se lay toi da
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    month = request.vars.Month
    year = request.vars.Year
    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)
    if month and year:
        month = int(month)
        year = int(year)
        from_date = datetime(year=year, month=month, day=1)
        to_date = date_util.get_first_day_next_month(from_date)
        to_date = datetime.combine(to_date, datetime.min.time())

    if not station_id:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station for viewing data'), success=True)
    aaData = []  # Du lieu json se tra ve
    # lay du lieu aqi hour
    if data_type == 2:
        table = 'aqi_data_adjust_hour'
    else:
        table = 'aqi_data_hour'
    conditions = (db[table]['id'] > 0)
    conditions &= (db[table]['station_id'] == station_id)
    if from_date:
        conditions &= (db[table]['get_time'] >= from_date)
    if to_date:
        conditions &= (db[table]['get_time'] < to_date)
    if not from_date and not to_date:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=365)
        conditions &= (db[table]['get_time'] >= from_date)
        conditions &= (db[table]['get_time'] < to_date)

    list_data = db(conditions).select(db[table].id,
                                      db[table].get_time,
                                      db[table].data,
                                      orderby=~db[table].get_time)
    # Tong so ban ghi khong thuc hien phan trang

    aqi_hour = dict()
    # Duyet tung phan tu trong mang du lieu vua truy van duoc
    for i, item in enumerate(list_data):
        day = item.get_time.day
        hour = item.get_time.hour
        key = (day, hour)
        v = item.data['aqi'] if item.data.has_key('aqi') else ''
        if v == '' or v is None:
            aqi_hour[key] = '-'
        else:
            try:
                v = float(v)
                aqi_hour[key] = "{0:.2f}".format(v)
            except:
                aqi_hour[key] = '-'
    #####################################################
    # lay du lieu aqi day
    if data_type == 2:
        table = 'aqi_data_adjust_24h'
    else:
        table = 'aqi_data_24h'
    conditions = (db[table]['id'] > 0)
    conditions &= (db[table]['station_id'] == station_id)
    if from_date:
        conditions &= (db[table]['get_time'] >= from_date)
    if to_date:
        conditions &= (db[table]['get_time'] < to_date)
    if not from_date and not to_date:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=365)
        conditions &= (db[table]['get_time'] >= from_date)
        conditions &= (db[table]['get_time'] < to_date)

    list_data = db(conditions).select(db[table].id,
                                      db[table].get_time,
                                      db[table].data_24h,
                                      orderby=~db[table].get_time)
    # Tong so ban ghi khong thuc hien phan trang

    aqi_day = dict()
    # Duyet tung phan tu trong mang du lieu vua truy van duoc
    for i, item in enumerate(list_data):
        day = item.get_time.day
        v = item.data_24h['aqi'] if item.data_24h.has_key('aqi') else ''
        if v == '' or v is None:
            aqi_day[day] = '-'
        else:
            try:
                v = float(v)
                aqi_day[day] = "{0:.2f}".format(v)
            except:
                aqi_day[day] = '-'

    #####################################################
    iTotalRecords = monthrange(year, month)[1]
    # Thu tu ban ghi
    # iRow = iDisplayStart + 1
    for i_count in range(1, iTotalRecords + 1):
        row = [
            str(i_count),
        ]
        for j in range(0, 24):
            key = (i_count, j)
            if aqi_hour.has_key(key):
                row.append(aqi_hour[key])
            else:
                row.append('-')
        if aqi_day.has_key(i_count):
            row.append(aqi_day[i_count])
        else:
            row.append('-')
        aaData.append(row)

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    headers_1 = []
    headers_2 = []

    headers = []
    headers.append('#')
    for i in range(0, 24):
        headers.append('AQI Hour')
    headers.append('AQI Day')
    headers_1.append(headers)

    headers = []
    headers.append('#')
    for i in range(0, 24):
        headers.append(i)
    headers.append('AQI Day')
    headers_2.append(headers)

    for header in headers_1:
        ws2.append(header)
    for header in headers_2:
        ws2.append(header)
    for row in aaData:
        ws2.append(row)
    ws2.merge_cells('A1:A2')
    ws2.merge_cells('Z1:Z2')
    ws2.merge_cells('B1:Y1')
    for col in ws2.columns:
        for cell in col:
            alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
            cell.alignment = alignment_obj

    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace(",", " ")

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


# ################################################################################
@service.json
def export_excel_11():
    import os.path, openpyxl, pyexcelerate
    # get search parameters
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    # month = request.vars.Month
    year = request.vars.Year
    # if month and year:
    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)
    if year:
        # month = int(month)
        year = int(year)
        from_date = datetime(year=year, month=1, day=1)
        to_date = datetime(year=year + 1, month=1, day=1)
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if not station_id:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station for viewing data'), success=True)
    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc

    if data_type == 2:
        table = 'data_mon_adjust'
    else:
        table = 'data_mon'

    conditions = (db[table]['id'] > 0)
    conditions &= (db[table]['station_id'] == station_id)
    if from_date:
        conditions &= (db[table]['get_time'] >= from_date)
    if to_date:
        conditions &= (db[table]['get_time'] < to_date)
    if not from_date and not to_date:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=365)
        conditions &= (db[table]['get_time'] >= from_date)
        conditions &= (db[table]['get_time'] < to_date)

    list_data = db(conditions).select(db[table].id,
                                      db[table].get_time,
                                      db[table].data,
                                      orderby=db[table].get_time)
    # Tong so ban ghi khong thuc hien phan trang
    iTotalRecords = 1
    # Thu tu ban ghi
    si_dict = dict()
    conditions = (db.station_indicator.station_id == station_id)
    conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
    rows = db().select(db.station_indicator.ALL)
    indicator_ids = []
    for row in rows:
        indicator_ids.append(row.indicator_id)
        si_dict[str(row.indicator_id)] = row.as_dict()
    indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                    db.indicators.unit)
    indicator_dict = dict()
    for it in indicators:
        if not indicator_dict.has_key(it['indicator']):
            indicator_dict[it['indicator']] = it.unit

    first_headers = ['Thông số']
    second_headers = ['Đơn vị']
    third_headers = ['Tháng']
    headers = []
    for item in added_columns:
        first_headers.append(str(item))
    for indicator in added_columns:
        if indicator_dict.has_key(indicator):
            second_headers.append(indicator_dict[indicator])
    headers.append(first_headers)
    headers.append(second_headers)
    headers.append(third_headers)
    for header in headers:
        aaData.append(header)

    result_dict = dict()
    total_result = dict()
    count_result = dict()
    # Duyet tung phan tu trong mang du lieu vua truy van duoc
    for i, item in enumerate(list_data):
        added_item = dict()
        if added_columns:
            for indicator in indicators:
                i_name = str(indicator.indicator)
                i_name_decode = i_name.decode('utf-8')
                v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                if v == '' or v is None:
                    added_item[i_name] = '-'
                else:
                    try:
                        v = float(v)
                        added_item[i_name] = "{0:.2f}".format(v)
                        if total_result.has_key(i_name):
                            total_result[i_name] += v
                            count_result[i_name] += 1
                        else:
                            total_result[i_name] = v
                            count_result[i_name] = 1
                    except:
                        added_item[i_name] = '-'
        row = [
            item.get_time.month
        ]
        for column in added_columns:
            if column and added_item.has_key(column):
                row.append(added_item[column])
        result_dict[item.get_time.month] = row
    for i in range(1, 13):
        if not result_dict.has_key(i):
            row = [
                i,
            ]
            for column in added_columns:
                row.append('-')
            aaData.append(row)
        else:
            aaData.append(result_dict[i])
    row = [
        'TB Năm'
    ]
    for column in added_columns:
        if total_result.has_key(column):
            avg = float(total_result[column]) / count_result[column]
            row.append("{0:.2f}".format(avg))
        else:
            row.append('-')
    aaData.append(row)
    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace(",", " ")

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb = pyexcelerate.Workbook()
    wb.new_sheet('Data', aaData)
    wb.save(file_path)
    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


# ################################################################################
@service.json
def export_excel_12():
    import os.path, openpyxl

    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()
    # get data
    station_id = request.vars.station_id
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    temp_headers = ['STT', 'Loại gián đoạn', 'Thông số', 'Thời gian bắt đầu', 'Thời gian kết thúc', 'Kéo dài(giây)']
    headers = []
    headers.append(temp_headers)
    for header in headers:
        ws2.append(header)

    from_date = request.vars.from_date
    to_date = request.vars.to_date
    indicator_array = request.vars.indicators
    if indicator_array:
        indicator_array = indicator_array.split(',')
    discontinuity_type = request.vars.discontinuity_type

    condition_fetch_indi = db.station_indicator.station_id == station_id
    condition_fetch_indi &= db.station_indicator.status == 1
    indicator_station = db(condition_fetch_indi).select(db.station_indicator.ALL)
    indicator_ids = [str(it.indicator_id) for it in indicator_station]
    condition_fetch_indi_1 = db.indicators.id.belongs(indicator_ids)
    condition_fetch_indi_1 &= db.indicators.indicator.belongs(indicator_array)
    indicator_fetch = db(condition_fetch_indi_1).select(db.indicators.indicator, db.indicators.id)
    indicator_ids = [str(i.id) for i in indicator_fetch]
    c = 0
    # disconnect
    if discontinuity_type in ['disconnect_type', 'all']:
        condition_search = get_condition_report_12(station_id, from_date, to_date,
                                                   'station_off_log', 'disconnect_type', 0)
        data = db(condition_search).select(db.station_off_log.start_off,
                                           db.station_off_log.end_off,
                                           db.station_off_log.duration,
                                           orderby=~db.station_off_log.start_off)
        if data:
            for c, item in enumerate(data):
                listToStr = ' '.join(map(str, indicator_array))
                row = [
                    c + 1,
                    'Mất kết nối',
                    listToStr,
                    item.start_off,
                    item.end_off,
                    item.duration
                ]
                ws2.append(row)
    # sensor error
    if discontinuity_type in ['device_failure', 'all']:
        condition_search = get_condition_report_12(station_id, from_date, to_date,
                                                   'sensor_trouble_history', 'device_failure', 2)
        condition_search &= db.sensor_trouble_history.indicator_id.belongs(indicator_ids)
        data = db(condition_search).select(db.sensor_trouble_history.indicator_name,
                                           db.sensor_trouble_history.indicator_id,
                                           db.sensor_trouble_history.get_time,
                                           orderby=db.sensor_trouble_history.get_time)
        data_sensor_trouble_history = dict()
        if data:
            for item in data:
                time = item.get_time - timedelta(minutes=5)
                if not data_sensor_trouble_history.has_key(item.indicator_id):
                    data_sensor_trouble_history[item.indicator_id] = {
                        item.get_time: dict(indicator_name=item.indicator_name,
                                            end_off=item.get_time + timedelta(minutes=5),
                                            start_off=item.get_time,
                                            duration=300,
                                            type='Lỗi thiết bị')}
                else:
                    if data_sensor_trouble_history[item.indicator_id].has_key(time):
                        data_sensor_trouble_history[item.indicator_id][time]['end_off'] = item.get_time + timedelta(
                            minutes=5)
                        data_sensor_trouble_history[item.indicator_id][time]['duration'] += 300
                        data_sensor_trouble_history[item.indicator_id][item.get_time] = data_sensor_trouble_history[
                            item.indicator_id].pop(time)

                    else:
                        data_sensor_trouble_history[item.indicator_id][item.get_time] = dict(
                            indicator_name=item.indicator_name,
                            end_off=item.get_time + timedelta(minutes=5), start_off=item.get_time, duration=300,
                            type='Lỗi thiết bị')
        inx = 0
        for indicator in data_sensor_trouble_history:
            data = data_sensor_trouble_history[indicator]
            for indx_sensor, key in enumerate(sorted(data)):
                row = [
                    indx_sensor + c + 1 + inx,
                    data[key]['type'],
                    data[key]['indicator_name'],
                    data[key]['start_off'],
                    data[key]['end_off'],
                    data[key]['duration'],
                ]
                ws2.append(row)
            inx = row[0] - 1
    # Calibration
    if discontinuity_type in ['calibration_type', 'all']:
        condition_search = get_condition_report_12(station_id, from_date, to_date,
                                                   'sensor_trouble_history', 'calibration_type', 1)
        condition_search &= db.sensor_trouble_history.indicator_id.belongs(indicator_ids)
        data = db(condition_search).select(db.sensor_trouble_history.indicator_name,
                                           db.sensor_trouble_history.indicator_id,
                                           db.sensor_trouble_history.get_time,
                                           orderby=db.sensor_trouble_history.get_time)
        data_calib = dict()
        if data:
            for item in data:
                time = item.get_time - timedelta(minutes=5)
                if not data_calib.has_key(item.indicator_id):
                    data_calib[item.indicator_id] = {
                        item.get_time: dict(indicator_name=item.indicator_name,
                                            end_off=item.get_time,
                                            start_off=item.get_time,
                                            duration=300,
                                            type='Hiệu chuẩn')}
                else:
                    if data_calib[item.indicator_id].has_key(time):
                        data_calib[item.indicator_id][time]['start_off'] = item.get_time
                        data_calib[item.indicator_id][time]['duration'] += 300
                        data_calib[item.indicator_id][item.get_time] = data_calib[
                            item.indicator_id].pop(time)

                    else:
                        data_calib[item.indicator_id][item.get_time] = dict(
                            indicator_name=item.indicator_name,
                            end_off=item.get_time, start_off=item.get_time, duration=300,
                            type='Hiệu chuẩn')
        for indicator in data_calib:
            data = data_calib[indicator]
            for indx_calib, key in enumerate(sorted(data)):
                row = [
                    indx_sensor + c + indx_calib + 1,
                    data[key]['type'],
                    data[key]['indicator_name'],
                    data[key]['start_off'],
                    data[key]['end_off'],
                    data[key]['duration'],
                ]
                ws2.append(row)

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name
    return data


################################################################################
@service.json
def export_excel_13():
    import os.path, openpyxl
    station_type = request.vars.station_type
    area_id = request.vars.area_id
    from_date = request.vars.from_date
    data_type = request.vars.data_type
    if from_date:
        from_date = date_util.string_to_datetime(from_date)
    to_date = request.vars.to_date
    if to_date:
        to_date = date_util.string_to_datetime(to_date)
    added_provinces = request.vars.added_stations or ''
    if not added_provinces:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('No Province choosen'), success=True)
    added_provinces = added_provinces.split(',')

    result_o = dict()
    inx = 0
    for it in added_provinces:
        result_o[it] = inx
        inx += 1

    aaData = []  # Du lieu json se tra ve
    table = 'data_min'
    if data_type:
        if int(data_type) == 2:
            table = 'data_adjust'
    conditions = (db.stations.id > 0)
    if station_type:
        conditions &= (db.stations.station_type == station_type)
    if area_id:
        conditions &= (db.stations.area_id == area_id)
    if added_provinces:
        conditions &= (db.stations.province_id.belongs(added_provinces))
    stations = db(conditions).select(
        db.stations.id,
        db.stations.station_name,
        db.stations.province_id,
        db.stations.station_type,
        db.stations.frequency_receiving_data
    )
    provinces = dict()
    # Xu ly du lieu tinh
    if added_provinces:
        province_tmp = db(db.provinces.id.belongs(added_provinces)).select(db.provinces.province_name, db.provinces.id)
        for p in province_tmp:
            provinces[str(p.id)] = p.province_name
    result_dict = {}

    ids = [str(it.id) for it in stations]
    conditions = (db[table]['id'] > 0)
    conditions &= (db[table]['station_id'].belongs(ids))
    if from_date:
        conditions &= (db[table]['get_time'] >= from_date)
    if to_date:
        conditions &= (db[table]['get_time'] < to_date + timedelta(days=1))
    if not from_date and not to_date:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=30)
        conditions &= (db[table]['get_time'] >= from_date)
        conditions &= (db[table]['get_time'] < to_date)
    count_func = db[table]['station_id'].count()
    data = db(conditions).select(db[table]['station_id'],
                                 db[table]['is_exceed'],
                                 count_func,
                                 groupby=db[table]['station_id'])
    data_rs = {}
    for item in data:
        if not data_rs.has_key(item[table]['station_id']):
            data_rs[item[table]['station_id']] = {'t': 0, 'e': 0}
        for k in item['_extra']:
            if item[table]['is_exceed']:
                data_rs[item[table]['station_id']]['e'] = item['_extra'][k]
            data_rs[item[table]['station_id']]['t'] += item['_extra'][k]
    # duyet tren danh sach tram tra ve ket qua cuoi cung
    for station in stations:
        province_id = str(station.province_id)
        station_id = str(station.id)
        delta = to_date - from_date
        if not station.frequency_receiving_data or station.frequency_receiving_data == 5:
            total_data = int((delta.days + 1) * 287)
        else:
            freq = station.frequency_receiving_data
            total_data = int((delta.days + 1) * 24 * 60 / freq)

        if not result_dict.has_key(province_id):
            province_name = provinces[province_id] if provinces.has_key(province_id) else ''
            result_dict[province_id] = {
                'order': result_o[province_id],
                'name': province_name,
                'total': 0,
                'stations': [],
                'info': {
                    'total': 0,  # Du lieu ly tuong
                    'received': 0,  # Du lieu nhan duoc
                    'exceed': 0,  # du lieu vuot chuan
                },
            }
        r_station = {'name': station.station_name, 'total': total_data, 'received': 0, 'exceed': 0}
        if data_rs.has_key(station_id):
            r_station['received'] = data_rs[station_id]['t']
            r_station['exceed'] = data_rs[station_id]['e']
            result_dict[province_id]['info']['received'] += data_rs[station_id]['t']
            result_dict[province_id]['info']['exceed'] += data_rs[station_id]['e']
        result_dict[province_id]['info']['total'] += total_data  # Du lieu ly tuong
        result_dict[province_id]['stations'].append(r_station)
        result_dict[province_id]['total'] += 1
    inx = 0
    ls = result_dict.values()

    def func_s(e):
        return e['order']

    ls.sort(key=func_s)

    for item in ls:
        inx += 1
        row = [
            str(inx),
            item['name'],
        ]
        row.append(item['info']['exceed'])
        content = '%s' % (item['info']['received'])
        row.append(content)
        ratio = float(item['info']['received']) / float(item['info']['total']) * 100
        ratio = "{0:.2f}".format(ratio)
        content = '%s' % (ratio)
        row.append(content)
        aaData.append(row)

    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()

    temp_headers = []
    headers = []
    temp_headers.append('#')
    temp_headers.append('Tỉnh')
    temp_headers.append('Số lượng dữ liệu vượt quy chuẩn')
    temp_headers.append('Tổng số dữ liệu nhận được')
    temp_headers.append('Tỉ lệ nhận được dữ liệu(%)')
    headers.append(temp_headers)
    for header in headers:
        ws2.append(header)
    # Write data
    for row in aaData:
        ws2.append(row)

    # Get station name

    file_name = request.now.strftime('_%Y%m%d_%H%M%S.xlsx')
    file_name = 'Thong ke ty le nhan du lieu theo tinh' + file_name
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


################################################################################
@service.json
def export_excel_14():
    import os.path, openpyxl
    station_type = request.vars.station_type
    area_id = request.vars.area_id
    to_date = request.vars.to_date
    if to_date:
        if to_date.find('/') == -1:
            to_date = datetime.strptime(to_date, '%Y-%m-%d')
        else:
            to_date = to_date.split(':')[0]
            to_date = datetime.strptime(to_date, '%Y/%m/%d %H')
    added_provinces = request.vars.added_stations or ''
    if not to_date:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=T('Date'), success=True)
    if not added_provinces:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=T('No Province choosen!'), success=True)
    added_provinces = added_provinces.split(',')
    result_o = dict()
    index = 0
    for it in added_provinces:
        result_o[it] = index
        index += 1
    aaData = []  # Du lieu json se tra ve
    table = 'data_hour_status_history'
    conditions = (db[table]['id'] > 0)
    if to_date:
        conditions = (db[table]['time'] == to_date)
    stations = {}
    data = db(conditions).select(db[table].data)
    if data:
        stations = data[0]['data']
    provinces = dict()
    # Xu ly du lieu tinh
    if added_provinces:
        province_tmp = db(db.provinces.id.belongs(added_provinces)).select(db.provinces.province_name, db.provinces.id)
        for p in province_tmp:
            provinces[str(p.id)] = p.province_name
    result_dict = {}
    for province_id in provinces:
        province_name = provinces[province_id] if provinces.has_key(province_id) else ''
        result_dict[province_id] = {
            'order': result_o[province_id],
            'name': province_name,
            'total': 0,
            'stations': [],
            'info': {
                'total': 0,  # Tong tram
                'connect': 0,  # Tong tram ket noi
                'adjusting': 0,  # Tong tram hieu chuan
                'exceed': 0,  # Tong tram vuot nguong
                'disconnect': 0,  # Tong tram mat ket noi
            },
        }
    # duyet tren danh sach tram tra ve ket qua cuoi cung
    for station in stations:
        station_id = str(station)
        is_get_station = True
        if station_type:
            if str(stations[station]['station_type']) != station_type:
                is_get_station = False
        if area_id:
            if str(stations[station]['area_id']) != area_id:
                is_get_station = False
        if added_provinces:
            if not str(stations[station]['province_id']) in added_provinces:
                is_get_station = False
        if is_get_station:
            province_id = str(stations[station]['province_id'])
            status = stations[station]['status']
            r_station = {'id': station_id, 'status': status}
            if status == 4:
                result_dict[province_id]['info']['disconnect'] += 1
            else:
                if status == 3:
                    result_dict[province_id]['info']['exceed'] += 1
                elif status == 5:
                    result_dict[province_id]['info']['adjusting'] += 1
                result_dict[province_id]['info']['connect'] += 1
            result_dict[province_id]['stations'].append(r_station)
            result_dict[province_id]['info']['total'] += 1
    inx = 0
    ls = result_dict.values()

    def func_s(e):
        return e['order']

    ls.sort(key=func_s)

    for item in ls:
        inx += 1
        row = [
            str(inx),
            item['name'],
        ]
        row.append(item['info']['total'])
        row.append(item['info']['connect'])
        row.append(item['info']['adjusting'])
        row.append(item['info']['exceed'])
        row.append(item['info']['disconnect'])
        if item['info']['total'] != 0:
            ratio = float(item['info']['connect']) / float(item['info']['total']) * 100
        else:
            ratio = 0.00
        ratio = "{0:.2f}".format(ratio)
        content = '%s (%s)' % (ratio, '%')
        row.append(content)
        aaData.append(row)
    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()

    temp_headers = []
    headers = []
    temp_headers.append('#')
    temp_headers.append('Tỉnh')
    temp_headers.append('Tổng số trạm')
    temp_headers.append('Số trạm kết nối')
    temp_headers.append('Số trạm hiệu chuẩn')
    temp_headers.append('Số trạm vượt ngưỡng')
    temp_headers.append('Số trạm mất kết nối')
    temp_headers.append('Tỉ lệ kết nối')
    headers.append(temp_headers)
    for header in headers:
        ws2.append(header)
    # Write data
    for row in aaData:
        ws2.append(row)

    # Get station name

    file_name = request.now.strftime('_%Y%m%d_%H%M%S.xlsx')
    file_name = 'Trang thai ket noi tai mot thoi diem theo tinh' + file_name
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


################################################################################
@service.json
def export_excel_15():
    import os.path, openpyxl
    from w2pex import date_util
    from datetime import datetime, date
    area_id = request.vars.area_id
    province_id = request.vars.province_id
    if not province_id and not area_id:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station type or an area for viewing data'), success=True)
    n_qty_this_month = 0
    expected_this_month = 0
    n_qty_last_month = 0
    expected_last_month = 0
    conditions = (db.stations.id > 0)
    if province_id and province_id != 'all_provinces':
        conditions &= (db.stations.province_id == province_id)
    if area_id:
        conditions &= (db.stations.area_id == area_id)
    # hungdx phan quyen quan ly trạm theo user issue 44
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id).select(
                db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))

    # station_ids = []
    # if area_id or province_id:
    ids = db(conditions).select(db.stations.id)
    station_ids = [str(item.id) for item in ids]

    first_date_in_this_month = date_util.get_first_day_current_month(date.today())
    first_date_in_this_month = datetime.combine(first_date_in_this_month, datetime.min.time())
    first_date_in_last_month = date_util.get_first_day_last_month(date.today())
    first_date_in_last_month = datetime.combine(first_date_in_last_month, datetime.min.time())
    first_date_in_next_month = date_util.get_first_day_next_month(date.today())
    first_date_in_next_month = datetime.combine(first_date_in_next_month, datetime.min.time())

    t1 = 0
    t2 = 0
    n_this_month = 0
    n_last_month = 0
    # alarm_level
    if True:
        alarm_level_this_month = dict()
        alarm_level_last_month = dict()
        for item in const.STATION_STATUS:
            alarm_level_this_month[item] = dict()
            alarm_level_this_month[item]['value'] = const.STATION_STATUS[item]['value']
            alarm_level_this_month[item]['name'] = str(T(const.STATION_STATUS[item]['name']))
            alarm_level_this_month[item]['color'] = const.STATION_STATUS[item]['color']
            alarm_level_this_month[item]['icon'] = const.STATION_STATUS[item]['icon']
            alarm_level_this_month[item]['qty'] = 0
            alarm_level_this_month[item]['percent'] = 0

            alarm_level_last_month[item] = dict()
            alarm_level_last_month[item]['value'] = const.STATION_STATUS[item]['value']
            alarm_level_last_month[item]['name'] = str(T(const.STATION_STATUS[item]['name']))
            alarm_level_last_month[item]['color'] = const.STATION_STATUS[item]['color']
            alarm_level_last_month[item]['icon'] = const.STATION_STATUS[item]['icon']
            alarm_level_last_month[item]['qty'] = 0
            alarm_level_last_month[item]['percent'] = 0
    # station_type
    if True:
        station_type_this_month = dict()
        station_type_last_month = dict()
        # for item in const.STATION_TYPE:
        for item in common.get_station_types():
            _key = str(item['value'])
            # This month
            station_type_this_month[_key] = dict()
            station_type_this_month[_key]['value'] = item['value']  # const.STATION_TYPE[item]['value']
            station_type_this_month[_key]['name'] = T(item['name'])  # str(T(const.STATION_TYPE[item]['name']))
            station_type_this_month[_key]['qty'] = 0
            # Last month
            station_type_last_month[_key] = dict()
            station_type_last_month[_key]['value'] = item['value']  # const.STATION_TYPE[item]['value']
            station_type_last_month[_key]['name'] = T(item['name'])  # str(T(const.STATION_TYPE[item]['name']))
            station_type_last_month[_key]['qty'] = 0

    alarm_level_this_month.pop('PREPARING')
    alarm_level_last_month.pop('PREPARING')
    alarm_level_this_month.pop('TENDENCY')
    alarm_level_last_month.pop('TENDENCY')

    stations = db(db.stations.id.belongs(station_ids)).select()
    ## tinh lượng datamin phải nhận trong tháng này
    conditions = (db.data_min_month_collect.station_id.belongs(station_ids))
    conditions &= (db.data_min_month_collect.date_month >= first_date_in_this_month)
    conditions &= (db.data_min_month_collect.date_month < first_date_in_next_month)
    data_min_month_this_month = db(conditions).select(db.data_min_month_collect.station_id,
                                                      db.data_min_month_collect.actual_datamin)
    data_min_month_this_month_dict = dict()
    for item in data_min_month_this_month:
        data_min_month_this_month_dict[item.station_id] = item
    expected_this_month = 0.0
    # số ngày trong tháng
    # days_this_month = (
    #         first_date_in_this_month.replace(month=first_date_in_this_month.month % 12 + 1, day=1) - timedelta(
    #     days=1)).day
    days_this_month = datetime.now().day - 1
    days_this_month += 1.0 / 24.0 * datetime.now().hour
    for row in stations:
        # tần suất nhận dữ liệu
        freq = row['frequency_receiving_data']
        # freq = 5
        indicator_number = db(db.station_indicator.station_id == row.id).count(db.station_indicator.indicator_id)
        if not indicator_number:
            indicator_number = 0
        if (not freq) or (freq == 0):
            freq = 5
        expected_this_month_each = (indicator_number * days_this_month * 24 * 60 / freq)
        if data_min_month_this_month_dict.has_key(str(row.id)):
            actual_this_month = data_min_month_this_month_dict[str(row.id)].actual_datamin

            if actual_this_month:
                if expected_this_month_each < actual_this_month:
                    expected_this_month_each = actual_this_month
        expected_this_month += expected_this_month_each
    # Count number records of data_min in this month
    if data_min_month_this_month:
        # conditions = (db.data_min_month_collect.station_id.belongs(station_ids))
        # conditions &= (db.data_min_month_collect.date_month >= first_date_in_this_month)
        # conditions &= (db.data_min_month_collect.date_month < first_date_in_next_month)
        data = db(conditions).select(
            db.data_min_month_collect.actual_datamin.sum().with_alias('actual_datamin'),
            db.data_min_month_collect.qty_good.sum().with_alias('qty_good'),
            db.data_min_month_collect.qty_exceed.sum().with_alias('qty_exceed'),
            db.data_min_month_collect.qty_adjusting.sum().with_alias('qty_adjusting'),
            db.data_min_month_collect.qty_error.sum().with_alias('qty_error'),
        ).first()
        n_this_month = data['actual_datamin'] if data['actual_datamin'] else 0
        t1 = n_this_month
        n_qty_this_month = data['actual_datamin'] if data['actual_datamin'] else 0
        # Tinh theo %
        if expected_this_month and expected_this_month > 0:
            n_this_month = round(100 * data['actual_datamin'] / expected_this_month, 2) if data[
                'actual_datamin'] else 0
        else:
            n_this_month = 0
        # n_this_month = "{:,}".format(n_this_month)
        alarm_level_this_month['GOOD']['qty'] = common.convert_data(data['qty_good'])
        alarm_level_this_month['GOOD']['percent'] = round(
            100.0 * data['qty_good'] / expected_this_month if expected_this_month > 0 and data['qty_good'] else 0,
            2)

        alarm_level_this_month['EXCEED']['qty'] = common.convert_data(data['qty_exceed'])
        alarm_level_this_month['EXCEED']['percent'] = round(
            100.0 * data['qty_exceed'] / expected_this_month if expected_this_month > 0 and data[
                'qty_exceed'] else 0,
            2)

        alarm_level_this_month['ADJUSTING']['qty'] = common.convert_data(data['qty_adjusting'])
        alarm_level_this_month['ADJUSTING']['percent'] = round(
            100.0 * data['qty_adjusting'] / expected_this_month if expected_this_month > 0 and data[
                'qty_adjusting'] else 0,
            2)
        alarm_level_this_month['ERROR']['qty'] = manh_test.moneyfmt(data['qty_error'])
        alarm_level_this_month['ERROR']['percent'] = round(
            100.0 * data['qty_error'] / expected_this_month if expected_this_month > 0 and data['qty_error'] else 0,
            2)

        offline_count = expected_this_month - data['qty_good'] - data['qty_exceed'] - data['qty_adjusting'] - data[
            'qty_error']
        alarm_level_this_month['OFFLINE']['qty'] = manh_test.moneyfmt(offline_count)
        alarm_level_this_month['OFFLINE']['percent'] = round(
            100.0 * offline_count / expected_this_month if expected_this_month > 0 and offline_count else 0,
            2)

    conditions = (db.data_min_month_collect.station_id.belongs(station_ids))
    conditions &= (db.data_min_month_collect.date_month >= first_date_in_last_month)
    conditions &= (db.data_min_month_collect.date_month < first_date_in_this_month)
    data_min_month_last_month = db(conditions).select(db.data_min_month_collect.station_id,
                                                      db.data_min_month_collect.actual_datamin)
    data_min_month_last_month_dict = dict()
    for item in data_min_month_last_month:
        data_min_month_last_month_dict[item.station_id] = item
    expected_last_month = 0.0
    # số ngày trong tháng
    days_last_month = (
            first_date_in_last_month.replace(month=first_date_in_last_month.month % 12 + 1, day=1) - timedelta(
        days=1)).day
    for row in stations:
        # tần suất nhận dữ liệu
        freq = row['frequency_receiving_data']
        # freq = 5
        indicator_number = db(db.station_indicator.station_id == row.id).count(db.station_indicator.indicator_id)
        if not indicator_number:
            indicator_number = 0
        if (not freq) or (freq == 0):
            freq = 5
        expected_last_month_each = (indicator_number * days_last_month * 24 * 60 / freq)
        # Trường hợp data nhận được nhiều hơn dự kiến (expected)
        if data_min_month_last_month_dict.has_key(str(row.id)):
            actual_last_month = data_min_month_last_month_dict[str(row.id)].actual_datamin
            if actual_last_month:
                if expected_last_month_each < actual_last_month:
                    expected_last_month_each = actual_last_month
            expected_last_month += expected_last_month_each

    # Count number records of data_min in last month
    if data_min_month_last_month:
        # conditions = (db.data_min_month_collect.station_id.belongs(station_ids))
        # conditions &= (db.data_min_month_collect.date_month >= first_date_in_last_month)
        # conditions &= (db.data_min_month_collect.date_month < first_date_in_this_month)
        data = db(conditions).select(
            db.data_min_month_collect.actual_datamin.sum().with_alias('actual_datamin'),
            db.data_min_month_collect.qty_good.sum().with_alias('qty_good'),
            db.data_min_month_collect.qty_exceed.sum().with_alias('qty_exceed'),
            db.data_min_month_collect.qty_adjusting.sum().with_alias('qty_adjusting'),
            db.data_min_month_collect.qty_error.sum().with_alias('qty_error'),
        ).first()
        n_last_month = data['actual_datamin'] if data['actual_datamin'] else 0
        t2 = n_last_month
        n_qty_last_month = data['actual_datamin'] if data['actual_datamin'] else 0
        if expected_last_month and expected_last_month > 0:
            n_last_month = round(100 * data['actual_datamin'] / expected_last_month, 2) if data[
                'actual_datamin'] else 0
        else:
            n_last_month = 0

        # alarm_level_last_month['TENDENCY']['qty'] = data['qty_good']
        # alarm_level_last_month['TENDENCY']['percent'] = round(
        #     100.0 * data['qty_good'] / data['actual_datamin'] if data['actual_datamin'] and data['qty_good'] else 0, 2)
        # alarm_level_last_month['TENDENCY']['name'] = T('Good')
        # alarm_level_last_month['EXCEED']['qty'] = data['qty_exceed']
        # alarm_level_last_month['EXCEED']['percent'] = round(
        #     100.0 * data['qty_exceed'] / data['actual_datamin'] if data['actual_datamin'] and data['qty_exceed'] else 0, 2)
        alarm_level_last_month['GOOD']['qty'] = manh_test.moneyfmt(data['qty_good'])
        alarm_level_last_month['GOOD']['percent'] = round(
            100.0 * data['qty_good'] / expected_last_month if expected_last_month > 0 and data['qty_good'] else 0,
            2)

        alarm_level_last_month['EXCEED']['qty'] = manh_test.moneyfmt(data['qty_exceed'])
        alarm_level_last_month['EXCEED']['percent'] = round(
            100.0 * data['qty_exceed'] / expected_last_month if expected_last_month > 0 and data[
                'qty_exceed'] else 0,
            2)

        alarm_level_last_month['ADJUSTING']['qty'] = manh_test.moneyfmt(data['qty_adjusting'])
        alarm_level_last_month['ADJUSTING']['percent'] = round(
            100.0 * data['qty_adjusting'] / expected_last_month if expected_last_month > 0 and data[
                'qty_adjusting'] else 0,
            2)
        alarm_level_last_month['ERROR']['qty'] = manh_test.moneyfmt(data['qty_error'])
        alarm_level_last_month['ERROR']['percent'] = round(
            100.0 * data['qty_error'] / expected_last_month if expected_last_month > 0 and data['qty_error'] else 0,
            2)

        offline_count = expected_last_month - data['qty_good'] - data['qty_exceed'] - data['qty_adjusting'] - data[
            'qty_error']
        alarm_level_last_month['OFFLINE']['qty'] = manh_test.moneyfmt(offline_count)
        alarm_level_last_month['OFFLINE']['percent'] = round(
            100.0 * offline_count / expected_last_month if expected_last_month > 0 and offline_count else 0,
            2)
    # So sanh gtri collect tong cong
    collect_icon = 'fa fa-arrow-up text-info'
    if t1 < t2:
        collect_icon = 'fa fa-arrow-down text-info'
    elif t1 == t2:
        collect_icon = 'fa fa-pause text-warning'

    # So sanh gtri 3 nguong thang nay va thang truoc de display Icon len/xuong cho dung
    for item in alarm_level_last_month:
        if alarm_level_this_month[item]['percent'] > alarm_level_last_month[item]['percent']:
            alarm_level_this_month[item]['icon'] = 'fa fa-arrow-up text-danger'
        if alarm_level_this_month[item]['percent'] < alarm_level_last_month[item]['percent']:
            alarm_level_this_month[item]['icon'] = 'fa fa-arrow-down text-info'
        else:
            alarm_level_this_month[item]['icon'] = 'fa fa-pause text-warning'

    # Count number records of station_off_log in this month
    if True:
        conditions = (db.data_min_month_collect.station_id.belongs(station_ids))
        conditions &= (db.data_min_month_collect.date_month >= first_date_in_this_month)
        conditions &= (db.data_min_month_collect.date_month < first_date_in_next_month)
        conditions &= (db.data_min_month_collect.number_off_log > 0)
        rows = db(conditions).select(
            db.data_min_month_collect.id,
            db.data_min_month_collect.station_type,
            db.data_min_month_collect.number_off_log
        )
        n2_this_month = 0
        for row in rows:
            n2_this_month += row.number_off_log
            for item in station_type_this_month:
                if row.station_type == station_type_this_month[item]['value']:
                    station_type_this_month[item]['qty'] += row.number_off_log
                    break
        for item in station_type_this_month:
            if station_type_this_month[item]['qty']:
                station_type_this_month[item]['qty'] = manh_test.moneyfmt(station_type_this_month[item]['qty'])
        t1 = n2_this_month
        n2_this_month = "{:,}".format(n2_this_month)

    # Count number records of station_off_log in last month
    if True:
        conditions = (db.data_min_month_collect.station_id.belongs(station_ids))
        conditions &= (db.data_min_month_collect.date_month >= first_date_in_last_month)
        conditions &= (db.data_min_month_collect.date_month < first_date_in_this_month)
        conditions &= (db.data_min_month_collect.number_off_log > 0)
        rows = db(conditions).select(
            db.data_min_month_collect.id,
            db.data_min_month_collect.station_type,
            db.data_min_month_collect.number_off_log
        )
        n2_last_month = 0
        for row in rows:
            n2_last_month += row.number_off_log
            for item in station_type_last_month:
                if row.station_type == station_type_last_month[item]['value']:
                    station_type_last_month[item]['qty'] += row.number_off_log
                    break
        for item in station_type_last_month:
            if station_type_last_month[item]['qty']:
                station_type_last_month[item]['qty'] = manh_test.moneyfmt(station_type_last_month[item]['qty'])
        t2 = n2_last_month
        n2_last_month = "{:,}".format(n2_last_month)

    # So sanh gtri offline cua station
    offline_icon = 'fa fa-arrow-up text-danger'
    if t1 < t2:
        collect_icon = 'fa fa-arrow-down text-info'
    elif t1 == t2:
        collect_icon = 'fa fa-pause text-warning'
    aaData = []
    aaData.append(['I', 'Dữ liệu thu thập',
                   str(n_this_month) + '%' + ' (' + manh_test.moneyfmt(n_qty_this_month) + ' dữ liệu đã thu thập trên '
                   + manh_test.moneyfmt(expected_this_month) + ' dữ liệu cần thu thập)',
                   str(n_last_month) + '%' + ' (' + manh_test.moneyfmt(n_qty_last_month) + ' dữ liệu đã thu thập trên '
                   + manh_test.moneyfmt(expected_last_month) + ' dữ liệu cần thu thập)'])
    aaData.append(['1', 'Hoạt động tốt',
                   str(alarm_level_this_month['GOOD']['percent']) + '% (' + alarm_level_this_month['GOOD'][
                       'qty'] + ' dữ liệu)',
                   str(alarm_level_last_month['GOOD']['percent']) + '% (' + alarm_level_last_month['GOOD'][
                       'qty'] + ' dữ liệu)'])
    aaData.append(['2', 'Vuợt quy chuẩn',
                   str(alarm_level_this_month['EXCEED']['percent']) + '% (' + alarm_level_this_month['EXCEED'][
                       'qty'] + ' dữ liệu)',
                   str(alarm_level_last_month['EXCEED']['percent']) + '% (' + alarm_level_last_month['EXCEED'][
                       'qty'] + ' dữ liệu)'])
    aaData.append(['3', 'Hiệu chuẩn',
                   str(alarm_level_this_month['ADJUSTING']['percent']) + '% (' + alarm_level_this_month['ADJUSTING'][
                       'qty'] + ' dữ liệu)',
                   str(alarm_level_last_month['ADJUSTING']['percent']) + '% (' + alarm_level_last_month['ADJUSTING'][
                       'qty'] + ' dữ liệu)'])
    aaData.append(['4', 'Lỗi thiết bị',
                   str(alarm_level_this_month['ERROR']['percent']) + '% (' + alarm_level_this_month['ERROR'][
                       'qty'] + ' dữ liệu)',
                   str(alarm_level_last_month['ERROR']['percent']) + '% (' + alarm_level_last_month['ERROR'][
                       'qty'] + ' dữ liệu)'])
    aaData.append(['5', 'Mất kết nối',
                   str(alarm_level_this_month['OFFLINE']['percent']) + '% (' + alarm_level_this_month['OFFLINE'][
                       'qty'] + ' dữ liệu)',
                   str(alarm_level_last_month['OFFLINE']['percent']) + '% (' + alarm_level_last_month['OFFLINE'][
                       'qty'] + ' dữ liệu)'])
    aaData.append(['II', 'Sự kiện mất dữ liệu',
                   n2_this_month + ' Các sự kiện mất dữ liệu',
                   n2_last_month + ' Các sự kiện mất dữ liệu'])
    aaData.append(['1', 'Nuớc mặt', station_type_this_month['1']['qty'], station_type_last_month['1']['qty']])
    aaData.append(['2', 'Nuớc thải', station_type_this_month['0']['qty'], station_type_last_month['0']['qty']])
    aaData.append(['3', 'Khí thải', station_type_this_month['3']['qty'], station_type_last_month['3']['qty']])
    # Lang update loi du lieu thi sai
    aaData.append(['4', 'Không khí', station_type_this_month['4']['qty'], station_type_last_month['4']['qty']])
    aaData.append(['5', 'Nuớc ngầm', station_type_this_month['2']['qty'], station_type_last_month['2']['qty']])

    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()

    temp_headers = []
    headers = []
    temp_headers.append('#')
    temp_headers.append('Nội dung')
    temp_headers.append('Tháng này')
    temp_headers.append('Tháng truớc')
    headers.append(temp_headers)
    for header in headers:
        ws2.append(header)
    # Write data
    for row in aaData:
        ws2.append(row)

    # Get station name

    file_name = request.now.strftime('_%Y%m%d_%H%M%S.xlsx')
    file_name = 'Thong ke du lieu theo tinh' + file_name
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


##################################################################################
@service.json
def export_excel_16():
    import os.path, pyexcelerate
    station_id = request.vars.station_id
    data_type = request.vars.data_type
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    ws2 = []

    temp_headers = ['Thông số', 'Số ngày có giá trị 1 giờ vượt QCNV', 'Số ngày trung bình 1 giờ vượt QCVN',
                    'Tỷ lệ giá trị trung bình 1 giờ vượt QCVN (%)']
    headers = []
    headers.append(temp_headers)
    for header in headers:
        ws2.append(header)
    table = 'data_hour'
    if data_type:
        if int(data_type) == 2:
            table = 'data_hour_adjust'
    conditions = {'station_id': station_id}
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    month = request.vars.Month
    year = request.vars.Year
    if month and year:
        year = int(year)
        if "quarter" in month:
            quarter = month.replace("quarter", "")
            quarter = int(quarter)
            if quarter == 1:
                from_date = datetime(year=year, month=1, day=1)
                to_date = datetime(year=year, month=4, day=1)
            elif quarter == 2:
                from_date = datetime(year=year, month=4, day=1)
                to_date = datetime(year=year, month=7, day=1)
            elif quarter == 3:
                from_date = datetime(year=year, month=7, day=1)
                to_date = datetime(year=year, month=10, day=1)
            elif quarter == 4:
                from_date = datetime(year=year, month=10, day=1)
                to_date = datetime(year=year + 1, month=1, day=1)
        else:
            month = int(month)
            from_date = datetime(year=year, month=month, day=1)
            to_date = date_util.get_first_day_next_month(from_date)
            to_date = datetime.combine(to_date, datetime.min.time())
    else:
        if from_date:
            from_date = date_util.string_to_datetime(from_date)
        if to_date:
            to_date = date_util.string_to_datetime(to_date) + timedelta(days=1)
    if from_date:
        if not conditions.has_key('get_time'):
            conditions['get_time'] = {}
        conditions['get_time']['$gte'] = from_date
    if to_date:
        if not conditions.has_key('get_time'):
            conditions['get_time'] = {}
        conditions['get_time']['$lte'] = to_date
    duration = (to_date - from_date).days
    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc
    iTotalRecords = len(added_columns)  # Tong so ban ghi
    rows = db(db.station_indicator.station_id == station_id).select(db.station_indicator.ALL,
                                                                    orderby=db.station_indicator.station_id)

    dic_station_indicator = dict()
    indicators = common.get_indicator_dict()

    for row in rows:
        if not dic_station_indicator.has_key(row['station_id']):
            dic_station_indicator[str(row['station_id'])] = dict()
        if indicators.has_key(row['indicator_id']):
            indicator_name = indicators[row['indicator_id']]
            dic_station_indicator[str(row['station_id'])][indicator_name] = row.as_dict()

    qcvn_dict = dic_station_indicator[station_id]

    dic_data_c = {}
    attrs = {'get_time': 1, 'data': 1}
    list_data = pydb[table].find(conditions, attrs).sort('get_time', 1)
    for item in list_data:
        timeconvert = item['get_time'].strftime("%m/%d/%Y")
        for i in added_columns:
            if item['data'].has_key(i):
                x = str(item['data'][i])
                if x == 'None' or x == 'NULL' or x == '-' or x.find(',') != -1:
                    continue
                else:
                    z = float(x)
                    name_decode = i.encode('utf-8')
                    qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                        name_decode) else None
                    qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                        name_decode) else None
                    check_qcvn = True
                    if z or z == 0:
                        if qcvn_max or qcvn_min:
                            check_qcvn = True
                        if qcvn_max and z > qcvn_max:
                            check_qcvn = False
                        if qcvn_min and z < qcvn_min:
                            check_qcvn = False
                        if not check_qcvn:
                            if not dic_data_c.has_key(timeconvert):
                                dic_data_c[timeconvert] = {i: {'total': 1, 'status': True}}
                            else:
                                if dic_data_c[timeconvert].has_key(i):
                                    dic_data_c[timeconvert][i]['total'] += 1
                                else:
                                    dic_data_c[timeconvert][i] = {'total': 1, 'status': True}
    for c, it in enumerate(added_columns):
        row = [
            it,
            0,
            0,
            0
        ]
        for i in dic_data_c:
            if dic_data_c[i].has_key(it):
                row[2] += dic_data_c[i][it]['total']
                row[1] += 1
        row[3] = round(float(row[2]) / float(duration * 24), 2)
        ws2.append(row)

    # Write data
    wb = pyexcelerate.Workbook()
    wb.new_sheet('Data', ws2)

    # Get station name

    file_name = request.now.strftime('_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


################################################################################
@service.json
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def get_indicators(*args, **kwargs):
    try:
        html = ''
        station_type = request.vars.type
        province_id = request.vars.province_id
        from_public = request.vars.from_public
        # conditions = (db.station_indicator.station_id == station_id)
        conditions = (db.indicators.id > 0)
        if station_type:
            conditions &= (db.indicators.indicator_type == station_type)
        rows = db(conditions).select(db.indicators.ALL)
        # Lay tram theo dieu kien
        stationIds = []
        cons = (db.stations.station_type == station_type)
        if province_id:
            cons &= (db.stations.province_id == province_id)
        # check tinhr

        stations = db(cons).select(db.stations.id)
        for s in stations:
            stationIds.append(str(s['id']))

        # lay thong so thuoc tram
        indiDic = []
        station_indicators = db(db.station_indicator.station_id.belongs(stationIds)).select(
            db.station_indicator.indicator_id)
        for ind in station_indicators:
            ind_key = ind['indicator_id']
            if ind_key not in indiDic:
                indiDic.append(ind_key)

        for row in rows:
            if str(row['id']) in indiDic:
                html += '<option value="%(value)s" selected>%(name)s (%(unit)s)</option>' % dict(value=row.indicator,
                                                                                                 name=row.indicator,
                                                                                                 unit=row.unit)
        return dict(success=True, html=html)
    except Exception as ex:
        return dict(success=False, message=str(ex))


################################################################################
@service.json
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def get_indicators_and_stations(*args, **kwargs):
    try:
        html = ''
        station_type = request.vars.type
        province_id = request.vars.province_id
        careers = request.vars.careers
        # conditions = (db.station_indicator.station_id == station_id)
        conditions = (db.indicators.id > 0)
        if station_type:
            conditions &= (db.indicators.indicator_type == station_type)
        rows = db(conditions).select(db.indicators.ALL)
        # Lay tram theo dieu kien
        stationIds = []
        cons = (db.stations.station_type == station_type)
        if province_id:
            cons &= (db.stations.province_id == province_id)
        if careers:
            career = careers.split(",")
            cons &= (db.stations.careers.belongs(career))
        # check tinhr
        if current_user:
            if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
                list_station_manager = db(db.manager_stations.user_id == current_user.id).select(
                    db.manager_stations.station_id)
                station_ids = [str(item.station_id) for item in list_station_manager]
                cons &= (db.stations.id.belongs(station_ids))
        stations = db(cons).select(db.stations.id)
        for s in stations:
            stationIds.append(str(s['id']))

        # lay thong so thuoc tram
        indiDic = []
        station_indicators = db(db.station_indicator.station_id.belongs(stationIds)).select(
            db.station_indicator.indicator_id)
        for ind in station_indicators:
            ind_key = ind['indicator_id']
            if ind_key not in indiDic:
                indiDic.append(ind_key)

        for row in rows:
            if str(row['id']) in indiDic:
                html += '<option value="%(value)s" selected>%(name)s (%(unit)s)</option>' % dict(value=row.indicator,
                                                                                                 name=row.indicator,
                                                                                                 unit=row.unit)

        html2 = ''
        # conditions = (db.stations.id > 0)
        conditions = (db.stations.id > 0)
        if station_type:
            conditions &= (db.stations.station_type == station_type)
        if province_id:
            conditions &= (db.stations.province_id == province_id)
        rows = db(conditions).select(
            db.stations.id,
            db.stations.station_name,
        )
        for row in rows:
            html2 += '<option value="%(value)s" selected>%(name)s</option>' % dict(value=row.id,
                                                                                   name=row.station_name)
        return dict(success=True, html=html, html_2=html2)
    except Exception as ex:
        return dict(success=False, message=str(ex))


################################################################################
@service.json
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def get_stations_and_province(*args, **kwargs):
    try:
        html = ''
        station_type = request.vars.type
        html2 = ''
        conditions = (db.stations.id > 0)
        if station_type:
            conditions &= (db.stations.station_type == station_type)
        if current_user:
            if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
                list_station_manager = db(db.manager_stations.user_id == current_user.id).select(
                    db.manager_stations.station_id)
                station_ids = [str(item.station_id) for item in list_station_manager]
                conditions &= (db.stations.id.belongs(station_ids))
        rows = db(conditions).select(
            db.stations.id,
            db.stations.station_name,
            db.stations.province_id
        )
        province = [str(it.province_id) for it in rows]
        province_dic = db(db.provinces.id.belongs(province)).select(db.provinces.id,
                                                                    db.provinces.province_name)
        if rows:
            html = "<option value='' selected>%s</option>" % T('-- Select province --')
            for item in province_dic:
                html += '<option value="%(value)s">%(name)s</option>' % dict(value=item.id, name=item.province_name)
        else:
            html = "<option value='' selected>%s</option>" % T('-- No data --')

        for row in rows:
            html2 += '<option value="%(value)s" selected>%(name)s</option>' % dict(value=row.id,
                                                                                   name=row.station_name)

        return dict(success=True, html=html, html_2=html2)
    except Exception as ex:
        return dict(success=False, message=str(ex))


#####################################################################
@service.json
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def get_provinces(*args, **kwargs):
    try:
        html = ''
        station_type = request.vars.type
        area = request.vars.area
        db = current.db
        conditions = db.stations.id > 0
        if station_type:
            conditions &= db.stations.station_type == station_type
        if area:
            conditions &= db.stations.area_id == area
        if current_user:
            if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
                list_station_manager = db(db.manager_stations.user_id == current_user.id).select(
                    db.manager_stations.station_id)
                station_ids = [str(item.station_id) for item in list_station_manager]
                conditions &= (db.stations.id.belongs(station_ids))
        stations = db(conditions).select(db.stations.province_id)
        province_ids = [station.province_id for station in stations]

        provinces = db(db.provinces.id.belongs(province_ids)).select()

        for row in provinces:
            html += '<option value="%(value)s" selected>%(name)s</option>' % dict(value=row.id, name=row.province_name)

        return dict(success=True, html=html)
    except Exception as ex:
        return dict(success=False, message=str(ex))


################################################################################
@service.json
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def get_indicators_and_stations_for_report_2(*args, **kwargs):
    try:
        html5 = ''
        html = ''
        html3 = ''
        station_type = request.vars.type
        province_id = request.vars.province_id
        area_ids = request.vars.area_ids
        careers = request.vars.careers
        conditions = (db.indicators.id > 0)
        province_dic = []
        areas = []

        area_id = []
        career = []
        if area_ids:
            area_id = area_ids.split(",")
        if careers:
            career = careers.split(",")

        # Lay tram theo dieu kien
        stationIds = []
        cons = db.stations.id > 0
        if station_type:
            cons &= (db.stations.station_type == station_type)
        if province_id:
            cons &= (db.stations.province_id == province_id)
        if len(area_id) > 0:
            cons &= (db.stations.area_ids.belongs(area_id))
        if len(career) > 0:
            cons &= (db.stations.career.belongs(career))

        # check tinh
        if current_user:
            if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
                list_station_manager = db(db.manager_stations.user_id == current_user.id).select(
                    db.manager_stations.station_id)
                station_ids = [str(item.station_id) for item in list_station_manager]
                cons &= (db.stations.id.belongs(station_ids))
        stations = db(cons).select(db.stations.id,
                                   db.stations.station_name)
        for s in stations:
            stationIds.append(str(s['id']))
        if station_type:
            conditions &= (db.indicators.indicator_type == station_type)
            query = db.stations.station_type == station_type
            query &= db.stations.id.belongs(stationIds)
            if len(stationIds) > 0:
                province_dic = db(query).select(db.stations.province_id, distinct=True)
        else:
            if len(stationIds) > 0:
                province_dic = db(db.stations.id.belongs(stationIds)).select(db.stations.province_id, distinct=True)
        rows = db(conditions).select(db.indicators.ALL)
        # lay thong so thuoc tram
        indiDic = []
        conds = db.station_indicator.station_id.belongs(stationIds)
        conds &= db.station_indicator.status == const.SI_STATUS['IN_USE']['value']
        station_indicators = db(conds).select(db.station_indicator.indicator_id)
        for ind in station_indicators:
            ind_key = ind['indicator_id']
            if ind_key not in indiDic:
                indiDic.append(ind_key)

        for row in rows:
            if str(row['id']) in indiDic:
                html += '<option value="%(value)s" selected>%(name)s (%(unit)s)</option>' % dict(value=row.indicator,
                                                                                                 name=row.indicator,
                                                                                                 unit=row.unit)

        html2 = ''
        # conditions = (db.stations.id > 0)
        for row in stations:
            html2 += '<option value="%(value)s" selected>%(name)s</option>' % dict(value=row.id,
                                                                                   name=row.station_name)
        if province_dic:
            province_ids = [str(item['province_id']) for item in province_dic]
            province_name = db(db.provinces.id.belongs(province_ids)).select(db.provinces.id,
                                                                             db.provinces.province_name)
            if province_name:
                html4 = "<option value=''>%s</option>" % T('-- Select province --')
                html3 = ["<option value='%s'>%s</option>" % (it['id'], it['province_name']) for it in province_name]
                html3 = html4 + ''.join(html3)

        areas = db(db.areas.id > 0).select(db.areas.id, db.areas.area_code, db.areas.area_name)
        for row in areas:
            html5 += '<option value="%(value)s" selected>%(name)s</option>' % dict(value=row.id,
                                                                                   name=row.area_name)

        return dict(success=True, html=html, html_2=html2, html3=html3, html5=html5, areas=areas)
    except Exception as ex:
        return dict(success=False, message=str(ex))


################################################################################
# hungdx issue 44 add (co the ap dung cho cac chon tinh, vung ... fix sau
@service.json
def dropdown_content(table, filter_field, get_value_field, get_dsp_field, *args, **kwargs):
    try:
        filter_value = request.vars.filter_value
        filter_value = filter_value.split(';')
        filter_field = filter_field.split('-')
        conditions = (db[table]['id'] > 0)
        t1 = len(filter_field)
        t2 = len(filter_value)
        station_ids = []
        if t1 == t2:
            for i in range(0, t1):
                if filter_field[i] and filter_value[i] != '':
                    conditions &= (db[table][filter_field[i]] == filter_value[i])
        if current_user:
            if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
                list_station_manager = db(db.manager_stations.user_id == current_user.id).select(
                    db.manager_stations.station_id)
                station_ids = [str(item.station_id) for item in list_station_manager]
                conditions &= (db[table]['id'].belongs(station_ids))

        records = db(conditions).select(db[table][get_value_field], db[table][get_dsp_field])

        if records:
            html1 = "<option value=''>%s</option>" % T('-- Select station --')
            html = ["<option value='%s'>%s</option>" % (item[get_value_field], item[get_dsp_field]) for item in records]
            html = html1 + ''.join(html)
        else:
            html = "<option value=''>%s</option>" % T('-- No data --')
        provinces = []
        if filter_value[0]:
            query = db.stations.station_type == filter_value[0]
            if station_ids:
                query &= db.stations.id.belongs(station_ids)
            data = db(query).select(db.stations.province_id, distinct=True)
            province_ids = [it.province_id for it in data]
            provinces = db(db.provinces.id.belongs(province_ids, null=True)).select(db.provinces.province_name,
                                                                                    db.provinces.id)
        if provinces:
            html3 = "<option value=''>%s</option>" % T('-- Select province --')
            html2 = ["<option value='%s'>%s</option>" % (name['id'], name['province_name']) for name in provinces]
            html2 = html3 + ''.join(html2)
        else:
            html2 = "<option value=''>%s</option>" % T('-- No data --')
        return dict(success=True, html=html, html2=html2)
    except Exception as ex:
        return dict(success=False, message=str(ex))


###########################################################
def get_condition_report_12(station_id, from_date, to_date, table_name, type, status):
    if station_id:
        condition_search = db[table_name].station_id == station_id
    if type == 'disconnect_type':
        if from_date:
            condition_search &= db[table_name].start_off >= datetime.strptime(from_date, '%Y-%m-%d')
        if to_date:
            condition_search &= db[table_name].end_off <= datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1)
    if type in ['device_failure', 'calibration_type']:
        if status:
            condition_search &= db[table_name].status == status
        if from_date:
            condition_search &= db[table_name].get_time >= datetime.strptime(from_date, '%Y-%m-%d')
        if to_date:
            condition_search &= db[table_name].get_time <= datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1)
    return condition_search


###########################################################
@service.json
def get_detail_for_report_2(*args, **kwargs):
    try:
        indicator = request.vars.indicator
        station_id = request.vars.station_id
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        data_type = request.vars.data_type
        table = 'data_min'
        if data_type:
            if int(data_type) == 2:
                table = 'data_adjust'
        conditions = {'station_id': station_id}
        if from_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            conditions['get_time']['$gte'] = date_util.string_to_datetime(from_date)
        if to_date:
            if not conditions.has_key('get_time'):
                conditions['get_time'] = {}
            conditions['get_time']['$lte'] = date_util.string_to_datetime(to_date) + timedelta(days=1)
        conditions['is_exceed'] = True
        attrs = {'get_time': 1, 'data': 1}
        list_data = pydb[table].find(conditions, attrs).sort('get_time', -1)

        rows = db(db.station_indicator.station_id == station_id).select(db.station_indicator.ALL,
                                                                        orderby=db.station_indicator.station_id)

        dic_station_indicator = dict()
        indicators = common.get_indicator_dict()

        for row in rows:
            if not dic_station_indicator.has_key(row['station_id']):
                dic_station_indicator[str(row['station_id'])] = dict()
            if indicators.has_key(row['indicator_id']):
                indicator_name = indicators[row['indicator_id']]
                dic_station_indicator[str(row['station_id'])][indicator_name] = row.as_dict()

        qcvn_dict = dic_station_indicator[station_id]
        record = []
        for item in list_data:
            x = str(item['data'][indicator])
            if x == 'None' or x == 'NULL' or x == '-' or x.find(',') != -1:
                continue
            else:
                z = float(x)
                name_decode = indicator.encode('utf-8')
                qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                    name_decode) else None
                qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                    name_decode) else None
                check_qcvn = True
                if z or z == 0:
                    if qcvn_max or qcvn_min:
                        check_qcvn = True
                    if qcvn_max and z > qcvn_max:
                        check_qcvn = False
                    if qcvn_min and z < qcvn_min:
                        check_qcvn = False
                    if not check_qcvn:
                        row = [
                            item['get_time'].strftime("%d/%m/%Y %H:%M"),
                            z
                        ]
                        record.append(row)
        return dict(aaData=record, success=True)
    except Exception as ex:
        return dict(message=str(ex), success=False)


###########################################################
@service.json
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def get_provinces_and_area_and_stations(*args, **kwargs):
    try:
        html = "<option value='' selected>%s</option>" % T('-- Select area --')
        html1 = "<option value='' selected>%s</option>" % T('-- Select province --')
        html2 = "<option value='' selected>%s</option>" % T('-- Select station --')
        station_type = request.vars.station_type
        careers = request.vars.careers

        conditions = db.stations.id > 0
        if current_user:
            if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
                list_station_manager = db(db.manager_stations.user_id == current_user.id).select(
                    db.manager_stations.station_id)
                station_ids = [item.station_id for item in list_station_manager]
                conditions &= (db['stations']['id'].belongs(station_ids, null=True))
        if station_type:
            conditions &= (db.stations.station_type == int(station_type))

        if careers:
            career = careers.split(",")
            conditions &= (db.stations.career.belongs(career))

        stations = db(conditions).select(db.stations.id,
                                         db.stations.station_name,
                                         db.stations.province_id,
                                         db.stations.area_id)
        area_ids = []
        for it in stations:
            html2 += '<option value="%(value)s">%(name)s</option>' % dict(value=it.id, name=it.station_name)
            area_ids.append(it.area_id)
        data_area = db(db.areas.id.belongs(area_ids, null=True)).select(db.areas.id,
                                                                        db.areas.area_name)
        for item in data_area:
            html += '<option value="%(value)s">%(name)s</option>' % dict(value=item.id, name=item.area_name)

        province_ids = [station.province_id for station in stations]
        provinces = db(db.provinces.id.belongs(province_ids, null=True)).select()

        for row in provinces:
            html1 += '<option value="%(value)s">%(name)s</option>' % dict(value=row.id, name=row.province_name)

        return dict(success=True, html=html, html1=html1, html2=html2)
    except Exception as ex:
        return dict(success=False, message=str(ex))


@service.json
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def get_provinces_and_stations_by_station_type(*args, **kwargs):
    try:
        html1 = "<option value='' selected>%s</option>" % T('-- Select province --')
        html2 = "<option value='' selected>%s</option>" % T('-- Select station --')
        station_type = request.vars.station_type
        area_id = request.vars.area_id
        # conditions = (db.stations.id > 0)
        conditions = (db.stations.id > 0)
        if area_id:
            conditions &= (db.stations.area_ids.belongs([area_id]))
        if station_type:
            conditions &= (db.stations.station_type == int(station_type))
        if current_user:
            if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
                list_station_manager = db(db.manager_stations.user_id == current_user.id).select(
                    db.manager_stations.station_id)
                station_ids = [str(item.station_id) for item in list_station_manager]
                conditions &= (db.stations.id.belongs(station_ids))
        rows = db(conditions).select(
            db.stations.id,
            db.stations.station_name,
            db.stations.province_id
        )
        province = [str(it.province_id) for it in rows]
        province_dic = db(db.provinces.id.belongs(province)).select(db.provinces.id,
                                                                    db.provinces.province_name)
        for item in province_dic:
            html1 += '<option value="%(value)s">%(name)s</option>' % dict(value=item.id, name=item.province_name)
        for row in rows:
            html2 += '<option value="%(value)s">%(name)s</option>' % dict(value=row.id, name=row.station_name)
        return dict(success=True, html1=html1, html2=html2)
    except Exception as ex:
        return dict(success=False, message=str(ex))


@service.json
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def get_stations_by_station_type_area_province(*args, **kwargs):
    try:
        html1 = "<option value='' selected>%s</option>" % T('-- Select station --')
        station_type = request.vars.station_type
        area_id = request.vars.area_id
        province_id = request.vars.province_id
        conditions = (db.stations.id > 0)
        if area_id:
            conditions &= (db.stations.area_id == area_id)
        if station_type:
            conditions &= (db.stations.station_type == int(station_type))
        if province_id:
            conditions &= (db.stations.province_id == province_id)
        if current_user:
            if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
                list_station_manager = db(db.manager_stations.user_id == current_user.id).select(
                    db.manager_stations.station_id)
                station_ids = [str(item.station_id) for item in list_station_manager]
                conditions &= (db.stations.id.belongs(station_ids))
        rows = db(conditions).select(
            db.stations.id,
            db.stations.station_name,
            db.stations.province_id
        )
        for row in rows:
            html1 += '<option value="%(value)s">%(name)s</option>' % dict(value=row.id, name=row.station_name)
        return dict(success=True, html1=html1)
    except Exception as ex:
        return dict(success=False, message=str(ex))


@service.json
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def get_provinces_and_area(*args, **kwargs):
    try:
        html = ''
        html1 = "<option value='' selected>%s</option>" % T('-- Select area --')
        station_type = request.vars.type
        area = request.vars.area
        db = current.db
        conditions = db.stations.id > 0
        if station_type:
            conditions &= db.stations.station_type == station_type
        if area:
            conditions &= db.stations.area_id == area
        if current_user:
            if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
                list_station_manager = db(db.manager_stations.user_id == current_user.id).select(
                    db.manager_stations.station_id)
                station_ids = [str(item.station_id) for item in list_station_manager]
                conditions &= (db.stations.id.belongs(station_ids))
        stations = db(conditions).select(db.stations.province_id,
                                         db.stations.area_id)
        area_ids = []
        province_ids = []
        for it in stations:
            area_ids.append(it.area_id)
            province_ids.append(it.province_id)
        provinces = db(db.provinces.id.belongs(province_ids)).select()
        areas = db(db.areas.id.belongs(area_ids)).select()
        for row in provinces:
            html += '<option value="%(value)s" selected>%(name)s</option>' % dict(value=row.id, name=row.province_name)
        for i in areas:
            html1 += '<option value="%(value)s" >%(name)s</option>' % dict(value=i.id, name=i.area_name)
        return dict(success=True, html=html, html1=html1)
    except Exception as ex:
        return dict(success=False, message=str(ex))


@service.json
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def get_provinces_for_report_15(*args, **kwargs):
    try:
        html = "<option value='' selected>%s</option>" % T('-- Select province --')
        station_type = request.vars.type
        area = request.vars.area
        db = current.db
        conditions = db.stations.id > 0
        if station_type:
            conditions &= db.stations.station_type == station_type
        if area:
            conditions &= db.stations.area_id == area
        if current_user:
            if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
                list_station_manager = db(db.manager_stations.user_id == current_user.id).select(
                    db.manager_stations.station_id)
                station_ids = [str(item.station_id) for item in list_station_manager]
                conditions &= (db.stations.id.belongs(station_ids))
        stations = db(conditions).select(db.stations.province_id)
        province_ids = [station.province_id for station in stations]

        provinces = db(db.provinces.id.belongs(province_ids)).select()

        for row in provinces:
            html += '<option value="%(value)s">%(name)s</option>' % dict(value=row.id, name=row.province_name)
        return dict(success=True, html=html)
    except Exception as ex:
        return dict(success=False, message=str(ex))


def get_all_areas(*args, **kwargs):
    try:
        # areas
        areas = common.get_area_by_station_dict()
        return dict(success=True, areas=areas)
    except Exception as ex:
        return dict(success=False, message=str(ex))
