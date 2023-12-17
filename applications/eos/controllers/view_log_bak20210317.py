# coding=utf-8
# encoding: utf-8
# encoding=utf8
# -*- coding: utf-8 -*-
###############################################################################
# Author :
# Date   :
#
# Description :
#
###############################################################################

import sys
reload(sys)
sys.setdefaultencoding('utf8')
from applications.eos.modules import common
from w2pex import date_util
from datetime import datetime, timedelta


def call():
    return service()


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_log')))
def index():
    now = datetime.now()
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value']
    station_id = request.vars.station_id
    provinces = common.get_province_have_station()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly trạm theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select(orderby=db.stations.order_no)
    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces, month=now.month, year=now.year)

################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_log')))
def advanced():
    now = datetime.now()
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value']
    station_id = request.vars.station_id
    provinces = common.get_province_have_station()
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select(db.stations.id, orderby=db.stations.order_no)
    station_ids = [str(item.id) for item in stations]
    data_indicator = db(db.station_indicator.station_id.belongs(station_ids)).select(db.station_indicator.indicator_id)
    indicator_ids = [str(it.indicator_id) for it in data_indicator]
    indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.indicator,db.indicators.id)
    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces, indicators=indicators)

################################################################################
@service.json
def get_list_log(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_id = request.vars.station_id
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        data_type = request.vars.data_type
        if data_type:
            data_type = int(data_type)

        month = request.vars.Month
        year = request.vars.Year


        duration = request.vars.durationf
        is_exceed = request.vars.is_exceed
        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        view_type = request.vars.view_type or const.VIEW_BY['MINUTE']['value']
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        aaData = []  # Du lieu json se tra ve

        view_by_info = common.get_const_by_value(const.VIEW_BY, view_type)
        table = view_by_info['table']

        if data_type and (
                data_type == const.DATA_TYPE['YET_APPROVED']['value'] or data_type == const.DATA_TYPE['APPROVED'][
            'value']):
            view_by_info_adjust = common.get_const_by_value(const.VIEW_BY_ADJUST, view_type)
            table = view_by_info_adjust['table']
            # table = 'data_adjust'

        # conditions = (db[table]['id'] > 0)
        # conditions &= (db[table]['station_id'] == station_id)
        conditions = {'station_id': station_id}
        if duration:  # Todo: Not clear
            duration = int(duration)
            duration = datetime.now() - timedelta(days=duration)
            # conditions &= (db[table]['get_time'] >= duration)
            conditions['get_time'] = {'$gte': duration}
        else:
            if month and year:
                month = int(month)
                year = int(year)
                from_date = datetime(year=year, month=month, day=1)
                to_date = date_util.get_first_day_next_month(from_date)
                to_date = datetime.combine(to_date, datetime.min.time())
                if from_date:
                    if not conditions.has_key('get_time'):
                        conditions['get_time'] = {}
                    conditions['get_time']['$gte'] = from_date
                    # conditions &= (db[table]['get_time'] >= from_date)
                if to_date:
                    if not conditions.has_key('get_time'):
                        conditions['get_time'] = {}
                    # conditions &= (db[table]['get_time'] < to_date)
                    conditions['get_time']['$lte'] = to_date
            else:
                if from_date:
                    if not conditions.has_key('get_time'):
                        conditions['get_time'] = {}
                    # conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
                    conditions['get_time']['$gte'] = date_util.string_to_datetime(from_date)
                if to_date:
                    if not conditions.has_key('get_time'):
                        conditions['get_time'] = {}
                    # conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
                    conditions['get_time']['$lte'] = date_util.string_to_datetime(to_date) + timedelta(days=1)
                if not from_date and not to_date:
                    to_date = datetime.now()
                    from_date = to_date - timedelta(days=365)
                    conditions['get_time'] = {'$gte': from_date, '$lte': to_date}

        if is_exceed and table == 'data_min':
            conditions['is_exceed'] = True
            # conditions &= (db[table]['is_exceed'] == True)
        attrs = {'get_time': 1}
        if table != 'aqi_data_24h' and table != 'aqi_data_adjust_24h':
            attrs['data'] = 1
            # list_data = db(conditions).select(db[table].id,
            #                                   db[table].get_time,
            #                                   db[table].data,
            #                                   orderby=~db[table].get_time,
            #                                   limitby=limitby)
        else:
            attrs['data_24h'] = 1
            # list_data = db(conditions).select(db[table].id,
            #                                   db[table].get_time,
            #                                   db[table].data_24h,
            #                                   orderby=~db[table].get_time,
            #                                   limitby=limitby)
        list_data = pydb[table].find(conditions, attrs).sort('get_time', -1)\
            .skip(iDisplayStart).limit(iDisplayLength)
        # Tong so ban ghi khong thuc hien phan trang
        iTotalRecords = pydb[table].count(conditions)
        # Thu tu ban ghi
        iRow = iDisplayStart + 1
        rows = db(db.station_indicator.station_id == station_id).select(db.station_indicator.ALL,
                                                                        orderby=db.station_indicator.station_id)

        dic_station_indicator = dict()
        indicators = common.get_indicator_dict()

        for row in rows:
            if not dic_station_indicator.has_key(row['station_id']):
                dic_station_indicator[str(row['station_id'])] = dict()
            if indicators.has_key(row['indicator_id']):
                indicator_name = indicators[row['indicator_id']]
                dic_station_indicator[str(row['station_id'])][indicator_name] = row.as_dict()

        qcvn_dict = dic_station_indicator[station_id]
        # Duyet tung phan tu trong mang du lieu vua truy van duoc
        if is_exceed:
            for c, item in enumerate(list_data):
                row = [
                    c + 1,
                    item['get_time'].strftime("%H:%M %d/%m/%Y")
                ]
                for indicator in added_columns:
                    x = str(item['data'][indicator])
                    if x == 'None' or x == 'NULL' or x == '-' or x.find(',') != - \
                            1:
                        row.append('-')
                    else:
                        z = float(x)
                        name_decode = indicator.encode('utf-8')
                        qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        check_qcvn = True
                        if z or z == 0:
                            if qcvn_max or qcvn_min:
                                check_qcvn = True
                                title = 'QCVN Min: %s' % qcvn_min + ' - QCVN Max: %s' % qcvn_max
                            if qcvn_max and z > qcvn_max:
                                check_qcvn = False
                                title = 'QCVN Max: %s' % qcvn_max
                            if qcvn_min and z < qcvn_min:
                                check_qcvn = False
                                title = 'QCVN Min: %s' % qcvn_min
                            if not check_qcvn:
                                row.append(
                                    SPAN(z, _style="background:#EA3223; color:white", _class="badge", _title=title))
                            else:
                                row.append(z)
                aaData.append(row)
        else:
            for i, item in enumerate(list_data):
                row = [str(iRow + i)]
                if table != 'data_mon' and table != 'data_mon_adjust':
                    row.append(item['get_time'].strftime(datetime_format_vn))
                else:
                    list = item['get_time'].strftime(datetime_format_vn).split(' ')[1].split('/')
                    list.pop(0)
                    row.append(list[0] + '/' + list[1])
                for indicator in added_columns:
                    i_name = indicator
                    i_name_decode = i_name.decode('utf-8')
                    if table != 'aqi_data_24h' and table != 'aqi_data_adjust_24h':
                        v = item['data'][i_name_decode] if item['data'].has_key(i_name_decode) else ''
                    else:
                        v = item['data_24h'][i_name_decode] if item['data_24h'].has_key(i_name_decode) else ''
                    if v == '' or v is None:
                        val = '-'
                    else:
                        try:
                            val = "{0:.2f}".format(float(v))
                        except:
                            val = '-'
                    row.append(val)
                aaData.append(row)

        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


################################################################################
@service.json
def get_list_max_min(*args, **kwargs):
    try:
        from datetime import datetime, timedelta
        from w2pex import date_util
        station_type = request.vars.station_type
        station_id = request.vars.station_id
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        duration = request.vars.duration
        is_exceed = request.vars.is_exceed
        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns[0].split(',')

        month = request.vars.Month
        year = request.vars.Year
        if month and year:
            month = int(month)
            year = int(year)
            from_date = datetime(year=year, month=month, day=1)
            to_date = date_util.get_first_day_next_month(from_date)
            to_date = datetime.combine(to_date, datetime.min.time())

        view_type = request.vars.view_type or const.VIEW_BY['MINUTE']['value']
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        aaData = []  # Du lieu json se tra ve
        view_by_info = common.get_const_by_value(const.VIEW_BY, view_type)
        table = view_by_info['table']
        # conditions = (db[table]['id'] > 0)
        conditions = {}# (db[table].id > 0)
        if duration:  # Neu co duration thi 2 dk from/to bi discard
            duration = int(duration)
            duration = datetime.now() - timedelta(days=duration)
            # conditions = (db[table]['get_time'] >= duration)
            conditions['get_time'] = {'$gte': duration}
        else:
            # Chi xuat toi da 30 ngay du lieu gan nhat
            if from_date and to_date:
                from_date = date_util.string_to_datetime(from_date)
                to_date = date_util.string_to_datetime(to_date)
                # if (to_date - from_date).days > 30:
                #     to_date = from_date + timedelta(days=30)
            elif from_date:
                from_date = date_util.string_to_datetime(from_date)
                if (request.now - from_date).days > 30:
                    to_date = from_date + timedelta(days=30)
                else:
                    to_date = request.now
            elif to_date:
                to_date = date_util.string_to_datetime(to_date)
                from_date = to_date - timedelta(days=30)
            else:
                to_date = request.now
                from_date = to_date - timedelta(days=30)
            conditions['get_time'] = {'$gte': from_date, '$lt': to_date + timedelta(days=1)}
            # conditions &= (db[table]['get_time'] >= from_date)
            # conditions &= (db[table]['get_time'] < to_date + timedelta(days=1))

        if is_exceed and table == 'data_min':
            conditions['is_exceed'] = True
            # conditions &= (db[table]['is_exceed'] == True)

        conditions2 = (db.station_indicator.id > 0)
        conditions2 &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        if station_id:
            conditions2 &= (db.station_indicator.station_id == station_id)
        if station_type:
            conditions2 &= (db.station_indicator.station_type == station_type)
        # lấy id của added_columns
        rows = db(conditions2).select(db.station_indicator.ALL)
        id_indicator_station = []
        for id in rows:
            id_indicator_station.append(id.indicator_id)
        query = (db.indicators.id.belongs(id_indicator_station))
        query &= (db.indicators.indicator.belongs(added_columns))
        indicatorOrigin = db(query).select(db.indicators.id)
        id_addcolumns = []
        for item in indicatorOrigin:
            id_addcolumns.append(item.id)

        indicator_ids = id_addcolumns
        station_ids = []
        for row in rows:
            if row.station_id not in station_ids:
                station_ids.append(row.station_id)
            # if row.indicator_id not in indicator_ids:
            #     indicator_ids.append(row.indicator_id)

        if station_ids:
            # conditions &= (db[table]['station_id'].belongs(station_ids))
            conditions['station_id'] = {'$in': station_ids}
        attrs = {'get_time': 1}
        if table != 'aqi_data_24h':
            attrs['data'] = 1
            # list_data = db(conditions).select(db[table].id, db[table].get_time, db[table].data)
        else:
            attrs['data_24h'] = 1
            # list_data = db(conditions).select(db[table].id, db[table].get_time, db[table].data_24h)

        list_data = pydb[table].find(conditions, attrs).sort('get_time', -1)

        rows = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                  db.indicators.unit)

        indicators = dict()
        for row in rows:
            indicators[u'{}'.format(row.indicator)] = {
                'min_value': 999999999,
                'min_time': '',
                'has_min': False,
                'max_value': -999999999,
                'max_time': '',
                'has_max': False,
                'total': 0,
                'qty': 0,
                'unit': row.unit,
            }

        # Duyet tung phan tu trong mang du lieu vua truy van duoc
        for item in list_data:
            if table != 'aqi_data_24h':
                data = item['data']
            else:
                data = item['data_24h']
            for key_tmp in data:
                k = u'{}'.format(key_tmp)
                if indicators.has_key(k):
                    v = data[k]
                    if v or v == 0:
                        try:
                            v = float(v)
                            indicators[k]['total'] += v
                            indicators[k]['qty'] += 1
                            if not indicators[k]['has_min'] or indicators[k]['min_value'] > v:
                                indicators[k]['has_min'] = True
                                indicators[k]['min_value'] = v
                                indicators[k]['min_time'] = item['get_time']
                            if not indicators[k]['has_max'] or indicators[k]['max_value'] < v:
                                indicators[k]['has_max'] = True
                                indicators[k]['max_value'] = v
                                indicators[k]['max_time'] = item['get_time']
                        except:
                            pass
        iTotalRecords = 0
        for k in indicators:
            iTotalRecords += 1
            row = [
                str(iTotalRecords),
                '%s(%s)' % (k, indicators[k]['unit']),
                '%0.2f' % indicators[k]['max_value'] if indicators[k]['has_max'] else '',
                indicators[k]['max_time'] if indicators[k]['has_max'] else '',
                '%0.2f' % indicators[k]['min_value'] if indicators[k]['has_min'] else '',
                indicators[k]['min_time'] if indicators[k]['has_min'] else '',
                '%0.2f' % (indicators[k]['total'] / indicators[k]['qty']) if indicators[k]['qty'] else '',
            ]
            aaData.append(row)
        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


# ################################################################################
# @decor.requires_login()
def export_excel():
    import os.path, pyexcelerate
    # get search parameters
    ######################################
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    month = request.vars.Month
    year = request.vars.Year
    data_type = request.vars.data_type
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if data_type:
        data_type = int(data_type)
    added_columns_decode = []
    for name in added_columns:
        name_decode = name.decode('utf-8')
        added_columns_decode.append(name_decode)

    duration = request.vars.duration
    is_exceed = request.vars.is_exceed
    view_type = request.vars.view_type or const.VIEW_BY['MINUTE']['value']

    if not station_id:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station for export data'), success=True)
    aaData = []  # Du lieu json se tra ve
    view_by_info = common.get_const_by_value(const.VIEW_BY, view_type)
    table = view_by_info['table']

    if data_type and (
            data_type == const.DATA_TYPE['YET_APPROVED']['value'] or data_type == const.DATA_TYPE['APPROVED'][
        'value']):
        view_by_info_adjust = common.get_const_by_value(const.VIEW_BY_ADJUST, view_type)
        table = view_by_info_adjust['table']

    conditions = {'station_id': station_id}#(db[table]['station_id'] == station_id)
    if duration:  # Duration duoc uu tien nhat so voi from/to
        duration = int(duration)
        # duration = datetime.now() - timedelta(days=duration)
        # Chi xuat toi da 7 ngay du lieu
        # duration = datetime.now() - timedelta(days=7)
        if duration == 1:
            duration = datetime.now() - timedelta(days=1)
        elif duration == 7:
            duration = datetime.now() - timedelta(days=7)
        else:
            duration = datetime.now() - timedelta(days=15)
        # conditions &= (db[table]['get_time'] >= duration)
        conditions['get_time'] = {'$gte': duration}
    else:
        # Chi xuat toi da 7 ngay du lieu
        if from_date and to_date and not year:
            from_date = date_util.string_to_datetime(from_date)
            to_date = date_util.string_to_datetime(to_date) + timedelta(days=1)
        # if (to_date - from_date).days > 7:
        #    to_date = from_date + timedelta(days=7)
        elif from_date and not year:
            from_date = date_util.string_to_datetime(from_date)
            if table != 'data_min' or table != 'data_adjust':
                to_date = from_date + timedelta(days=365)
            else:
                if (request.now - from_date).days > 7:
                    to_date = from_date + timedelta(days=30)
        elif to_date and not year:
            to_date = date_util.string_to_datetime(to_date)
            if table != 'data_min' or table != 'data_adjust':
                from_date = to_date - timedelta(days=365)
            else:
                from_date = to_date - timedelta(days=30)
        if month and year:
            month = int(month)
            year = int(year)
            from_date = datetime(year=year, month=month, day=1)
            to_date = date_util.get_first_day_next_month(from_date)
            to_date = datetime.combine(to_date, datetime.min.time())
        # conditions &= (db[table]['get_time'] >= from_date)
        # conditions &= (db[table]['get_time'] < to_date)
        conditions['get_time'] = {'$gte': from_date, '$lt': to_date}
    if is_exceed and table == 'data_min':
        conditions['is_exceed'] = True# &= (db[table]['is_exceed'] == True)
    if data_type:
        if data_type == const.DATA_TYPE['YET_APPROVED']['value']:
            if table != 'data_hour_adjust' and table != 'data_day_adjust' and table != 'data_mon_adjust':
                #conditions &= (db[table]['is_approved'] == False)
                conditions['is_approved'] = False
        elif data_type == const.DATA_TYPE['APPROVED']['value']:
            if table != 'data_hour_adjust' and table != 'data_day_adjust' and table != 'data_mon_adjust':
                # conditions &= (db[table]['is_approved'] == True)
                conditions['is_approved'] = True
    attrs = {'get_time': 1}
    if table != 'aqi_data_24h' and table != 'aqi_data_adjust_24h':
        attrs['data'] = 1
    else:
        attrs['data_24h'] = 1

    list_data = pydb[table].find(conditions, attrs).sort('get_time', 1)
    ws2 = []
    if list_data is None:
        raise HTTP(400, "Không đủ dữ liệu xuất excel, vui lòng thử lại!")
    else:
        temp_headers = []
        headers = []
        temp_headers.append('Datetime')
        for i, indicator in enumerate(added_columns):
            temp_headers.append(indicator.upper())
        headers.append(temp_headers)
        for header in headers:
            ws2.append(header)
        for i, item in enumerate(list_data):  # bo dem
            temp_datas = []
            if table == 'data_mon' or table == 'data_mon_adjust':
                list = item['get_time'].strftime(datetime_format_vn).split(' ')[1].split('/')
                list.pop(0)
                temp_datas.append(list[0] + '/' + list[1])
            else:
                temp_datas.append(item['get_time'].strftime(datetime_format_vn))
            for j, indicator in enumerate(added_columns_decode):
                item_data = 'data_24h'
                if table != 'aqi_data_24h' and table != 'aqi_data_adjust_24h':
                    item_data = 'data'
                item_data = item[item_data]
                if item_data.has_key(indicator):
                    if item_data[indicator] is None or item_data[indicator] == 'NULL':
                        temp_datas.append('-')
                    else:
                        temp_datas.append(item_data[indicator])
                else:
                    temp_datas.append('-')
            ws2.append(temp_datas)
    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace(",", " ")
    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb = pyexcelerate.Workbook()
    wb.new_sheet('Data', ws2)
    wb.save(file_path)
    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data

# ################################################################################
# @decor.requires_login()
def export_excel_min_max():
    import os.path, pyexcelerate
    # get search parameters
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)

    duration = request.vars.duration
    is_exceed = request.vars.is_exceed
    view_type = request.vars.view_type or const.VIEW_BY['MINUTE']['value']
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')

    if not station_id:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station for export data'), success=True)

    view_by_info = common.get_const_by_value(const.VIEW_BY, view_type)
    table = view_by_info['table']

    conditions = {'station_id': station_id}#(db[table]['station_id'] == station_id)
    if duration:  # Duration duoc uu tien nhat so voi from/to
        duration = int(duration)
        # duration = datetime.now() - timedelta(days=duration)
        # Chi xuat toi da 7 ngay du lieu
        # duration = datetime.now() - timedelta(days=7)
        if duration == 1:
            duration = datetime.now() - timedelta(days=1)
        elif duration == 7:
            duration = datetime.now() - timedelta(days=7)
        else:
            duration = datetime.now() - timedelta(days=15)
        # conditions &= (db[table]['get_time'] >= duration)
        conditions['get_time'] = {'$gte': duration}
    else:
        # Chi xuat toi da 7 ngay du lieu
        if from_date and to_date:
            from_date = date_util.string_to_datetime(from_date)
            to_date = date_util.string_to_datetime(to_date) + timedelta(days=1)
        # if (to_date - from_date).days > 7:
        #    to_date = from_date + timedelta(days=7)
        elif from_date:
            from_date = date_util.string_to_datetime(from_date)
            if (request.now - from_date).days > 7:
                to_date = from_date + timedelta(days=30)
            else:
                if table == 'data_mon':
                    to_date = request.now
                else:
                    to_date = request.now
        elif to_date:
            to_date = date_util.string_to_datetime(to_date)
            from_date = to_date - timedelta(days=30)
        else:
            if table == 'data_mon':
                to_date = request.now
                from_date = to_date - timedelta(days=365)
            else:
                to_date = request.now
                from_date = to_date - timedelta(days=30)

        # conditions &= (db[table]['get_time'] >= from_date)
        # conditions &= (db[table]['get_time'] <= to_date)
        conditions['get_time'] = {'$gte': from_date, '$lte': to_date}

    if is_exceed and table == 'data_min':
        # conditions &= (db[table]['is_exceed'] == True)
        conditions['is_exceed'] = True

    if data_type:
        if data_type == const.DATA_TYPE['YET_APPROVED']['value']:
            # conditions &= (db[table]['is_approved'] == False)
            conditions['is_approved'] = False
        elif data_type == const.DATA_TYPE['APPROVED']['value']:
            # conditions &= (db[table]['is_approved'] == True)
            conditions['is_approved'] = True
    attrs = {'get_time': 1}
    if table != 'aqi_data_24h' and table != 'aqi_data_adjust_24h':
        attrs['data'] = 1
        # list_data = db(conditions).select(db[table].id,
        #                                   db[table].get_time,
        #                                   db[table].data,
        #                                   orderby=db[table].get_time)
    else:
        attrs['data_24h'] = 1
        # list_data = db(conditions).select(db[table].id,
        #                                   db[table].get_time,
        #                                   db[table].data_24h,
        #                                   orderby=db[table].get_time)
    list_data = pydb[table].find(conditions, attrs).sort('get_time', 1)
    c_len = pydb[table].count(conditions)
    conditions = (db.station_indicator.station_id == station_id)
    conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])

    if c_len == 0 is None:
        raise HTTP(400, "Không đủ dữ liệu xuất excel, vui lòng thử lại!")
    else:
        #lấy id của added_columns
        indicator_ids_data = db(conditions).select(db.station_indicator.indicator_id)
        id_indicator_station = []
        for id in indicator_ids_data:
            id_indicator_station.append(id.indicator_id)
        query = (db.indicators.id.belongs(id_indicator_station))
        query &= (db.indicators.indicator.belongs(added_columns))
        indicatorOrigin = db(query).select(db.indicators.id)
        id_addcolumns = []
        for item in indicatorOrigin:
            id_addcolumns.append(item.id)

        indicator_ids = id_addcolumns
        # for row in indicator_ids_data:
        #     indicator_ids.append(row.indicator_id)
        rows = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id,
                                                                               db.indicators.indicator,
                                                                               db.indicators.unit)
        indicators = dict()
        for row in rows:
            name_str = str(row.indicator)
            name_decode = name_str.decode('utf-8')
            indicators[name_decode] = {
                'min_value': 999999999,
                'min_time': '',
                'has_min': False,
                'max_value': -999999999,
                'max_time': '',
                'has_max': False,
                'total': 0,
                'qty': 0,
                'unit': row.unit,
            }

            # Duyet tung phan tu trong mang du lieu vua truy van duoc
        data_sheet = []
        for item in list_data:
            if table != 'aqi_data_24h' and table != 'aqi_data_adjust_24h':
                data = item['data']
            else:
                data = item['data_24h']
            for k in data:
                if indicators.has_key(k):
                    v = data[k]
                    if v:
                        try:
                            v = float(v)
                            indicators[k]['total'] += v
                            indicators[k]['qty'] += 1
                            if not indicators[k]['has_min'] or indicators[k]['min_value'] > v:
                                indicators[k]['has_min'] = True
                                indicators[k]['min_value'] = v
                                indicators[k]['min_time'] = item['get_time']
                            if not indicators[k]['has_max'] or indicators[k]['max_value'] < v:
                                indicators[k]['has_max'] = True
                                indicators[k]['max_value'] = v
                                indicators[k]['max_time'] = item['get_time']
                        except:
                            pass

        iTotalRecords = 0
        # Write header

        data_sheet.append(['No.', 'Thông số', 'Giá trị Max', 'Thời gian Max', 'Giá trị Min', 'Thời gian Min',
                   'Giá trị trung bình'])
        # Write data

        for k in indicators:
            iTotalRecords += 1
            row = [
                str(iTotalRecords),
                '%s(%s)' % (k, indicators[k]['unit']),
                '%0.2f' % indicators[k]['max_value'] if indicators[k]['has_max'] else '',
                indicators[k]['max_time'] if indicators[k]['has_max'] else '',
                '%0.2f' % indicators[k]['min_value'] if indicators[k]['has_min'] else '',
                indicators[k]['min_time'] if indicators[k]['has_min'] else '',
                '%0.2f' % (indicators[k]['total'] / indicators[k]['qty']) if indicators[k]['qty'] else '',
            ]

            data_sheet.append(row)

    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace("," , " ")

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb = pyexcelerate.Workbook()
    wb.new_sheet('Data', data_sheet)
    wb.save(file_path)
    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data
# ################################################################################

# @decor.requires_login()
def export_excel_chart():
    import os.path, openpyxl
    from datetime import date
    from openpyxl import Workbook
    from openpyxl.chart import (
        LineChart,
        Reference,
    )
    from openpyxl.chart.axis import DateAxis
    # get search parameters
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    data_type = request.vars.data_type
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if data_type:
        data_type = int(data_type)
    added_columns_decode = []
    for name in added_columns:
        name_decode = name.decode('utf-8')
        added_columns_decode.append(name_decode)
    duration = request.vars.duration
    is_exceed = request.vars.is_exceed
    view_type = request.vars.view_type or const.VIEW_BY['MINUTE']['value']

    if not station_id:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station for export data'), success=True)
    aaData = []  # Du lieu json se tra ve
    view_by_info = common.get_const_by_value(const.VIEW_BY, view_type)
    table = view_by_info['table']

    if data_type and (
            data_type == const.DATA_TYPE['YET_APPROVED']['value'] or data_type == const.DATA_TYPE['APPROVED'][
        'value']):
        view_by_info_adjust = common.get_const_by_value(const.VIEW_BY_ADJUST, view_type)
        table = view_by_info_adjust['table']

    conditions = (db[table]['station_id'] == station_id)
    if duration:  # Duration duoc uu tien nhat so voi from/to
        duration = int(duration)
        if duration == 1:
            duration = datetime.now() - timedelta(days=1)
        elif duration == 7:
            duration = datetime.now() - timedelta(days=7)
        else:
            duration = datetime.now() - timedelta(days=15)
        conditions &= (db[table]['get_time'] >= duration)
    else:
        # Chi xuat toi da 7 ngay du lieu
        if from_date and to_date:
            from_date = date_util.string_to_datetime(from_date)
            to_date = date_util.string_to_datetime(to_date) + timedelta(days=1)
        elif from_date:
            from_date = date_util.string_to_datetime(from_date)
            if table != 'data_min' or table != 'data_adjust':
                to_date = from_date + timedelta(days=365)
            else:
                if (request.now - from_date).days > 7:
                    to_date = from_date + timedelta(days=30)
        elif to_date:
            to_date = date_util.string_to_datetime(to_date)
            if table != 'data_min' or table != 'data_adjust':
                from_date = to_date - timedelta(days=365)
            else:
                from_date = to_date - timedelta(days=30)
        else:
            if table == 'data_mon' or table == 'data_adjust':
                to_date = request.now
                from_date = to_date - timedelta(days=365)
            else:
                to_date = request.now
                from_date = to_date - timedelta(days=30)

        conditions &= (db[table]['get_time'] >= from_date)
        conditions &= (db[table]['get_time'] <= to_date)

    if is_exceed and table == 'data_min':
        conditions &= (db[table]['is_exceed'] == True)

    if data_type:
        if data_type == const.DATA_TYPE['YET_APPROVED']['value']:
            if table != 'data_hour_adjust' and table != 'data_day_adjust' and table != 'data_mon_adjust':
                conditions &= (db[table]['is_approved'] == False)
        elif data_type == const.DATA_TYPE['APPROVED']['value']:
            if table != 'data_hour_adjust' and table != 'data_day_adjust' and table != 'data_mon_adjust':
                conditions &= (db[table]['is_approved'] == True)
    if table != 'aqi_data_24h' and table != 'aqi_data_adjust_24h':
        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].data,
                                          orderby=db[table].get_time)
    else:
        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].data_24h,
                                          orderby=db[table].get_time)

    si_dict = dict()
    conditions = (db.station_indicator.station_id == station_id)
    conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
    rows = db(conditions).select(db.station_indicator.ALL)
    indicator_ids = []
    for row in rows:
        indicator_ids.append(row.indicator_id)
        si_dict[str(row.indicator_id)] = row.as_dict()
    indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                    db.indicators.unit)
    for i, item in enumerate(list_data):
        added_item = dict()
        if added_columns:
            for indicator in indicators:
                i_name = str(indicator.indicator)
                i_name_decode = i_name.decode('utf-8')
                v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                if v == '' or v is None:
                    added_item[i_name] = '-'
                else:
                    try:
                        v = float(v)
                        added_item[i_name] = "{0:.2f}".format(v)
                    except:
                        added_item[i_name] = '-'
    temp_headers = []
    temp_datas = []
    max_col = len(added_columns)
    added_columns.insert(0,'Date')
    datas = [added_columns]
    #datas = [['Date', 'NOx', 'SO2', 'O3', 'PM-10', 'PM-2-5', 'NO', 'NO2', 'CO', 'Barometer', 'Radiation', 'WindDir', 'Wind Spd', 'Rain', 'NH3', 'H2S']]
    for i, item in enumerate(list_data):  # bo dem
            if table == 'data_mon':
                list = item.get_time.strftime(datetime_format_vn).split(' ')[1].split('/')
                list.pop(0)
                    # ws['A' + str(2 + i)] = list[0] + '/' + list[1]
                temp_datas.append(list[0] + '/' + list[1])
            else:
                    # ws['A' + str(2 + i)] = item.get_time.strftime(datetime_format_vn)
                temp_datas.append(item.get_time.strftime(datetime_format_vn))
            for j,indicator in enumerate(added_columns_decode):
                    if table != 'aqi_data_24h' and table != 'aqi_data_adjust_24h':
                        if item.data.has_key(indicator):
                            if item.data[indicator] is None or item.data[indicator] == 'NULL':
                                    # ws[chr(ord('A') + j + 1) + str(2 + i)] = '-'
                                    temp_datas.append('-')
                            else:
                                    # ws[chr(ord('A') + j + 1) + str(2 + i)] = item.data[indicator]
                                    temp_datas.append(item.data[indicator])
                        else:
                                # ws[chr(ord('A') + j + 1) + str(2 + i)] = '-'
                                temp_datas.append('-')
                    else:
                        if item.data_24h.has_key(indicator):
                            if item.data_24h[indicator] is None or item.data_24h[indicator] == 'NULL':
                                    # ws[chr(ord('A') + j + 1) + str(2 + i)] = '-'
                                temp_datas.append('-')
                            else:
                                    # ws[chr(ord('A') + j + 1) + str(2 + i)] = item.data[indicator]
                                temp_datas.append(item.data_24h[indicator])
                        else:
                                # ws[chr(ord('A') + j + 1) + str(2 + i)] = '-'
                            temp_datas.append('-')

            datas.append(temp_datas)
            temp_datas = []
    wb = Workbook()
    ws = wb.active
    max_row = len(datas)

    for row in datas:
        ws.append(row)
    data = Reference(ws, min_col=2, min_row=1, max_col=max_col +1, max_row=max_row)
    c2 = LineChart()
    c2.style = 12
    c2.y_axis.title = "Values"
    c2.y_axis.crossAx = 500
    c2.x_axis = DateAxis(crossAx=100)
    # c2.x_axis.number_format = 'd-mmm'
    # c2.x_axis.majorTimeUnit = "days"

    c2.add_data(data, titles_from_data=True)
    dates = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    c2.set_categories(dates)

    ws.add_chart(c2, "I1")
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace(",", " ")

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data

################################################################################
# hungdx issue 44 add (co the ap dung cho cac chon tinh, vung ... fix sau
@service.json
def dropdown_content(table, filter_field, get_value_field, get_dsp_field, *args, **kwargs):
    try:
        filter_value = request.vars.filter_value
        filter_value = filter_value.split(';')
        filter_field = filter_field.split('-')
        conditions = (db[table]['id'] > 0)
        t1 = len(filter_field)
        t2 = len(filter_value)
        if t1 == t2:
            for i in range(0, t1):
                if filter_field[i] and filter_value[i] != '':
                    conditions &= (db[table][filter_field[i]] == filter_value[i])
        if current_user:
            if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
                list_station_manager = db(db.manager_stations.user_id == current_user.id).select(
                    db.manager_stations.station_id)
                station_ids = [str(item.station_id) for item in list_station_manager]
                conditions &= (db[table]['id'].belongs(station_ids))

        records = db(conditions).select(db[table][get_value_field], db[table][get_dsp_field],
                                        orderby=db[table].order_no)

        if records:
            html1 = "<option value=''>%s</option>" % T('-- Select an option --')
            html = ["<option value='%s'>%s</option>" % (item[get_value_field], item[get_dsp_field]) for item in records]
            html = html1 + ''.join(html)
        else:
            html = "<option value=''>%s</option>" % T('-- No data --')

        return dict(success=True, html=html)
    except Exception as ex:
        return dict(success=False, message=str(ex))

################################################################################
@service.json
@auth.requires(lambda: (auth.has_permission('view', 'view_log')))
def get_indicators_and_stations(*args, **kwargs):
    try:
        html = ''
        station_type = request.vars.type
        province_id = request.vars.province_id

        conditions = (db.indicators.id > 0)
        if station_type:
            conditions &= (db.indicators.indicator_type == station_type)
        rows = db(conditions).select(db.indicators.ALL)
        # Lay tram theo dieu kien
        stationIds = []
        cons = (db.stations.station_type == station_type)
        if province_id:
            cons &= (db.stations.province_id == province_id)
        # check tinhr

        stations = db(cons).select(db.stations.id)
        stationIds = [str(it.id) for it in stations]

        #lay thong so thuoc tram
        data_indicator = db(db.station_indicator.station_id.belongs(stationIds)).select(db.station_indicator.indicator_id)
        indicator_ids = [str(it.indicator_id) for it in data_indicator]
        indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.indicator, db.indicators.id)

        if indicators:
            html1 = "<option value=''>%s</option>" % T('-- Select an option --')
            html = ["<option value='%s'>%s</option>" % (item.id, item['indicator']) for item in indicators]
            html = html1 + ''.join(html)
        else:
            html = "<option value=''>%s</option>" % T('-- No data --')

        html2 = ''
        conditions = (db.stations.id > 0)
        if station_type:
            conditions &= (db.stations.station_type == station_type)
        if province_id:
            conditions &= (db.stations.province_id == province_id)
        rows = db(conditions).select(
            db.stations.id,
            db.stations.station_name,
        )
        for row in rows:
            html2 += '<option value="%(value)s" selected>%(name)s</option>' % dict(value=row.id,
                                                                                   name=row.station_name)
        return dict(success=True, html=html, html_2=html2)
    except Exception as ex:
        return dict(success=False, message=str(ex))

################################################################################
@service.json
def get_list_log_advanced(*args, **kwargs):
    try:
        indicator_id = request.vars.indicator_id
        station_id = request.vars.add_stations
        if station_id:
            station_id = station_id.split(',')
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        aaData = []
        view_type = request.vars.view_type
        data_type = request.vars.data_type
        table = 'data_min'
        if data_type == '3' :
            table = view_type
        else:
            if view_type == 'data_min':
                table = 'data_adjust'
            else:
                table = view_type + '_adjust'
        conditions = db[table]['id'] > 0
        conditions &= db[table]['station_id'].belongs(station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
        if to_date:
            conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
        data_indicator = db(db.indicators.id == indicator_id).select(db.indicators.indicator)
        indicator_name = [str(it.indicator) for it in data_indicator]
        rows = db(conditions).select(db[table]['id'],
                                     db[table]['get_time'],
                                     db[table]['station_id'],
                                     db[table]['data'],
                                     orderby=db[table]['get_time'])
        stations = db(db.stations.id.belongs(station_id)).select(db.stations.id,
                                                                 db.stations.station_name)
        station = dict()
        for i in stations:
            if not station.has_key(str(i.id)):
                station[str(i.id)] = i.station_name
        data = dict()
        for row in rows:
            time = row.get_time.strftime("%H:%M %d/%m/%Y")
            station_id = row.station_id
            if not data.has_key(time):
                try:
                    data[time] = {station_id: row['data'][indicator_name[0]]}
                except:
                    data[time] = {station_id: '-'}
            else:
                try:
                    data[time][station_id] = row['data'][indicator_name[0]]
                except:
                    data[time][station_id] = '-'
        for item in sorted(data.keys(), reverse=False):
            row = [item]
            for it in station:
                if data[item].has_key(it):
                    try:
                        value = float(data[item][it])
                    except:
                        value = '-'
                    row.append(value)
                else:
                    row.append('-')
            aaData.append(row)
        station_name = list(station.values())
        return dict(iTotalRecords=10, iTotalDisplayRecords=1, aaData=aaData, success=True, station_name= station_name)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)