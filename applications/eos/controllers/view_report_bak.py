# -*- coding: utf-8 -*-
###############################################################################
# Author : 
# Date   : 
# 
# Description : 
# 
###############################################################################
from applications.eos.modules import common
from datetime import datetime, timedelta
from w2pex import date_util
from datetime import datetime, timedelta
import collections
from operator import attrgetter


def call():
    return service()


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def index():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_exceed_indicator_1():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces)


################################################################################
# @service.json
# def get_list_report_exceed_indicator_1(*args, **kwargs):
#     try:
#         iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
#         iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
#         station_id = request.vars.station_id
#         from_date = request.vars.from_date
#         to_date = request.vars.to_date
#
#         added_columns = request.vars.added_columns or ''
#         if added_columns:
#             added_columns = added_columns.split(',')
#         if not station_id:
#             return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
#                         message=T('Select a station for viewing data'), success=True)
#         aaData = []  # Du lieu json se tra ve
#         list_data = None  # Du lieu truy van duoc
#         iTotalRecords = 0  # Tong so ban ghi
#         limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
#         table = 'data_adjust'
#
#         conditions = (db[table]['id'] > 0)
#         conditions &= (db[table]['station_id'] == station_id)
#         if from_date:
#             conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
#         if to_date:
#             conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
#         if not from_date and not to_date:
#             to_date = datetime.now()
#             from_date = to_date - timedelta(days=365)
#             conditions &= (db[table]['get_time'] >= from_date)
#             conditions &= (db[table]['get_time'] < to_date)
#         conditions &= (db[table]['is_exceed'] == True)
#         # conditions &= (db[table]['is_approved'] == True)
#
#         list_data = db(conditions).select(db[table].id,
#                                           db[table].get_time,
#                                           db[table].data,
#                                           orderby=~db[table].get_time,
#                                           limitby=limitby)
#         # Tong so ban ghi khong thuc hien phan trang
#         iTotalRecords = db(conditions).count(db[table].id)
#         # Thu tu ban ghi
#         iRow = iDisplayStart + 1
#
#         si_dict = dict()
#         conditions = (db.station_indicator.station_id == station_id)
#         conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
#         rows = db().select(db.station_indicator.ALL)
#         indicator_ids = []
#         for row in rows:
#             indicator_ids.append(row.indicator_id)
#             si_dict[str(row.indicator_id)] = row.as_dict()
#         indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
#                                                                         db.indicators.unit)
#
#         # Duyet tung phan tu trong mang du lieu vua truy van duoc
#         for i, item in enumerate(list_data):
#             added_item = dict()
#             if added_columns:
#                 for indicator in indicators:
#                     i_name = str(indicator.indicator)
#                     i_name_decode = i_name.decode('utf-8')
#                     v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
#                     if v == '' or v is None:
#                         added_item[i_name] = '-'
#                     else:
#                         try:
#                             v = float(v)
#                             added_item[i_name] = "{0:.2f}".format(v)
#                         except:
#                             added_item[i_name] = '-'
#             row = [
#                     str(iRow + i),
#                     # item.get_time.strftime('%Y-%m-%d %H:%M:%S'),
#                     item.get_time.strftime(datetime_format_vn),
#                 ]
#             for column in added_columns:
#                 if column and added_item.has_key(column):
#                     row.append(added_item[column])
#             aaData.append(row)
#
#         return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
#     except Exception as ex:
#         return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)

###############################################################################
@service.json
def get_list_report_exceed_indicator_1(*args, **kwargs):
    try:
        from datetime import datetime, timedelta
        from w2pex import date_util
        iDisplayStart = int(request.vars.iDisplayStart)
        iDisplayLength = int(request.vars.iDisplayLength)
        selected_st = request.vars.station_type
        station_id = request.vars.station_id
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        province_id = request.vars.province_id
        data_type = request.vars.data_type
        if data_type:
            data_type = int(data_type)
        data_filter_by = request.vars.data_filter_by
        if data_filter_by:
            data_filter_by = data_filter_by.split(';')
        status = request.vars.alarm_level
        s_search = request.vars.sometext
        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        aaData = []
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        if province_id:
            if not station_id:
                return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], success=True)
            table = 'data_min'
            if data_type:
                if data_type == const.DATA_TYPE['YET_APPROVED']['value']:
                    table = 'data_adjust'
                elif data_type == const.DATA_TYPE['APPROVED']['value']:
                    table = 'data_adjust'
        conditions = (db[table]['id'] > 0)
        if data_type:
            if data_type == const.DATA_TYPE['YET_APPROVED']['value']:
                conditions &= (db[table]['is_approved'] == False)
        if from_date:
            conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
        if to_date:
            conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))

        station_status = dict()
        if selected_st:
            conditions2 = (db.stations.station_type == selected_st)
            if status:
                conditions2 &= (db.stations.status == status)
            if province_id:
                conditions2 &= (db.stations.province_id == province_id)

            rows = db(conditions2).select(db.stations.id, db.stations.status)
            station_ids = ['']
            for row in rows:
                station_ids.append(str(row.id))
                station_status[str(row.id)] = row.status
            conditions &= (db[table]['station_id'].belongs(station_ids))

        # Neu chuoi tim kiem khac rong thi them dieu kien truy van voi chuoi tim kiem
        if s_search:
            conditions &= ((db[table]['station_id'].contains(s_search)) |
                           (db[table]['data'].contains(s_search)))
        if station_id:
            conditions &= (db[table]['station_id'] == station_id)
        if table == 'data_adjust':
            conditions &= (db[table]['del_flag'] != True)

        list_data = db(conditions).select(db[table]['id'],
                                          db[table]['station_id'],
                                          db[table]['get_time'],
                                          db[table]['is_exceed'],
                                          db[table]['data'],
                                          orderby=~db[table]['get_time'],
                                          limitby=limitby)
        if data_type == const.DATA_TYPE['YET_APPROVED']['value']:
            # If is 'YET_APPROVED', load data_adjust in order to compare on list
            # Todo: Toi uu de lay dung nhung ban ghi nhu list_data
            table2 = 'data_min'
            conditions2 = (db[table2]['id'] > 0)
            conditions2 &= (db[table2]['station_id'] == station_id)
            if from_date:
                conditions2 &= (db[table2]['get_time'] >= date_util.string_to_datetime(from_date))
            if to_date:
                conditions2 &= (db[table2]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
            list_data2 = db(conditions2).select(db[table2]['id'],
                                                db[table2]['station_id'],
                                                db[table2]['get_time'],
                                                db[table2]['is_exceed'],
                                                db[table2]['data'],
                                                orderby=~db[table]['get_time'])

        # Tong so ban ghi khong thuc hien phan trang
        iTotalRecords = db(conditions).count()
        iRow = iDisplayStart + 1
        res = common.get_station_dict()
        station_dict_name = res[0]
        station_dict_status = res[2]

        si_dict = dict()
        conditions = (db.station_indicator.id > 0)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        if selected_st:
            conditions &= (db.station_indicator.station_type == selected_st)
        if station_id:
            conditions &= (db.station_indicator.station_id == station_id)
        rows = db(conditions).select(db.station_indicator.ALL)
        si_ids = []  # list of indicators id
        for row in rows:
            si_dict[str(row.indicator_id)] = row.as_dict()
            si_ids.append(row.indicator_id)
        conditions = (db.indicators.id > 0)
        if selected_st:
            conditions &= (db.indicators.indicator_type == selected_st)
        if si_ids:
            conditions &= (db.indicators.id.belongs(si_ids))
        indicators = db(conditions).select(db.indicators.id, db.indicators.indicator, db.indicators.unit)

        # Duyet tung phan tu trong mang du lieu vua truy van duoc
        for i, item in enumerate(list_data):
            added_item = dict()
            if added_columns:
                for indicator in indicators:
                    i_id = str(indicator.id)
                    i_name = str(indicator.indicator)
                    i_name_decode = i_name.decode('utf-8')
                    if data_type == const.DATA_TYPE['YET_APPROVED']['value']:
                        v = ''
                        adjust_data = get_adjust_value_by_original_record(item, list_data2, i_name_decode)
                        if adjust_data != False:
                            try:
                                v = float(adjust_data)
                            except:
                                v = ''
                    else:
                        v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                    if not v:
                        html = SPAN('-')
                        if province_id:
                            if data_type == const.DATA_TYPE['ORIGINAL_DATA']['value'] or data_type == \
                                    const.DATA_TYPE['YET_APPROVED']['value']:
                                html += INPUT(_class='inline_adjust', _style="display: none;", _value=v,
                                              data=dict(table=table, indicator=i_name, id=item.id, oldValue=''))
                        added_item[i_name] = html
                    else:
                        try:
                            v = float(v)
                            if data_filter_by:
                                is_valid = False
                                for data_filter in data_filter_by:
                                    if str(const.DATA_FILTER_BY['IS_ZERO']['value']) == data_filter:
                                        if v == 0:
                                            is_valid = True
                                            break
                                    elif str(const.DATA_FILTER_BY['NEGATIVE']['value']) == data_filter:
                                        if v < 0:
                                            is_valid = True
                                            break
                                    elif str(const.DATA_FILTER_BY['OUT_OF_RANGE']['value']) == data_filter:
                                        pass
                                if not is_valid:
                                    html = SPAN(
                                        '-')  # Todo: Gia tri k theo dung dieu kien loc thi cung khong cho phep thay doi
                                    added_item[i_name] = html
                                    continue
                            # Lay thong tin cac nguong de display title khi hover len gtri
                            title = ''
                            if si_dict.has_key(i_id):
                                data = si_dict[i_id]
                                qcvn_detail_min_value = data['qcvn_detail_min_value']
                                qcvn_detail_max_value = data['qcvn_detail_max_value']

                                if qcvn_detail_min_value and qcvn_detail_max_value:
                                    # So sanh <= ... <=
                                    title = 'QCVN Min: %s' % si_dict[i_id]['qcvn_detail_min_value']
                                    title = title + ' - QCVN Max: %s' % si_dict[i_id]['qcvn_detail_max_value']
                                else:
                                    if qcvn_detail_max_value:
                                        title = 'QCVN Max: %s' % si_dict[i_id]['qcvn_detail_max_value']
                                    elif qcvn_detail_min_value:
                                        title = 'QCVN Min: %s' % si_dict[i_id]['qcvn_detail_min_value']

                            c = common.getColorByIndicatorQcvn(si_dict, i_id, v)
                            html = SPAN("{0:.2f}".format(v), _style="background:%s; color:white" % c, _class="badge",
                                        _title=title)
                            if province_id:
                                if data_type == const.DATA_TYPE['ORIGINAL_DATA']['value'] or data_type == \
                                        const.DATA_TYPE['YET_APPROVED']['value']:
                                    html += INPUT(_class='inline_adjust', _style="display: none;", _value=v,
                                                  data=dict(table=table, indicator=i_name, id=item.id, oldValue=v))
                            # adjust_data
                            if data_type == const.DATA_TYPE['YET_APPROVED']['value']:
                                v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                                if v:
                                    try:
                                        v = float(v)
                                        html += SPAN('({0:.2f})'.format(v))
                                    except:
                                        pass
                            added_item[i_name] = html
                        except:
                            html = SPAN('-')
                            if province_id:
                                if data_type == const.DATA_TYPE['ORIGINAL_DATA']['value'] or data_type == \
                                        const.DATA_TYPE['YET_APPROVED']['value']:
                                    html += INPUT(_class='inline_adjust', _style="display: none;", _value=v,
                                                  data=dict(table=table, indicator=i_name, id=item.id, oldValue=''))
                            added_item[i_name] = html

            # Get status info for item
            status_info = const.STATION_STATUS['GOOD']
            for ss in const.STATION_STATUS:
                if station_status.has_key(str(item.station_id)):
                    if station_status[str(item.station_id)] == const.STATION_STATUS[ss]['value']:
                        status_info = const.STATION_STATUS[ss]
            # hungdx comment 28/5/2019
            # if item.get_time > request.now:
            #     item.get_time = request.now
            row = [
                str(iRow + i),
            ]
            if province_id:
                pass
            else:
                row.append(A(I(_class=status_info['icon'], _style='color: %s' % (status_info['color'])),
                             _href=URL('detail', args=[item.station_id])))
                row.append(
                    A(I(_class='fa fa-bar-chart-o text-info'), _href=URL('detail_graph', args=[item.station_id])))
                row.append(A(I(_class='fa fa-video-camera text-success'),
                             _href=URL('camera_links', 'index', args=[item.station_id])))
                row.append(
                    A(I(_class='fa fa-map-marker text-danger'), _href=URL('station', 'map', args=[item.station_id])))
            # row.append(item.get_time.strftime('%Y-%m-%d %H-%M'))
            row.append(item.get_time.strftime(datetime_format_vn))
            idx_column = 0
            for column in added_columns:
                if column and added_item.has_key(column):
                    if data_type == const.DATA_TYPE['YET_APPROVED']['value']:
                        idx_column += 1
                        row.append(added_item[column] + INPUT(_type='checkbox', _class='column_item row_item', _row=i,
                                                              _column=idx_column, _group=0, _value=column))
                    else:
                        row.append(added_item[column])
            aaData.append(row)

        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)

#####################################################################################################



@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_exceed_times_2():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces)


################################################################################
@service.json
def get_list_report_exceed_ratio_0(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_type = request.vars.station_type
        station_id = request.vars.station_id
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        province_id = request.vars.province_id
        data_type = request.vars.data_type
        table = 'data_min'
        if data_type:
            if data_type == const.DATA_TYPE['YET_APPROVED']['value']:
                    table = 'data_adjust'
            elif data_type == const.DATA_TYPE['APPROVED']['value']:
                    table = 'data_adjust'
        if province_id:
            if not station_id:
                return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], success=True)
        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')

        conditions = (db[table]['id'] > 0)
        if from_date:
            conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
        if to_date:
            conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)

        conditions &= (db[table]['station_id'] == station_id)
        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)

        rows = db(conditions).select(db[table]['get_time'], db[table]['data'])
        dict_data = {}
        for row in rows:
            key = row['get_time'].strftime('%Y-%m-%d')
            if not dict_data.get(key): dict_data[key] = {'data': []}
            dict_data[key]['data'].append(row.data)
        iTotalRecords = len(dict_data)

        stt = 0

        iRow = iDisplayStart + 1
        indicators = common.get_indicator_dict()
        conditions = (db.station_indicator.station_id == station_id)
        rows = db(conditions).select(db.station_indicator.ALL)
        qcvn_dict = dict()
        for row in rows:
            indicator_name = indicators[row.indicator_id]
            qcvn_dict[indicator_name] = row.as_dict()

        b = dict_data.keys()
        b.sort(reverse=True)
        if b:
            for i in range(iDisplayStart, iDisplayStart + iDisplayLength):
                if i == iTotalRecords:
                    break;
                else:
                    key = b[i]
                    a = dict_data[key]['data']
                    countTotal = len(a)
                    count = dict()
                    for i in a:
                        c = i
                        for indicator_name in c:
                            v = str(c[indicator_name])
                            x = v.replace(",", "")
                            if x == 'NULL':
                                break
                            if x != '-':
                                z = float(x)
                                name_decode = indicator_name.encode('utf-8')
                                qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                                    name_decode) else None
                                qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                                    name_decode) else None
                                check_qcvn = True
                                if z:
                                    if qcvn_max or qcvn_min:
                                        check_qcvn = True
                                    if qcvn_max and z > qcvn_max:
                                        check_qcvn = False
                                    if qcvn_min and z < qcvn_min:
                                        check_qcvn = False
                                    if not check_qcvn:
                                        if not count.has_key(name_decode):
                                            count[name_decode] = 1
                                        else:
                                            count[name_decode] = int(count[name_decode]) + 1

                    row = [
                        str(iRow + stt),
                        key,
                    ]
                    stt = stt + 1
                    for column in added_columns:
                        if column and count.has_key(column):
                            m = int(count[column])
                            n = float(m) / float(countTotal)
                            n = "{0:.2f}".format(n)
                            content = '%s (%s)' % (n, '%')
                            row.append(content)
                        else:
                            content = '%s (%s)' % (0, '%')
                            row.append(content)
                    aaData.append(row)

        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


####################################################################################3
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_max_min_4():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)

    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces)


################################################################################
@service.json
def get_list_report_max_min_4(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_type = request.vars.station_type
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        province_id = request.vars.province_id
        data_type = request.vars.data_type

        conditions = (db.stations.id > 0)

        added_columns = request.vars.custom_added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        if station_type:
            conditions &= (db.stations.station_type == station_type)
        if province_id:
            conditions &= (db.stations.province_id == province_id)
        if not station_type:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station type for viewing data'), success=True)
        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        table = 'data_min'
        if data_type:
            if int(data_type) == 2:
                table = 'data_adjust'

        stations = db(conditions).select(db.stations.id,
                                         db.stations.station_name,
                                         db.stations.station_type,
                                         db.stations.province_id,
                                         limitby=limitby)
        query = db.stations.station_type == station_type
        query &= db.stations.province_id == province_id
        iTotalRecords = db(query).count()
        iRow = iDisplayStart + 1
        indicators = common.get_indicator_dict()
        for c, station in enumerate(stations):
            station_id = station.id
            conditions = (db[table]['id'] > 0)
            conditions &= (db[table]['station_id'] == station_id)
            if from_date:
                conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
            if to_date:
                conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
            if not from_date and not to_date:
                to_date = datetime.now()
                from_date = to_date - timedelta(days=365)
                conditions &= (db[table]['get_time'] >= from_date)
                conditions &= (db[table]['get_time'] < to_date)
            # conditions &= (db[table]['is_exceed'] == True)

            list_data = db(conditions).select(db[table].id,
                                              db[table].get_time,
                                              db[table].station_id,
                                              db[table].data,
                                              )
            # Tong so ban ghi khong thuc hien phan trang
            conditions = (db.station_indicator.station_id == station_id)
            conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
            rows = db(conditions).select(db.station_indicator.ALL)
            qcvn_dict = dict()
            for row in rows:
                indicator_name = indicators[row.indicator_id]
                qcvn_dict[indicator_name] = row.as_dict()
            max = dict()
            min = dict()
            total = dict()
            count = dict()
            # for in
            for i, item in enumerate(list_data):
                if item.data:
                    for indicator_name in item.data:
                        z = str(item.data[indicator_name])
                        if z == 'NULL' or z == '-' or z == 'None':
                            break
                        else:
                            x = z.replace(",", "")
                            v = float(x)
                        name_decode = indicator_name.encode('utf-8')
                        if v:
                            # max
                            if not max.has_key(name_decode):
                                max[name_decode] = v
                            else:
                                if v > max[name_decode]:
                                    max[name_decode] = v
                            # min
                            if not min.has_key(name_decode):
                                min[name_decode] = v
                            else:
                                if v < min[name_decode]:
                                    min[name_decode] = v
                            # trung binh
                            if not total.has_key(name_decode):
                                total[name_decode] = v
                                count[name_decode] = 1
                            else:
                                total[name_decode] += v
                                count[name_decode] += 1;
            row = [
                str(iRow + c),
                station.station_name,
            ]
            for column in added_columns:
                if column and count.has_key(column):
                    row.append("{0:.2f}".format(max[column]))
                    row.append("{0:.2f}".format(min[column]))
                    row.append("{0:.2f}".format(total[column] / count[column]))
                else:
                    row.append('-')
                    row.append('-')
                    row.append('-')
            aaData.append(row)
        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


###############################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_exceed_ratio_0():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces)


##############################################################################
@service.json
def get_list_report_exceed_times_2(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_type = request.vars.station_type
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        province_id = request.vars.province_id
        type = request.vars.type
        data_type = request.vars.data_type

        conditions = (db.stations.id > 0)

        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        if station_type:
            conditions &= (db.stations.station_type == station_type)
        if province_id:
            conditions &= (db.stations.province_id == province_id)
        if not station_type:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station type for viewing data'), success=True)
        table = 'data_min'
        if data_type:
            if int(data_type) == 2:
                table = 'data_adjust'
        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)

        stations = db(conditions).select(db.stations.id,
                                         db.stations.station_name,
                                         db.stations.station_type,
                                         db.stations.province_id,
                                         limitby=limitby)

        iTotalRecords = len(stations)
        iRow = iDisplayStart + 1
        indicators = common.get_indicator_dict()
        for c, station in enumerate(stations):
            station_id = station.id
            conditions = (db[table]['id'] > 0)
            conditions &= (db[table]['station_id'] == station_id)
            if from_date:
                conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
            if to_date:
                conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
            if not from_date and not to_date:
                to_date = datetime.now()
                from_date = to_date - timedelta(days=365)
                conditions &= (db[table]['get_time'] >= from_date)
                conditions &= (db[table]['get_time'] < to_date)
            # conditions &= (db[table]['is_exceed'] == True)

            list_data = db(conditions).select(db[table].id,
                                              db[table].get_time,
                                              db[table].station_id,
                                              db[table].data,
                                              )
            # Tong so ban ghi khong thuc hien phan trang
            conditions = (db.station_indicator.station_id == station_id)
            conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
            rows = db(conditions).select(db.station_indicator.ALL)
            qcvn_dict = dict()
            for row in rows:
                indicator_name = indicators[row.indicator_id]
                qcvn_dict[indicator_name] = row.as_dict()
            count = dict()
            # for in
            for i, item in enumerate(list_data):
                if item.data:
                    for indicator_name in item.data:
                        v = str(item.data[indicator_name])
                        x = v.replace(",", "")
                        if x == 'None' or x == 'NULL' or x == '-':
                            break
                        else:
                            z = float(x)
                            name_decode = indicator_name.encode('utf-8')
                            qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                                name_decode) else None
                            qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                                name_decode) else None
                            check_qcvn = True
                            if z or z == 0:
                                if qcvn_max or qcvn_min:
                                    check_qcvn = True
                                if qcvn_max and z > qcvn_max:
                                    check_qcvn = False
                                if qcvn_min and z < qcvn_min:
                                    check_qcvn = False
                                if not check_qcvn:
                                    if not count.has_key(name_decode):
                                        count[name_decode] = 1
                                    else:
                                        count[name_decode] = int(count[name_decode]) + 1
            row = [
                str(iRow + c),
                station.station_name,
            ]
            for column in added_columns:
                if column and count.has_key(column):
                    row.append(count[column])
                else:
                    row.append(0)
            aaData.append(row)
        return dict(iTotalRecords=len(aaData), iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


###############################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_data_time_3():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces)


###############################################################################
@service.json
def get_list_report_data_time_3(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_type = request.vars.station_type
        province_id = request.vars.province_id
        from_date = request.vars.from_date
        data_type = request.vars.data_type
        if from_date:
            from_date = date_util.string_to_datetime(from_date)
        to_date = request.vars.to_date
        if to_date:
            to_date = date_util.string_to_datetime(to_date)
        added_stations = request.vars.added_stations or ''
        if added_stations:
            added_stations = added_stations.split(',')
        if not station_type:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station type for viewing data'), success=True)
        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        table = 'data_min'
        if data_type:
            if int(data_type) == 2:
                table = 'data_adjust'
        conditions = db.stations.station_type == station_type
        if added_stations:
            conditions &= db.stations.id.belongs(added_stations)
        stations = db(conditions).select(
            db.stations.id,
            db.stations.station_name,
            db.stations.frequency_receiving_data,
            limitby=limitby
        )
        x = db.stations.province_id == province_id
        x &= db.stations.station_type == station_type
        iTotalRecords = db(x).count()
        iRow = iDisplayStart + 1
        for c, station in enumerate(stations):
            station_id = station.id
            delta = to_date - from_date
            freq = station.frequency_receiving_data if station.frequency_receiving_data else 5
            total_data = int((delta.days +1 ) * 24 * 60 / freq)
            conditions = (db[table]['id'] > 0)
            conditions &= (db[table]['station_id'] == station_id)
            if from_date:
                conditions &= (db[table]['get_time'] >= from_date)
            if to_date:
                conditions &= (db[table]['get_time'] < to_date + timedelta(days=1))
            if not from_date and not to_date:
                to_date = datetime.now()
                from_date = to_date - timedelta(days=365)
                conditions &= (db[table]['get_time'] >= from_date)
                conditions &= (db[table]['get_time'] < to_date)
            list_data = len(db(conditions).select(
                                              db[table].id,
                                              ))
            row = [
                str(iRow + c),
                station.station_name,
            ]
            content = '%s %s' % (list_data, T('data'))
            row.append(content)
            v = float(list_data) / float(total_data) * 100
            v = "{0:.2f}".format(v)
            content = '%s (%s)' % (v, '%')
            row.append(content)

            aaData.append(row)
        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_data_info_5():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces)


################################################################################
@service.json
def get_list_report_data_info_5(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_type = request.vars.station_type
        from_date = request.vars.from_date
        if from_date:
            from_date = date_util.string_to_datetime(from_date)
        to_date = request.vars.to_date
        if to_date:
            to_date = date_util.string_to_datetime(to_date)
        view_type = request.vars.view_type
        added_stations = request.vars.added_stations or ''
        if added_stations:
            added_stations = added_stations.split(',')
        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        if not station_type:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station type for viewing data'), success=True)
        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        if view_type == '1':
            table = 'data_min'
        else:
            table = 'data_adjust'
        conditions = db.stations.station_type == station_type
        if added_stations:
            conditions &= db.stations.id.belongs(added_stations)
        stations = db(conditions).select(
            db.stations.id,
            db.stations.station_name,
            db.stations.frequency_receiving_data,
            limitby=limitby
        )

        iTotalRecords = len(stations)
        iRow = iDisplayStart + 1
        indicators = common.get_indicator_dict()
        for c, station in enumerate(stations):
            station_id = station.id
            delta = to_date - from_date
            freq = station.frequency_receiving_data if station.frequency_receiving_data else 5
            total_data = int((delta.days +1 )* 24 * 60 / freq)
            # station_ids = [str(item.id) for item in stations
            conditions = (db[table]['id'] > 0)
            conditions &= (db[table]['station_id'] == station_id)
            if from_date:
                conditions &= (db[table]['get_time'] >= from_date)
            if to_date:
                conditions &= (db[table]['get_time'] < to_date + timedelta(days=1))
            if not from_date and not to_date:
                to_date = datetime.now()
                from_date = to_date - timedelta(days=365)
                conditions &= (db[table]['get_time'] >= from_date)
                conditions &= (db[table]['get_time'] < to_date)
            # conditions &= (db[table]['is_exceed'] == True)
            # conditions &= (db[table]['is_approved'] == True)

            list_data = db(conditions).select(db[table].id,
                                              db[table].get_time,
                                              db[table].station_id,
                                              db[table].data,
                                              )
            # Tong so ban ghi khong thuc hien phan trang
            count = dict()
            # for in
            for i, item in enumerate(list_data):
                if item.data:
                    for indicator_name in item.data:
                        try:
                            z = str(item.data[indicator_name])
                            if z == 'NULL' or z == 'None' or z == '-':
                                break
                            else:
                                x = z.replace(",", "")
                                v = float(x)
                            name_decode = indicator_name.encode('utf-8')
                            if not count.has_key(name_decode):
                                count[name_decode] = 1
                            else:
                                count[name_decode] = int(count[name_decode]) + 1
                        except:
                            pass
            row = [
                str(iRow + c),
                station.station_name,
            ]
            content = '%s %s' % (total_data, T('data'))
            row.append(content)
            if view_type == '1':
                content = '%s (%s)' % (T('Percent Data Recieved'), '%')
                row.append(content)
            else:
                content = '%s (%s)' % (T('Percent Data Adjusted'), '%')
                row.append(content)
            for column in added_columns:
                if column and count.has_key(column):
                    try:
                        v = float(count[column]) / float(total_data) * 100
                        v = "{0:.2f}".format(v)
                        row.append(v)
                    except:
                        row.append(0)
                else:
                    row.append(0)
            aaData.append(row)
        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_data_hour_6():
    station_id = request.vars.station_id
    provinces = common.get_province_have_station()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, station_id=station_id, default_provinces=default_provinces)


################################################################################
@service.json
def get_list_report_data_hour_6(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_id = request.vars.station_id
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        data_type = request.vars.data_type
        if data_type:
            data_type = int(data_type)

        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        if data_type == 2:
            table = 'data_hour_adjust'
        else:
            table = 'data_hour'

        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
        if to_date:
            conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)

        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].data,
                                          orderby=~db[table].get_time,
                                          limitby=limitby)
        # Tong so ban ghi khong thuc hien phan trang
        iTotalRecords = db(conditions).count(db[table].id)
        # Thu tu ban ghi
        iRow = iDisplayStart + 1

        si_dict = dict()
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db().select(db.station_indicator.ALL)
        indicator_ids = []
        for row in rows:
            indicator_ids.append(row.indicator_id)
            si_dict[str(row.indicator_id)] = row.as_dict()
        indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                        db.indicators.unit)

        # Duyet tung phan tu trong mang du lieu vua truy van duoc
        for i, item in enumerate(list_data):
            added_item = dict()
            if added_columns:
                for indicator in indicators:
                    i_name = str(indicator.indicator)
                    i_name_decode = i_name.decode('utf-8')
                    v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                    if v == '' or v is None:
                        added_item[i_name] = '-'
                    else:
                        try:
                            v = float(v)
                            added_item[i_name] = "{0:.2f}".format(v)
                        except:
                            added_item[i_name] = '-'
            row = [
                str(iRow + i),
                # item.get_time.strftime('%Y-%m-%d %H:%M:%S'),
                item.get_time.strftime(datetime_format_vn),
            ]
            for column in added_columns:
                if column and added_item.has_key(column):
                    row.append(added_item[column])
            list_indicator = []
            # for i in added_columns:
            #             #     indicator = i
            #             #     for indicator in list:

            aaData.append(row)

        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_data_hour_max_7():
    import datetime
    now = datetime.datetime.now()
    # print now.year
    station_id = request.vars.station_id
    provinces = common.get_province_have_station()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, station_id=station_id, default_provinces=default_provinces,
                year=now.year, month=now.month)


################################################################################
@service.json
def get_list_report_data_hour_max_7(*args, **kwargs):
    from applications.eos.modules.w2pex import date_util
    from calendar import monthrange
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_id = request.vars.station_id
        from_date = request.vars.from_date
        data_type = request.vars.data_type
        if data_type:
            data_type = int(data_type)
        to_date = request.vars.to_date
        month = request.vars.Month
        year = request.vars.Year
        if month and year:
            month = int(month)
            year = int(year)
            from_date = datetime(year=year, month=month, day=1)
            to_date = date_util.get_first_day_next_month(from_date)
            to_date = datetime.combine(to_date, datetime.min.time())

        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        if data_type == 2:
            table = 'data_hour_adjust'
        else:
            table = 'data_hour'
        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= from_date)
        if to_date:
            conditions &= (db[table]['get_time'] < to_date)

        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].data,
                                          orderby=~db[table].get_time)
        # limitby=limitby
        # Tong so ban ghi khong thuc hien phan trang
        iTotalRecords = monthrange(year, month)[1]
        # Thu tu ban ghi
        iRow = iDisplayStart + 1

        si_dict = dict()
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db().select(db.station_indicator.ALL)
        indicator_ids = []
        for row in rows:
            indicator_ids.append(row.indicator_id)
            si_dict[str(row.indicator_id)] = row.as_dict()
        indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                        db.indicators.unit)


        for i_count in range(iDisplayStart + 1, iDisplayLength + 1):
            if i_count <= iTotalRecords:
                added_item = dict()
                for i, item in enumerate(list_data):
                    if item.get_time.day == i_count:
                        if added_columns:
                            for indicator in indicators:
                                i_name = str(indicator.indicator)
                                i_name_decode = i_name.decode('utf-8')
                                v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                                if v == '' or v is None:
                                    pass
                                else:
                                    try:
                                        v = float(v)
                                        if not added_item.has_key(i_name):
                                            added_item[i_name] = "{0:.2f}".format(v)
                                        else:
                                            if v > float(added_item[i_name]):
                                                added_item[i_name] = "{0:.2f}".format(v)
                                    except:
                                        pass
                row = [
                    str(i_count),
                ]


                for column in added_columns:
                    if column and added_item.has_key(column):
                        row.append(added_item[column])
                    else:
                        row.append('-')
                aaData.append(row)
        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_data_day_8():
    import datetime
    now = datetime.datetime.now()
    station_id = request.vars.station_id
    provinces = common.get_province_have_station()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, station_id=station_id, default_provinces=default_provinces,
                year=now.year, month=now.month
                )


################################################################################
@service.json
def get_list_report_data_day_8(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_id = request.vars.station_id
        from_date = request.vars.from_date
        data_type = request.vars.data_type
        if data_type:
            data_type = int(data_type)
        to_date = request.vars.to_date
        month = request.vars.Month
        year = request.vars.Year
        if month and year:
            month = int(month)
            year = int(year)
            from_date = datetime(year=year, month=month, day=1)
            to_date = date_util.get_first_day_next_month(from_date)
            to_date = datetime.combine(to_date, datetime.min.time())
        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        if data_type == 2:
            table = 'data_day_adjust'
        else:
            table = 'data_day'

        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= from_date)
        if to_date:
            conditions &= (db[table]['get_time'] < to_date)
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)

        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].data,
                                          orderby=db[table].get_time,
                                          limitby=limitby)
        # Tong so ban ghi khong thuc hien phan trang
        iTotalRecords = db(conditions).count(db[table].id)
        # Thu tu ban ghi
        iRow = iDisplayStart + 1

        si_dict = dict()
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db().select(db.station_indicator.ALL)
        indicator_ids = []
        for row in rows:
            indicator_ids.append(row.indicator_id)
            si_dict[str(row.indicator_id)] = row.as_dict()
        indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                        db.indicators.unit)

        # Duyet tung phan tu trong mang du lieu vua truy van duoc
        for i, item in enumerate(list_data):
            added_item = dict()
            if added_columns:
                for indicator in indicators:
                    i_name = str(indicator.indicator)
                    i_name_decode = i_name.decode('utf-8')
                    v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                    if v == '' or v is None:
                        added_item[i_name] = '-'
                    else:
                        try:
                            v = float(v)
                            added_item[i_name] = "{0:.2f}".format(v)
                        except:
                            added_item[i_name] = '-'
            row = [
                str(iRow + i),
                item.get_time.strftime(datetime_format_vn),
            ]
            for column in added_columns:
                if column and added_item.has_key(column):
                    row.append(added_item[column])
            aaData.append(row)

        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_data_hour_8h_max_9():
    import datetime
    now = datetime.datetime.now()
    # print now.year
    station_id = request.vars.station_id
    provinces = common.get_province_have_station()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, station_id=station_id, default_provinces=default_provinces,
                year=now.year, month=now.month)


################################################################################
@service.json
def get_list_report_data_hour_8h_max_9(*args, **kwargs):
    from applications.eos.modules.w2pex import date_util
    from calendar import monthrange
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_id = request.vars.station_id
        data_type = request.vars.data_type
        if data_type:
            data_type = int(data_type)
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        month = request.vars.Month
        year = request.vars.Year
        if month and year:
            month = int(month)
            year = int(year)
            from_date = datetime(year=year, month=month, day=1)
            to_date = date_util.get_first_day_next_month(from_date)
            to_date = datetime.combine(to_date, datetime.min.time())

        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        if data_type == 2:
            table = 'data_hour_8h_adjust'
        else:
            table = 'data_hour_8h'

        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= from_date)
        if to_date:
            conditions &= (db[table]['get_time'] < to_date)

        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].data,
                                          orderby=~db[table].get_time)
        # Tong so ban ghi khong thuc hien phan trang
        iTotalRecords = monthrange(year, month)[1]
        # Thu tu ban ghi
        iRow = iDisplayStart + 1

        si_dict = dict()
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db().select(db.station_indicator.ALL)
        indicator_ids = []
        for row in rows:
            indicator_ids.append(row.indicator_id)
            si_dict[str(row.indicator_id)] = row.as_dict()
        indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                        db.indicators.unit)

        for i_count in range(iDisplayStart + 1, iDisplayLength + 1):
            if i_count <= iTotalRecords:
                added_item = dict()
                for i, item in enumerate(list_data):
                    if item.get_time.day == i_count:
                        if added_columns:
                            for indicator in indicators:
                                i_name = str(indicator.indicator)
                                i_name_decode = i_name.decode('utf-8')
                                v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                                if v == '' or v is None:
                                    pass
                                else:
                                    try:
                                        v = float(v)
                                        if not added_item.has_key(i_name):
                                            added_item[i_name] = "{0:.2f}".format(v)
                                        else:
                                            if v > float(added_item[i_name]):
                                                added_item[i_name] = "{0:.2f}".format(v)
                                    except:
                                        pass
                row = [
                    str(i_count),
                ]
                for column in added_columns:
                    if column and added_item.has_key(column):
                        row.append(added_item[column])
                    else:
                        row.append('-')
                aaData.append(row)
        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_aqi_10():
    import datetime
    now = datetime.datetime.now()
    station_id = request.vars.station_id
    provinces = common.get_province_have_station()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, station_id=station_id, default_provinces=default_provinces,
                year=now.year, month=now.month)


################################################################################
@service.json
def get_list_report_aqi_10(*args, **kwargs):
    from applications.eos.modules.w2pex import date_util
    from calendar import monthrange
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_id = request.vars.station_id
        data_type = request.vars.data_type
        if data_type:
            data_type = int(data_type)
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        month = request.vars.Month
        year = request.vars.Year
        if month and year:
            month = int(month)
            year = int(year)
            from_date = datetime(year=year, month=month, day=1)
            to_date = date_util.get_first_day_next_month(from_date)
            to_date = datetime.combine(to_date, datetime.min.time())

        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        aaData = []  # Du lieu json se tra ve
        # lay du lieu aqi hour
        if data_type == 2:
            table = 'aqi_data_adjust_hour'
        else:
            table = 'aqi_data_hour'
        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= from_date)
        if to_date:
            conditions &= (db[table]['get_time'] < to_date)
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)

        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].data,
                                          orderby=~db[table].get_time)
        # Tong so ban ghi khong thuc hien phan trang

        aqi_hour = dict()
        # Duyet tung phan tu trong mang du lieu vua truy van duoc
        for i, item in enumerate(list_data):
            day = item.get_time.day
            hour = item.get_time.hour
            key = (day, hour)
            v = item.data['aqi'] if item.data.has_key('aqi') else ''
            if v == '' or v is None:
                aqi_hour[key] = '-'
            else:
                try:
                    v = float(v)
                    aqi_hour[key] = "{0:.2f}".format(v)
                except:
                    aqi_hour[key] = '-'
        #####################################################
        # lay du lieu aqi day
        table = 'aqi_data_adjust_24h'
        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= from_date)
        if to_date:
            conditions &= (db[table]['get_time'] < to_date)
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)

        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].data_24h,
                                          orderby=~db[table].get_time)
        # Tong so ban ghi khong thuc hien phan trang

        aqi_day = dict()
        # Duyet tung phan tu trong mang du lieu vua truy van duoc
        for i, item in enumerate(list_data):
            day = item.get_time.day
            v = item.data_24h['aqi'] if item.data_24h.has_key('aqi') else ''
            # v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
            if v == '' or v is None:
                aqi_day[day] = '-'
            else:
                try:
                    v = float(v)
                    aqi_day[day] = "{0:.2f}".format(v)
                except:
                    aqi_day[day] = '-'

        #####################################################
        iTotalRecords = iTotalRecords = monthrange(year, month)[1]
        # Thu tu ban ghi
        iRow = iDisplayStart + 1
        for i_count in range(iDisplayStart + 1, iDisplayLength + 1):
            if i_count <= iTotalRecords:
                row = [
                    str(i_count),
                ]
                for j in range(0, 24):
                    key = (i_count, j)
                    if aqi_hour.has_key(key):
                        row.append(aqi_hour[key])
                    else:
                        row.append('-')
                if aqi_day.has_key(i_count):
                    row.append(aqi_day[i_count])
                else:
                    row.append('-')
                aaData.append(row)

        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_data_month_11():
    import datetime
    now = datetime.datetime.now()
    station_id = request.vars.station_id
    provinces = common.get_province_have_station()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, station_id=station_id, default_provinces=default_provinces,
                year=now.year, month=now.month)


################################################################################
@service.json
def get_list_report_data_month_11(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_id = request.vars.station_id
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        data_type = request.vars.data_type
        if data_type:
            data_type = int(data_type)
        # month = request.vars.Month
        year = request.vars.Year
        # if month and year:
        if year:
            # month = int(month)
            year = int(year)
            from_date = datetime(year=year, month=1, day=1)
            to_date = datetime(year=year + 1, month=1, day=1)
        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        # iTotalRecords = 0  # Tong so ban ghi
        # limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        if data_type == 2:
            table = 'data_mon_adjust'
        else:
            table = 'data_mon'

        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= from_date)
        if to_date:
            conditions &= (db[table]['get_time'] < to_date)
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)

        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].data,
                                          orderby=db[table].get_time)
        # Tong so ban ghi khong thuc hien phan trang
        iTotalRecords = 1
        # Thu tu ban ghi
        iRow = iDisplayStart + 1

        si_dict = dict()
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db().select(db.station_indicator.ALL)
        indicator_ids = []
        for row in rows:
            indicator_ids.append(row.indicator_id)
            si_dict[str(row.indicator_id)] = row.as_dict()
        indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                        db.indicators.unit)
        result_dict = dict()
        total_result = dict()
        count_result = dict()
        # Duyet tung phan tu trong mang du lieu vua truy van duoc
        for i, item in enumerate(list_data):
            added_item = dict()
            if added_columns:
                for indicator in indicators:
                    i_name = str(indicator.indicator)
                    i_name_decode = i_name.decode('utf-8')
                    v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                    if v == '' or v is None:
                        added_item[i_name] = '-'
                    else:
                        try:
                            v = float(v)
                            added_item[i_name] = "{0:.2f}".format(v)
                            if total_result.has_key(i_name):
                                total_result[i_name] += v
                                count_result[i_name] += 1
                            else:
                                total_result[i_name] = v
                                count_result[i_name] = 1
                        except:
                            added_item[i_name] = '-'
            row = [
                item.get_time.month
            ]
            for column in added_columns:
                if column and added_item.has_key(column):
                    row.append(added_item[column])
            result_dict[item.get_time.month] = row
        for i in range(1, 13):
            if not result_dict.has_key(i):
                row = [
                    i,
                ]
                for column in added_columns:
                    row.append('-')
                aaData.append(row)
            else:
                aaData.append(result_dict[i])
        row = [
            T('AVG Year'),
        ]
        for column in added_columns:
            if total_result.has_key(column):
                avg = float(total_result[column]) / count_result[column]
                row.append("{0:.2f}".format(avg))
            else:
                row.append('-')
        aaData.append(row)
        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)


################################################################################
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def report_lost_indicator_12():
    view_type = request.vars.view_type
    try:
        view_type = int(view_type)
    except Exception as es:
        view_type = const.VIEW_BY['MINUTE']['value'];
    station_id = request.vars.station_id
    provinces = common.get_province_have_station()
    # stations = db(db.stations.id > 0).select() #hungdx comment issue 44
    # hungdx phan quyen quan ly tráº¡m theo user issue 44
    conditions = (db.stations.id > 0)
    if current_user:
        if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
            list_station_manager = db(db.manager_stations.user_id == current_user.id). \
                select(db.manager_stations.station_id)
            station_ids = [str(item.station_id) for item in list_station_manager]
            conditions &= (db.stations.id.belongs(station_ids))
    stations = db(conditions).select()

    default_provinces = db(db.provinces.default == 1).select()
    return dict(provinces=provinces, stations=stations, view_type=view_type, station_id=station_id,
                default_provinces=default_provinces)


################################################################################
@service.json
def get_list_report_lost_indicator_12(*args, **kwargs):
    try:
        iDisplayStart = int(request.vars.iDisplayStart)  # Vi tri bat dau lay du lieu
        iDisplayLength = int(request.vars.iDisplayLength)  # So luong ban ghi se lay toi da
        station_id = request.vars.station_id
        from_date = request.vars.from_date
        to_date = request.vars.to_date
        data_type = request.vars.data_type
        if data_type:
            data_type = int(data_type)

        added_columns = request.vars.added_columns or ''
        if added_columns:
            added_columns = added_columns.split(',')
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                        message=T('Select a station for viewing data'), success=True)
        aaData = []  # Du lieu json se tra ve
        list_data = None  # Du lieu truy van duoc
        iTotalRecords = 0  # Tong so ban ghi
        limitby = (iDisplayStart, iDisplayLength + 1)  # Tuple dung de phan trang (vtri bat dau - chieu dai)
        table = 'data_adjust'
        if data_type == 2:
            table = 'data_adjust'
        else:
            table = 'data_min'
        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
        if to_date:
            conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)

        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].data,
                                          orderby=~db[table].get_time,
                                          limitby=limitby)
        # Tong so ban ghi khong thuc hien phan trang
        # iTotalRecords = db(conditions).count(db[table].id)
        # Thu tu ban ghi
        iRow = iDisplayStart + 1

        si_dict = dict()
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db().select(db.station_indicator.ALL)
        indicator_ids = []
        for row in rows:
            indicator_ids.append(row.indicator_id)
            si_dict[str(row.indicator_id)] = row.as_dict()
        indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                        db.indicators.unit)

        # Duyet tung phan tu trong mang du lieu vua truy van duoc
        if from_date and to_date:
            start_time = date_util.string_to_datetime(from_date)
            end_time = date_util.string_to_datetime(to_date) + timedelta(days=1)
            freq_data = db(db.stations.id == station_id).select(db.stations.frequency_receiving_data)
            frequency = 60
            if freq_data:
                itemfreq = freq_data.first()['frequency_receiving_data']
                if itemfreq:
                    frequency = round(itemfreq * 60)
            iTotalRecords = (end_time - start_time).total_seconds() / frequency
            viewable = True
            for i in range(iDisplayStart, iDisplayStart + iDisplayLength):
                time = start_time + timedelta(seconds=i * frequency)
                added_item = dict()
                for j, item in enumerate(list_data):
                    if item.get_time == time:
                        viewable = False
                        if added_columns:
                            for indicator in indicators:
                                i_name = str(indicator.indicator)
                                i_name_decode = i_name.decode('utf-8')
                                v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                                if v == '' or v is None:
                                    viewable = True
                                    added_item[i_name] = '-'
                                else:
                                    try:
                                        v = float(v)
                                        added_item[i_name] = "{0:.2f}".format(v)
                                    except:
                                        viewable = True
                                        added_item[i_name] = '-'
                row = [
                    str(i + 1),
                    time.strftime(datetime_format_vn),
                ]
                for column in added_columns:
                    if column and added_item.has_key(column):
                        row.append(added_item[column])
                    else:
                        row.append('-')
                if viewable:
                    aaData.append(row)

        return dict(iTotalRecords=iTotalRecords, iTotalDisplayRecords=iTotalRecords, aaData=aaData, success=True)
    except Exception as ex:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], message=str(ex), success=False)
################################################################################
@service.json
def export_excel_0():
    import os.path, openpyxl
    from datetime import datetime, timedelta
    from w2pex import date_util
    station_type = request.vars.station_type
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)
    province_id = request.vars.province_id
    type = request.vars.type
    
    if data_type == 2:
        table = 'data_adjust'
    else:
        table = 'data_min'
    conditions = (db[table].id > 0)
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if station_id:
        conditions &= (db[table].station_id == station_id)
    if from_date:
        conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
    if to_date:
        conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
    if not from_date and not to_date:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=365)
        conditions &= (db[table]['get_time'] >= from_date)
        conditions &= (db[table]['get_time'] < to_date)
    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc
    iTotalRecords = 0  # Tong so ban ghi
    rows = db(conditions).select(orderby=db[table].get_time)
    dict_data = {}
    for row in rows:
        key = row['get_time'].strftime('%Y-%m-%d')
        if not dict_data.get(key): dict_data[key] = {'data': []}
        dict_data[key]['data'].append(row.data)

    iTotalRecords = db(db.stations.station_type == station_type).count()
    indicators = common.get_indicator_dict()
    list_data = db(conditions).select(db[table].id,
                                      db[table].get_time,
                                      db[table].station_id,
                                      db[table].data,
                                      )
    conditions = (db.station_indicator.station_id == station_id)
    conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
    rows = db(conditions).select(db.station_indicator.ALL)
    qcvn_dict = dict()
    for row in rows:
        indicator_name = indicators[row.indicator_id]
        qcvn_dict[indicator_name] = row.as_dict()

    for key in sorted(dict_data.keys(), reverse=True):
        count = dict()
        k = key
        a = dict_data[k]['data']
        countTotal = len(a)
        for i in a:
            c = i
            for indicator_name in c:
                v = str(c[indicator_name])
                x = v.replace(",", "")
                z = float(x)
                name_decode = indicator_name.encode('utf-8')
                qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                    name_decode) else None
                qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                    name_decode) else None
                check_qcvn = True
                if z:
                    if qcvn_max or qcvn_min:
                        check_qcvn = True
                    if qcvn_max and z > qcvn_max:
                        check_qcvn = False
                    if qcvn_min and z < qcvn_min:
                        check_qcvn = False
                    if not check_qcvn:
                        if not count.has_key(name_decode):
                            count[name_decode] = 1
                        else:
                            count[name_decode] = int(count[name_decode]) + 1

        row = [
            k
        ]
        for column in added_columns:
            if column and count.has_key(column):
                m = int(count[column])
                n = float(m) / float(countTotal)
                n = "{0:.2f}".format(n)
                content = '%s (%s)' % (n, '%')
                row.append(content)
            else:
                content = '%s (%s)' % (0, '%')
                row.append(content)
        aaData.append(row)

    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()
    if list_data.first() is None:
        raise HTTP(400, "KhÃ´ng Ä‘á»§ dá»¯ liá»‡u xuáº¥t excel, vui lÃ²ng thá»­ láº¡i!")
    else:
        temp_headers = []
        headers = []
        temp_headers.append('Ngay gio')
        for item in added_columns:
            temp_headers.append(str(item))
        headers.append(temp_headers)
        for header in headers:
            ws2.append(header)
        # Write data
        for row in aaData:
            ws2.append(row)

    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace("," , " ")

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data

###############################################################################
@service.json
def export_excel_1():
    import os.path, openpyxl
    from datetime import datetime, timedelta
    from w2pex import date_util
    selected_st = request.vars.station_type
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    province_id = request.vars.province_id
    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)
    data_filter_by = request.vars.data_filter_by
    if data_filter_by:
        data_filter_by = data_filter_by.split(';')
    status = request.vars.alarm_level
    s_search = request.vars.sometext
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    aaData = []
    if data_type == 2:
        table = 'data_adjust'
    else:
        table = 'data_min'
    if province_id:
        if not station_id:
            return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[], success=True)

    conditions = (db[table]['id'] > 0)
    if data_type:
        if data_type == const.DATA_TYPE['YET_APPROVED']['value']:
            conditions &= (db[table]['is_approved'] == False)
    if from_date:
        conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
    if to_date:
        conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))

    station_status = dict()
    if selected_st:
        conditions2 = (db.stations.station_type == selected_st)
        if status:
            conditions2 &= (db.stations.status == status)
        if province_id:
            conditions2 &= (db.stations.province_id == province_id)

        rows = db(conditions2).select(db.stations.id, db.stations.status)
        station_ids = ['']
        for row in rows:
            station_ids.append(str(row.id))
            station_status[str(row.id)] = row.status
        conditions &= (db[table]['station_id'].belongs(station_ids))

    # Neu chuoi tim kiem khac rong thi them dieu kien truy van voi chuoi tim kiem
    if s_search:
        conditions &= ((db[table]['station_id'].contains(s_search)) |
                       (db[table]['data'].contains(s_search)))
    if station_id:
        conditions &= (db[table]['station_id'] == station_id)
    if table == 'data_adjust':
        conditions &= (db[table]['del_flag'] != True)

    list_data = db(conditions).select(db[table]['id'],
                                      db[table]['station_id'],
                                      db[table]['get_time'],
                                      db[table]['is_exceed'],
                                      db[table]['data'],
                                      orderby=~db[table]['get_time'], )

    # Tong so ban ghi khong thuc hien phan trang
    res = common.get_station_dict()
    si_dict = dict()
    conditions = (db.station_indicator.id > 0)
    conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
    if selected_st:
        conditions &= (db.station_indicator.station_type == selected_st)
    if station_id:
        conditions &= (db.station_indicator.station_id == station_id)
    rows = db(conditions).select(db.station_indicator.ALL)
    si_ids = []  # list of indicators id
    for row in rows:
        si_dict[str(row.indicator_id)] = row.as_dict()
        si_ids.append(row.indicator_id)
    conditions = (db.indicators.id > 0)
    if selected_st:
        conditions &= (db.indicators.indicator_type == selected_st)
    if si_ids:
        conditions &= (db.indicators.id.belongs(si_ids))
    indicators = db(conditions).select(db.indicators.id, db.indicators.indicator, db.indicators.unit)

    # Duyet tung phan tu trong mang du lieu vua truy van duoc
    for i, item in enumerate(list_data):
        added_item = dict()
        if added_columns:
            for indicator in indicators:
                i_id = str(indicator.id)
                i_name = str(indicator.indicator)
                i_name_decode = i_name.decode('utf-8')
                if data_type == const.DATA_TYPE['YET_APPROVED']['value']:
                    v = ''
                    adjust_data = get_adjust_value_by_original_record(item, list_data2, i_name_decode)
                    if adjust_data != False:
                        try:
                            v = float(adjust_data)
                        except:
                            v = ''
                else:
                    v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                if not v:
                    html = ('-')
                    if province_id:
                        if data_type == const.DATA_TYPE['ORIGINAL_DATA']['value'] or data_type == \
                                const.DATA_TYPE['YET_APPROVED']['value']:
                            html += INPUT(_class='inline_adjust', _style="display: none;", _value=v,
                                          data=dict(table=table, indicator=i_name, id=item.id, oldValue=''))
                    added_item[i_name] = html
                else:
                    try:
                        v = float(v)
                        if data_filter_by:
                            is_valid = False
                            for data_filter in data_filter_by:
                                if str(const.DATA_FILTER_BY['IS_ZERO']['value']) == data_filter:
                                    if v == 0:
                                        is_valid = True
                                        break
                                elif str(const.DATA_FILTER_BY['NEGATIVE']['value']) == data_filter:
                                    if v < 0:
                                        is_valid = True
                                        break
                                elif str(const.DATA_FILTER_BY['OUT_OF_RANGE']['value']) == data_filter:
                                    pass
                            if not is_valid:
                                html = ('-')  # Todo: Gia tri k theo dung dieu kien loc thi cung khong cho phep thay doi
                                added_item[i_name] = html
                                continue
                        # Lay thong tin cac nguong de display title khi hover len gtri
                        title = ''
                        if si_dict.has_key(i_id):
                            data = si_dict[i_id]
                            qcvn_detail_min_value = data['qcvn_detail_min_value']
                            qcvn_detail_max_value = data['qcvn_detail_max_value']

                            if qcvn_detail_min_value and qcvn_detail_max_value:
                                # So sanh <= ... <=
                                title = 'QCVN Min: %s' % si_dict[i_id]['qcvn_detail_min_value']
                                title = title + ' - QCVN Max: %s' % si_dict[i_id]['qcvn_detail_max_value']
                            else:
                                if qcvn_detail_max_value:
                                    title = 'QCVN Max: %s' % si_dict[i_id]['qcvn_detail_max_value']
                                elif qcvn_detail_min_value:
                                    title = 'QCVN Min: %s' % si_dict[i_id]['qcvn_detail_min_value']

                        c = common.getColorByIndicatorQcvn(si_dict, i_id, v)
                        html = ("{0:.2f}".format(v))
                        if province_id:
                            if data_type == const.DATA_TYPE['ORIGINAL_DATA']['value'] or data_type == \
                                    const.DATA_TYPE['YET_APPROVED']['value']:
                                html += INPUT(_class='inline_adjust', _style="display: none;", _value=v,
                                              data=dict(table=table, indicator=i_name, id=item.id, oldValue=v))
                        # adjust_data
                        if data_type == const.DATA_TYPE['YET_APPROVED']['value']:
                            v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                            if v:
                                try:
                                    v = float(v)
                                    html += ('({0:.2f})'.format(v))
                                except:
                                    pass
                        added_item[i_name] = html
                    except:
                        html = SPAN('-')
                        if province_id:
                            if data_type == const.DATA_TYPE['ORIGINAL_DATA']['value'] or data_type == \
                                    const.DATA_TYPE['YET_APPROVED']['value']:
                                html += INPUT(_class='inline_adjust', _style="display: none;", _value=v,
                                              data=dict(table=table, indicator=i_name, id=item.id, oldValue=''))
                        added_item[i_name] = html

        # Get status info for item
        status_info = const.STATION_STATUS['GOOD']
        for ss in const.STATION_STATUS:
            if station_status.has_key(str(item.station_id)):
                if station_status[str(item.station_id)] == const.STATION_STATUS[ss]['value']:
                    status_info = const.STATION_STATUS[ss]


        row = []

        row.append(item.get_time.strftime(datetime_format_vn))
        idx_column = 0
        for column in added_columns:
            if column and added_item.has_key(column):
                if data_type == const.DATA_TYPE['YET_APPROVED']['value']:
                    idx_column += 1
                    row.append(added_item[column] + INPUT(_type='checkbox', _class='column_item row_item', _row=i,
                                                          _column=idx_column, _group=0, _value=column))
                else:
                    row.append(added_item[column])
        aaData.append(row)

    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()
    if list_data.first() is None:
        raise HTTP(400, "KhÃ´ng Ä‘á»§ dá»¯ liá»‡u xuáº¥t excel, vui lÃ²ng thá»­ láº¡i!")
    else:
        temp_headers = []
        headers = []
        temp_headers.append('NgÃ y giá»')
        for item in added_columns:
            temp_headers.append(str(item))
        headers.append(temp_headers)
        for header in headers:
            ws2.append(header)
        # Write data
        for row in aaData:
            ws2.append(row)

    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace("," , " ")

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data

##############################################################################
@service.json
def export_excel_2():
    import os.path, openpyxl
    # get search parameters
    station_type = request.vars.station_type
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    province_id = request.vars.province_id
    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if not station_type:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station type for viewing data'), success=True)
    conditions = db.stations.station_type == station_type

    if province_id:
        conditions &= (db.stations.province_id == province_id)

    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc
    iTotalRecords = 0  # Tong so ban ghi
    if data_type == 2:
        table = 'data_adjust'
    else:
        table = 'data_min'
    stations = db(conditions).select(
        db.stations.id,
        db.stations.station_name,
    )

    iTotalRecords = db(db.stations.station_type == station_type).count()
    indicators = common.get_indicator_dict()
    for c, station in enumerate(stations):
        station_id = station.id
        # station_ids = [str(item.id) for item in stations
        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
        if to_date:
            conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)
        # conditions &= (db[table]['is_exceed'] == True)
        # conditions &= (db[table]['is_approved'] == True)

        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].station_id,
                                          db[table].data,
                                          )
        # Tong so ban ghi khong thuc hien phan trang
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db(conditions).select(db.station_indicator.ALL)
        qcvn_dict = dict()
        for row in rows:
            indicator_name = indicators[row.indicator_id]
            qcvn_dict[indicator_name] = row.as_dict()
        count = dict()
        # for in
        for i, item in enumerate(list_data):
            if item.data:
                for indicator_name in item.data:
                    z = str(item.data[indicator_name])
                    x = z.replace(",", "")
                    if x == 'NULL' or x == 'None' or x == '-':
                        break
                    else:
                        v = float(x)
                        name_decode = indicator_name.encode('utf-8')
                        qcvn_min = qcvn_dict[name_decode]['qcvn_detail_min_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        qcvn_max = qcvn_dict[name_decode]['qcvn_detail_max_value'] if qcvn_dict.has_key(
                            name_decode) else None
                        check_qcvn = True
                        if v or v == 0:
                            if qcvn_max or qcvn_min:
                                check_qcvn = True
                            if qcvn_max and v > qcvn_max:
                                check_qcvn = False
                            if qcvn_min and v < qcvn_min:
                                check_qcvn = False
                            if not check_qcvn:
                                if not count.has_key(name_decode):
                                    count[name_decode] = 1
                                else:
                                    count[name_decode] = int(count[name_decode]) + 1
        row = [
            station.station_name,
        ]
        for column in added_columns:
            if column and count.has_key(column):
                row.append(count[column])
            else:
                row.append(0)
        aaData.append(row)

    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()

    temp_headers = []
    headers = []
    temp_headers.append(' ')
    for item in added_columns:
        temp_headers.append(str(item))
    headers.append(temp_headers)
    for header in headers:
        ws2.append(header)
    # Write data
    for row in aaData:
        ws2.append(row)

    # Get station name

    file_name = request.now.strftime('_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


##################################################################################
@service.json
def export_excel_3():
    import os.path, openpyxl
    station_type = request.vars.station_type
    from_date = request.vars.from_date
    if from_date:
        from_date = date_util.string_to_datetime(from_date)
    to_date = request.vars.to_date
    if to_date:
        to_date = date_util.string_to_datetime(to_date)
    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)
    added_stations = request.vars.added_stations or ''
    if added_stations:
        added_stations = added_stations.split(',')
    if not station_type:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station type for viewing data'), success=True)
    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc
    iTotalRecords = 0  # Tong so ban ghi # Tuple dung de phan trang (vtri bat dau - chieu dai)
    if data_type == 2:
        table = 'data_adjust'
    else:
        table = 'data_min'
    conditions = db.stations.station_type == station_type
    if added_stations:
        conditions &= db.stations.id.belongs(added_stations)
    stations = db(conditions).select(
        db.stations.id,
        db.stations.station_name,
        db.stations.frequency_receiving_data,
    )

    iTotalRecords = db(db.stations.station_type == station_type).count()
    for c, station in enumerate(stations):
        station_id = station.id
        delta = to_date - from_date
        freq = station.frequency_receiving_data if station.frequency_receiving_data else 5
        total_data = int((delta.days+1) * 24 * 60 / freq)
        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= from_date)
        if to_date:
            conditions &= (db[table]['get_time'] < to_date + timedelta(days=1))
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)

        list_data = len(db(conditions).select(
            db[table].id,
        ))
        row = [
            str(1 + c),
            station.station_name,
        ]
        content = '%s %s' % (list_data, T('data'))
        row.append(content)
        v = float(list_data) / float(total_data) * 100
        v = "{0:.2f}".format(v)
        content = '%s (%s)' % (v, '%')
        row.append(content)

        aaData.append(row)

    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()

    temp_headers = []
    headers = []
    temp_headers.append(' ')
    temp_headers.append('Ten tram')
    temp_headers.append('Tong so du lieu nhan duoc')
    temp_headers.append('Ti le nhan duoc de lieu')
    headers.append(temp_headers)
    for header in headers:
        ws2.append(header)
    # Write data
    for row in aaData:
        ws2.append(row)

    # Get station name

    file_name = request.now.strftime('_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


# ################################################################################
@service.json
def export_excel_4():
    import os.path, openpyxl
    # get search parameters
    station_type = request.vars.station_type
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    province_id = request.vars.province_id
    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)
    added_columns = request.vars.custom_added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if not station_type:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station type for viewing data'), success=True)
    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc
    iTotalRecords = 0  # Tong so ban ghi # Tuple dung de phan trang (vtri bat dau - chieu dai)
    if data_type == 2:
        table = 'data_adjust'
    else:
        table = 'data_min'
    conditions = db.stations.station_type == station_type
    if province_id:
        conditions &= (db.stations.province_id == province_id)
    stations = db(conditions).select(
        db.stations.id,
        db.stations.station_name,
    )

    iTotalRecords = db(db.stations.station_type == station_type).count()
    iRow = 1
    indicators = common.get_indicator_dict()
    for c, station in enumerate(stations):
        station_id = station.id
        # station_ids = [str(item.id) for item in stations
        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
        if to_date:
            conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)
        # conditions &= (db[table]['is_exceed'] == True)

        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].station_id,
                                          db[table].data,
                                          )
        # Tong so ban ghi khong thuc hien phan trang
        conditions = (db.station_indicator.station_id == station_id)
        conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
        rows = db(conditions).select(db.station_indicator.ALL)
        qcvn_dict = dict()
        for row in rows:
            indicator_name = indicators[row.indicator_id]
            qcvn_dict[indicator_name] = row.as_dict()
        max = dict()
        min = dict()
        total = dict()
        count = dict()
        # for in
        for i, item in enumerate(list_data):
            if item.data:
                for indicator_name in item.data:
                    z = str(item.data[indicator_name])
                    if z == 'NULL' or z == 'None' or z == '-':
                        break
                    else:
                        x = z.replace(",", "")
                        v = float(x)
                    name_decode = indicator_name.encode('utf-8')
                    if v:
                        # max
                        if not max.has_key(name_decode):
                            max[name_decode] = v
                        else:
                            if v > max[name_decode]:
                                max[name_decode] = v
                        # min
                        if not min.has_key(name_decode):
                            min[name_decode] = v
                        else:
                            if v < min[name_decode]:
                                min[name_decode] = v
                        # trung binh
                        if not total.has_key(name_decode):
                            total[name_decode] = v
                            count[name_decode] = 1
                        else:
                            total[name_decode] += v
                            count[name_decode] += 1;
        row = [
            str(iRow + c),
            station.station_name,
        ]
        for column in added_columns:
            if column and count.has_key(column):
                row.append("{0:.2f}".format(max[column]))
                row.append("{0:.2f}".format(min[column]))
                row.append("{0:.2f}".format(total[column] / count[column]))
            else:
                row.append('-')
                row.append('-')
                row.append('-')
        aaData.append(row)

    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()
    temp_headers = []
    temp_headers_2 = []
    headers = []
    headers_2 = []
    temp_headers.append('#')
    temp_headers.append('Ten Tram')
    temp_headers_2.append(' ')
    temp_headers_2.append(' ')

    for item in added_columns:
        temp_headers.append('')
        temp_headers.append(str(item))
        temp_headers.append('')
    headers.append(temp_headers)
    for header in headers:
        ws2.append(header)

    for item in added_columns:
        temp_headers_2.append('Max')
        temp_headers_2.append('Min')
        temp_headers_2.append('TB')
    headers_2.append(temp_headers_2)
    for header in headers_2:
        ws2.append(header)
    # Write data
    for row in aaData:
        ws2.append(row)

    # Get station name

    file_name = request.now.strftime('_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


# ################################################################################
@service.json
def export_excel_5():
    import os.path, openpyxl
    # get search parameters
    station_type = request.vars.station_type
    province_id = request.vars.province_id
    from_date = request.vars.from_date
    if from_date:
        from_date = date_util.string_to_datetime(from_date)
    to_date = request.vars.to_date
    if to_date:
        to_date = date_util.string_to_datetime(to_date)
    view_type = request.vars.view_type
    added_stations = request.vars.added_stations or ''
    if added_stations:
        added_stations = added_stations.split(',')
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if not station_type:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station type for viewing data'), success=True)

    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc
    iTotalRecords = 0  # Tong so ban ghiple dung de phan trang (vtri bat dau - chieu dai)
    if view_type == '1':
        table = 'data_min'
    else:
        table = 'data_adjust'
    conditions = db.stations.station_type == station_type
    if province_id:
        conditions &= (db.stations.province_id == province_id)
    if added_stations:
        conditions &= db.stations.id.belongs(added_stations)
    stations = db(conditions).select(
        db.stations.id,
        db.stations.station_name,
        db.stations.frequency_receiving_data,
    )

    iTotalRecords = db(db.stations.station_type == station_type).count()
    iRow = 1
    indicators = common.get_indicator_dict()
    for c, station in enumerate(stations):
        station_id = station.id
        delta = to_date - from_date
        freq = station.frequency_receiving_data if station.frequency_receiving_data else 5
        total_data = int((delta.days+1) * 24 * 60 / freq)
        # station_ids = [str(item.id) for item in stations
        conditions = (db[table]['id'] > 0)
        conditions &= (db[table]['station_id'] == station_id)
        if from_date:
            conditions &= (db[table]['get_time'] >= from_date)
        if to_date:
            conditions &= (db[table]['get_time'] < to_date + timedelta(days=1))
        if not from_date and not to_date:
            to_date = datetime.now()
            from_date = to_date - timedelta(days=365)
            conditions &= (db[table]['get_time'] >= from_date)
            conditions &= (db[table]['get_time'] < to_date)

        list_data = db(conditions).select(db[table].id,
                                          db[table].get_time,
                                          db[table].station_id,
                                          db[table].data,
                                          )
        # Tong so ban ghi khong thuc hien phan trang
        count = dict()
        # for in
        for i, item in enumerate(list_data):
            if item.data:
                for indicator_name in item.data:
                    z = str(item.data[indicator_name])
                    if z == "-" or z == 'NULL' or z == 'None':
                        break
                    else:
                        x = z.replace(",", "")
                        v = float(x)
                    name_decode = indicator_name.encode('utf-8')
                    if not count.has_key(name_decode):
                        count[name_decode] = 1
                    else:
                        count[name_decode] = int(count[name_decode]) + 1
        row = [
            str(iRow + c),
            station.station_name,
        ]
        content = '%s %s' % (total_data, T('data'))
        row.append(content)
        if view_type == '1':
            content = '%s (%s)' % (T('Percent Data Recieved'), '%')
            row.append(content)
        else:
            content = '%s (%s)' % (T('Percent Data Adjusted'), '%')
            row.append(content)
        for column in added_columns:
            if column and count.has_key(column):
                try:
                    v = float(count[column]) / float(total_data) * 100
                    v = "{0:.2f}".format(v)
                    row.append(v)
                except:
                    row.append(0)
            else:
                row.append(0)
        aaData.append(row)
    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()
    temp_headers = []
    headers = []
    temp_headers.append(' ')
    temp_headers.append('Ten Tram')
    temp_headers.append('Tong so du lieu')
    temp_headers.append('Noi Dung')
    for item in added_columns:
        temp_headers.append(str(item))
    headers.append(temp_headers)
    for header in headers:
        ws2.append(header)
    # Write data
    for row in aaData:
        ws2.append(row)

    # Get station name

    file_name = request.now.strftime('_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


# ################################################################################
@service.json
def export_excel_6():
    import os.path, openpyxl
    # get search parameters
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if not station_id:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station for viewing data'), success=True)
    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc
    if data_type == 2:
        table = 'data_hour_adjust'
    else:
        table = 'data_hour'
    conditions = (db[table]['id'] > 0)
    conditions &= (db[table]['station_id'] == station_id)
    if from_date:
        conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
    if to_date:
        conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
    if not from_date and not to_date:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=365)
        conditions &= (db[table]['get_time'] >= from_date)
        conditions &= (db[table]['get_time'] < to_date)

    list_data = db(conditions).select(db[table].id,
                                      db[table].get_time,
                                      db[table].data,
                                      orderby=~db[table].get_time)

    iRow = 1

    si_dict = dict()
    conditions = (db.station_indicator.station_id == station_id)
    conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
    rows = db().select(db.station_indicator.ALL)
    indicator_ids = []
    for row in rows:
        indicator_ids.append(row.indicator_id)
        si_dict[str(row.indicator_id)] = row.as_dict()
    indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                    db.indicators.unit)

    # Duyet tung phan tu trong mang du lieu vua truy van duoc
    for i, item in enumerate(list_data):
        added_item = dict()
        if added_columns:
            for indicator in indicators:
                i_name = str(indicator.indicator)
                i_name_decode = i_name.decode('utf-8')
                v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                if v == '' or v is None:
                    added_item[i_name] = '-'
                else:
                    try:
                        v = float(v)
                        added_item[i_name] = "{0:.2f}".format(v)
                    except:
                        added_item[i_name] = '-'
        row = [
            item.get_time.strftime(datetime_format_vn),
        ]
        for column in added_columns:
            if column and added_item.has_key(column):
                row.append(added_item[column])
        aaData.append(row)

    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()
    if list_data.first() is None:
        raise HTTP(400, "KhÃ´ng Ä‘á»§ dá»¯ liá»‡u xuáº¥t excel, vui lÃ²ng thá»­ láº¡i!")
    else:
        # Write header
        temp_headers = []
        headers = []
        temp_headers.append(' ')
        for item in added_columns:
            temp_headers.append(str(item))
        headers.append(temp_headers)
        for header in headers:
            ws2.append(header)
        # Write data
        for row in aaData:
            ws2.append(row)

    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace("," , " ")

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


# ################################################################################
@service.json
def export_excel_7():
    import os.path, openpyxl
    # get search parameters
    from applications.eos.modules.w2pex import date_util
    from calendar import monthrange
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    month = request.vars.Month
    year = request.vars.Year
    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)

    if month and year:
        month = int(month)
        year = int(year)
        from_date = datetime(year=year, month=month, day=1)
        to_date = date_util.get_first_day_next_month(from_date)
        to_date = datetime.combine(to_date, datetime.min.time())

    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if not station_id:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station for viewing data'), success=True)
    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc
    if data_type == 2:
        table = 'data_hour_adjust'
    else:
        table = 'data_hour'

    conditions = (db[table]['id'] > 0)
    conditions &= (db[table]['station_id'] == station_id)
    if from_date:
        conditions &= (db[table]['get_time'] >= from_date)
    if to_date:
        conditions &= (db[table]['get_time'] < to_date)

    list_data = db(conditions).select(db[table].id,
                                      db[table].get_time,
                                      db[table].data,
                                      orderby=db[table].get_time)
    # limitby=limitby
    # Tong so ban ghi khong thuc hien phan trang
    iTotalRecords = monthrange(year, month)[1]
    # Thu tu ban ghi

    si_dict = dict()
    conditions = (db.station_indicator.station_id == station_id)
    conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
    rows = db(conditions).select(db.station_indicator.ALL)
    indicator_ids = []
    for row in rows:
        indicator_ids.append(row.indicator_id)
        si_dict[str(row.indicator_id)] = row.as_dict()
    indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                    db.indicators.unit)
    for i_count in range(1, iTotalRecords + 1):
        added_item = dict()
        for i, item in enumerate(list_data):
            if item.get_time.day == i_count:
                if added_columns:
                    for indicator in indicators:
                        i_name = str(indicator.indicator)
                        i_name_decode = i_name.decode('utf-8')
                        v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                        if v == '' or v is None:
                            pass
                        else:
                            try:
                                v = float(v)
                                if not added_item.has_key(i_name):
                                    added_item[i_name] = "{0:.2f}".format(v)
                                else:
                                    if v > int(added_item[i_name]):
                                        added_item[i_name] = "{0:.2f}".format(v)
                            except:
                                pass
        row = [
            str(i_count),
        ]
        for column in added_columns:
            if column and added_item.has_key(column):
                row.append(added_item[column])
            else:
                row.append('-')
        aaData.append(row)

    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()
    if list_data.first() is None:
        raise HTTP(400, "KhÃ´ng Ä‘á»§ dá»¯ liá»‡u xuáº¥t excel, vui lÃ²ng thá»­ láº¡i!")
    else:
        # Write header
        # ws['A1'] = 'Datetime'
        temp_headers = []
        headers = []
        date_str = "%s/%s" % (month, year)
        temp_headers.append(date_str)
        for item in added_columns:
            temp_headers.append(str(item))
        headers.append(temp_headers)
        for header in headers:
            ws2.append(header)
        # Write data
        for row in aaData:
            ws2.append(row)

    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace("," , " ")

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb2.save(file_path)
    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


# ################################################################################
@service.json
def export_excel_8():
    import os.path, openpyxl
    # get search parameters
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    month = request.vars.Month
    year = request.vars.Year
    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)
    if month and year:
        month = int(month)
        year = int(year)
        from_date = datetime(year=year, month=month, day=1)
        to_date = date_util.get_first_day_next_month(from_date)
        to_date = datetime.combine(to_date, datetime.min.time())
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if not station_id:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station for viewing data'), success=True)
    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc
    iTotalRecords = 0  # Tong so ban ghi  # Tuple dung de phan trang (vtri bat dau - chieu dai)
    if data_type == 2:
        table = 'data_day_adjust'
    else:
        table = 'data_day'

    conditions = (db[table]['id'] > 0)
    conditions &= (db[table]['station_id'] == station_id)
    if from_date:
        conditions &= (db[table]['get_time'] >= from_date)
    if to_date:
        conditions &= (db[table]['get_time'] < to_date)
    if not from_date and not to_date:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=365)
        conditions &= (db[table]['get_time'] >= from_date)
        conditions &= (db[table]['get_time'] < to_date)

    list_data = db(conditions).select(db[table].id,
                                      db[table].get_time,
                                      db[table].data,
                                      orderby=db[table].get_time)
    # Tong so ban ghi khong thuc hien phan trang
    iTotalRecords = db(conditions).count(db[table].id)
    # Thu tu ban ghi
    si_dict = dict()
    conditions = (db.station_indicator.station_id == station_id)
    conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
    rows = db().select(db.station_indicator.ALL)
    indicator_ids = []
    for row in rows:
        indicator_ids.append(row.indicator_id)
        si_dict[str(row.indicator_id)] = row.as_dict()
    indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                    db.indicators.unit)

    # Duyet tung phan tu trong mang du lieu vua truy van duoc
    for i, item in enumerate(list_data):
        added_item = dict()
        if added_columns:
            for indicator in indicators:
                i_name = str(indicator.indicator)
                i_name_decode = i_name.decode('utf-8')
                v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                if v == '' or v is None:
                    added_item[i_name] = '-'
                else:
                    try:
                        v = float(v)
                        added_item[i_name] = "{0:.2f}".format(v)
                    except:
                        added_item[i_name] = '-'
        row = [
            item.get_time.strftime('%d-%m-%Y'),
        ]
        for column in added_columns:
            if column and added_item.has_key(column):
                row.append(added_item[column])
        aaData.append(row)
    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()
    if list_data.first() is None:
        raise HTTP(400, "KhÃ´ng Ä‘á»§ dá»¯ liá»‡u xuáº¥t excel, vui lÃ²ng thá»­ láº¡i!")
    else:
        # Write header
        temp_headers = []
        headers = []
        temp_headers.append(' ')
        for item in added_columns:
            temp_headers.append(str(item))
        headers.append(temp_headers)
        for header in headers:
            ws2.append(header)
        # Write data
        for row in aaData:
            ws2.append(row)

    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace("," , " ")

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


# ################################################################################
@service.json
def export_excel_9():
    import os.path, openpyxl
    # get search parameters
    from applications.eos.modules.w2pex import date_util
    from calendar import monthrange
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    month = request.vars.Month
    year = request.vars.Year
    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)

    # month_date = request.vars.month_date
    if month and year:
        month = int(month)
        year = int(year)
        from_date = datetime(year=year, month=month, day=1)
        to_date = date_util.get_first_day_next_month(from_date)
        to_date = datetime.combine(to_date, datetime.min.time())

    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if not station_id:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station for viewing data'), success=True)
    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc
    if data_type == 2:
        table = 'data_hour_8h_adjust'
    else:
        table = 'data_hour_8h'

    conditions = (db[table]['id'] > 0)
    conditions &= (db[table]['station_id'] == station_id)
    if from_date:
        conditions &= (db[table]['get_time'] >= from_date)
    if to_date:
        conditions &= (db[table]['get_time'] < to_date)

    list_data = db(conditions).select(db[table].id,
                                      db[table].get_time,
                                      db[table].data,
                                      orderby=db[table].get_time)
    # limitby=limitby
    # Tong so ban ghi khong thuc hien phan trang
    iTotalRecords = monthrange(year, month)[1]
    # Thu tu ban ghi

    si_dict = dict()
    conditions = (db.station_indicator.station_id == station_id)
    conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
    rows = db().select(db.station_indicator.ALL)
    indicator_ids = []
    for row in rows:
        indicator_ids.append(row.indicator_id)
        si_dict[str(row.indicator_id)] = row.as_dict()
    indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                    db.indicators.unit)

    for i_count in range(1, iTotalRecords + 1):
        added_item = dict()
        for i, item in enumerate(list_data):
            if (item.get_time.day == i_count) and (item.get_time.hour >= 8):
                if added_columns:
                    for indicator in indicators:
                        i_name = str(indicator.indicator)
                        i_name_decode = i_name.decode('utf-8')
                        v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                        if v == '' or v is None:
                            pass
                        else:
                            try:
                                v = float(v)
                                if not added_item.has_key(i_name):
                                    added_item[i_name] = "{0:.2f}".format(v)
                                else:
                                    if v > int(added_item[i_name]):
                                        added_item[i_name] = "{0:.2f}".format(v)
                            except:
                                pass
        row = [
            str(i_count),
            # item.get_time.strftime('%Y-%m-%d %H:%M:%S'),
            # item.get_time.strftime(datetime_format_vn),
        ]
        for column in added_columns:
            if column and added_item.has_key(column):
                row.append(added_item[column])
            else:
                row.append('-')
        aaData.append(row)

    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()
    if list_data.first() is None:
        raise HTTP(400, "KhÃ´ng Ä‘á»§ dá»¯ liá»‡u xuáº¥t excel, vui lÃ²ng thá»­ láº¡i!")
    else:
        # Write header
        temp_headers = []
        headers = []
        date_str = "%s/%s" % (month, year)
        temp_headers.append(date_str)
        for item in added_columns:
            temp_headers.append(str(item))
        headers.append(temp_headers)
        for header in headers:
            ws2.append(header)
        # Write data
        for row in aaData:
            ws2.append(row)

    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace("," , " ")

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


# ################################################################################
@service.json
def export_excel_10():
    from openpyxl.styles import Alignment
    import os.path, openpyxl
    # get search parameters
    from applications.eos.modules.w2pex import date_util
    from calendar import monthrange
    # So luong ban ghi se lay toi da
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    month = request.vars.Month
    year = request.vars.Year
    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)
    if month and year:
        month = int(month)
        year = int(year)
        from_date = datetime(year=year, month=month, day=1)
        to_date = date_util.get_first_day_next_month(from_date)
        to_date = datetime.combine(to_date, datetime.min.time())

    if not station_id:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station for viewing data'), success=True)
    aaData = []  # Du lieu json se tra ve
    # lay du lieu aqi hour
    if data_type == 2:
        table = 'aqi_data_adjust_hour'
    else:
        table = 'aqi_data_hour'
    conditions = (db[table]['id'] > 0)
    conditions &= (db[table]['station_id'] == station_id)
    if from_date:
        conditions &= (db[table]['get_time'] >= from_date)
    if to_date:
        conditions &= (db[table]['get_time'] < to_date)
    if not from_date and not to_date:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=365)
        conditions &= (db[table]['get_time'] >= from_date)
        conditions &= (db[table]['get_time'] < to_date)

    list_data = db(conditions).select(db[table].id,
                                      db[table].get_time,
                                      db[table].data,
                                      orderby=~db[table].get_time)
    # Tong so ban ghi khong thuc hien phan trang

    aqi_hour = dict()
    # Duyet tung phan tu trong mang du lieu vua truy van duoc
    for i, item in enumerate(list_data):
        day = item.get_time.day
        hour = item.get_time.hour
        key = (day, hour)
        v = item.data['aqi'] if item.data.has_key('aqi') else ''
        if v == '' or v is None:
            aqi_hour[key] = '-'
        else:
            try:
                v = float(v)
                aqi_hour[key] = "{0:.2f}".format(v)
            except:
                aqi_hour[key] = '-'
    #####################################################
    # lay du lieu aqi day
    table = 'aqi_data_adjust_24h'
    conditions = (db[table]['id'] > 0)
    conditions &= (db[table]['station_id'] == station_id)
    if from_date:
        conditions &= (db[table]['get_time'] >= from_date)
    if to_date:
        conditions &= (db[table]['get_time'] < to_date)
    if not from_date and not to_date:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=365)
        conditions &= (db[table]['get_time'] >= from_date)
        conditions &= (db[table]['get_time'] < to_date)

    list_data = db(conditions).select(db[table].id,
                                      db[table].get_time,
                                      db[table].data_24h,
                                      orderby=~db[table].get_time)
    # Tong so ban ghi khong thuc hien phan trang

    aqi_day = dict()
    # Duyet tung phan tu trong mang du lieu vua truy van duoc
    for i, item in enumerate(list_data):
        day = item.get_time.day
        v = item.data_24h['aqi'] if item.data_24h.has_key('aqi') else ''
        if v == '' or v is None:
            aqi_day[day] = '-'
        else:
            try:
                v = float(v)
                aqi_day[day] = "{0:.2f}".format(v)
            except:
                aqi_day[day] = '-'

    #####################################################
    iTotalRecords = monthrange(year, month)[1]
    # Thu tu ban ghi
    # iRow = iDisplayStart + 1
    for i_count in range(1, iTotalRecords + 1):
        row = [
            str(i_count),
        ]
        for j in range(0, 24):
            key = (i_count, j)
            if aqi_hour.has_key(key):
                row.append(aqi_hour[key])
            else:
                row.append('-')
        if aqi_day.has_key(i_count):
            row.append(aqi_day[i_count])
        else:
            row.append('-')
        aaData.append(row)


    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    headers_1 = []
    headers_2 = []

    headers = []
    headers.append('#')
    for i in range(0, 24):
        headers.append('AQI Hour')
    headers.append('AQI Day')
    headers_1.append(headers)

    headers = []
    headers.append('#')
    for i in range(0, 24):
        headers.append(i)
    headers.append('AQI Day')
    headers_2.append(headers)

    for header in headers_1:
        ws2.append(header)
    for header in headers_2:
        ws2.append(header)
    for row in aaData:
        ws2.append(row)
    ws2.merge_cells('A1:A2')
    ws2.merge_cells('Z1:Z2')
    ws2.merge_cells('B1:Y1')
    for col in ws2.columns:
        for cell in col:
            alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
            cell.alignment = alignment_obj

    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace("," , " ")

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


# ################################################################################
@service.json
def export_excel_11():
    import os.path, openpyxl
    # get search parameters
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    # month = request.vars.Month
    year = request.vars.Year
    # if month and year:
    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)
    if year:
        # month = int(month)
        year = int(year)
        from_date = datetime(year=year, month=1, day=1)
        to_date = datetime(year=year + 1, month=1, day=1)
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if not station_id:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station for viewing data'), success=True)
    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc

    if data_type == 2:
        table = 'data_mon_adjust'
    else:
        table = 'data_mon'

    conditions = (db[table]['id'] > 0)
    conditions &= (db[table]['station_id'] == station_id)
    if from_date:
        conditions &= (db[table]['get_time'] >= from_date)
    if to_date:
        conditions &= (db[table]['get_time'] < to_date)
    if not from_date and not to_date:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=365)
        conditions &= (db[table]['get_time'] >= from_date)
        conditions &= (db[table]['get_time'] < to_date)

    list_data = db(conditions).select(db[table].id,
                                      db[table].get_time,
                                      db[table].data,
                                      orderby=db[table].get_time)
    # Tong so ban ghi khong thuc hien phan trang
    iTotalRecords = 1
    # Thu tu ban ghi
    si_dict = dict()
    conditions = (db.station_indicator.station_id == station_id)
    conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
    rows = db().select(db.station_indicator.ALL)
    indicator_ids = []
    for row in rows:
        indicator_ids.append(row.indicator_id)
        si_dict[str(row.indicator_id)] = row.as_dict()
    indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                    db.indicators.unit)
    result_dict = dict()
    total_result = dict()
    count_result = dict()
    # Duyet tung phan tu trong mang du lieu vua truy van duoc
    for i, item in enumerate(list_data):
        added_item = dict()
        if added_columns:
            for indicator in indicators:
                i_name = str(indicator.indicator)
                i_name_decode = i_name.decode('utf-8')
                v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                if v == '' or v is None:
                    added_item[i_name] = '-'
                else:
                    try:
                        v = float(v)
                        added_item[i_name] = "{0:.2f}".format(v)
                        if total_result.has_key(i_name):
                            total_result[i_name] += v
                            count_result[i_name] += 1
                        else:
                            total_result[i_name] = v
                            count_result[i_name] = 1
                    except:
                        added_item[i_name] = '-'
        row = [
            item.get_time.month
        ]
        for column in added_columns:
            if column and added_item.has_key(column):
                row.append(added_item[column])
        result_dict[item.get_time.month] = row
    for i in range(1, 13):
        if not result_dict.has_key(i):
            row = [
                i,
            ]
            for column in added_columns:
                row.append('-')
            aaData.append(row)
        else:
            aaData.append(result_dict[i])
    row = [
        'TB Nam'
    ]
    for column in added_columns:
        if total_result.has_key(column):
            avg = float(total_result[column]) / count_result[column]
            row.append("{0:.2f}".format(avg))
        else:
            row.append('-')
    aaData.append(row)
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    if list_data.first() is None:
        raise HTTP(400, "KhÃ´ng Ä‘á»§ dá»¯ liá»‡u xuáº¥t excel, vui lÃ²ng thá»­ láº¡i!")
    else:
        # Write header
        temp_headers = []
        headers = []
        # date_str = T('Month')
        temp_headers.append(' ')
        for item in added_columns:
            # ws[chr(ord('A') + i + 1) + '1'] = indicator.upper()
            temp_headers.append(str(item))
        headers.append(temp_headers)
        for header in headers:
            ws2.append(header)
        # Write data
        for row in aaData:
            ws2.append(row)
    for col in ws2.columns:
        for cell in col:
            alignment_obj = cell.alignment.copy(horizontal='left', vertical='center')
            cell.alignment = alignment_obj
    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace("," , " ")

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


# ################################################################################
@service.json
def export_excel_12():
    import os.path, openpyxl
    # get search parameters
    station_id = request.vars.station_id
    from_date = request.vars.from_date
    to_date = request.vars.to_date
    data_type = request.vars.data_type
    if data_type:
        data_type = int(data_type)
    added_columns = request.vars.added_columns or ''
    if added_columns:
        added_columns = added_columns.split(',')
    if not station_id:
        return dict(iTotalRecords=0, iTotalDisplayRecords=0, aaData=[],
                    message=T('Select a station for viewing data'), success=True)
    aaData = []  # Du lieu json se tra ve
    list_data = None  # Du lieu truy van duoc
    if data_type == 2:
        table = 'data_adjust'
    else:
        table = 'data_min'

    conditions = (db[table]['id'] > 0)
    conditions &= (db[table]['station_id'] == station_id)
    if from_date:
        conditions &= (db[table]['get_time'] >= date_util.string_to_datetime(from_date))
    if to_date:
        conditions &= (db[table]['get_time'] < date_util.string_to_datetime(to_date) + timedelta(days=1))
    if not from_date and not to_date:
        to_date = datetime.now()
        from_date = to_date - timedelta(days=365)
        conditions &= (db[table]['get_time'] >= from_date)
        conditions &= (db[table]['get_time'] < to_date)


    list_data = db(conditions).select(db[table].id,
                                      db[table].get_time,
                                      db[table].data,
                                      orderby=~db[table].get_time)

    iRow = 1

    si_dict = dict()
    conditions = (db.station_indicator.station_id == station_id)
    conditions &= (db.station_indicator.status == const.SI_STATUS['IN_USE']['value'])
    rows = db().select(db.station_indicator.ALL)
    indicator_ids = []
    for row in rows:
        indicator_ids.append(row.indicator_id)
        si_dict[str(row.indicator_id)] = row.as_dict()
    indicators = db(db.indicators.id.belongs(indicator_ids)).select(db.indicators.id, db.indicators.indicator,
                                                                    db.indicators.unit)

    # Duyet tung phan tu trong mang du lieu vua truy van duoc
    if from_date and to_date:
        start_time = date_util.string_to_datetime(from_date)
        end_time = date_util.string_to_datetime(to_date) + timedelta(days=1)
        freq_data = db(db.stations.id == station_id).select(db.stations.frequency_receiving_data)
        frequency = 60
        if freq_data:
            itemfreq = freq_data.first()['frequency_receiving_data']
            if itemfreq:
                frequency = round(itemfreq * 60)
        iTotalRecords = int((end_time - start_time).total_seconds() / frequency)
        viewable = True
        for i in range(1, iTotalRecords):
            time = start_time + timedelta(seconds=i * frequency)
            added_item = dict()
            for j, item in enumerate(list_data):
                if item.get_time == time:
                    # viewable = False
                    if added_columns:
                        for indicator in indicators:
                            i_name = str(indicator.indicator)
                            i_name_decode = i_name.decode('utf-8')
                            v = item.data[i_name_decode] if item.data.has_key(i_name_decode) else ''
                            if v == '' or v is None:
                                added_item[i_name] = '-'
                            else:
                                try:
                                    v = float(v)
                                    added_item[i_name] = "{0:.2f}".format(v)
                                except:
                                    viewable = True
                                    added_item[i_name] = '-'
            row = [
                str(i + 1),
                time.strftime(datetime_format_vn),
            ]
            for column in added_columns:
                if column and added_item.has_key(column):
                    row.append(added_item[column])
                else:
                    row.append('-')
            if viewable:
                aaData.append(row)


    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet()
    if list_data.first() is None:
        raise HTTP(400, "KhÃ´ng Ä‘á»§ dá»¯ liá»‡u xuáº¥t excel, vui lÃ²ng thá»­ láº¡i!")
    else:
        # Write header
        temp_headers = []
        headers = []
        temp_headers.append(' ')
        for item in added_columns:
            temp_headers.append(str(item))
        headers.append(temp_headers)
        for header in headers:
            ws2.append(header)
        # Write data
        for row in aaData:
            ws2.append(row)

    # Get station name
    station = db.stations(station_id)
    station_name = station.station_name if station else ''
    if "," in station_name:
        station_name = station_name.replace("," , " ")

    file_name = request.now.strftime(station_name + '_%Y%m%d_%H%M%S.xlsx')
    file_path = os.path.join(request.folder, 'static', 'export', file_name)
    # wb.save(file_path)
    wb2.save(file_path)

    data = open(file_path, "rb").read()
    os.unlink(file_path)
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment;filename=%s' % file_name

    return data


################################################################################
@service.json
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def get_indicators(*args, **kwargs):
    try:
        html = ''
        station_type = request.vars.type
        from_public = request.vars.from_public
        # conditions = (db.station_indicator.station_id == station_id)
        conditions = (db.indicators.id > 0)
        if station_type:
            conditions &= (db.indicators.indicator_type == station_type)
        rows = db(conditions).select(db.indicators.ALL)
        for row in rows:
            html += '<option value="%(value)s" selected>%(name)s (%(unit)s)</option>' % dict(value=row.indicator,
                                                                                             name=row.indicator,
                                                                                             unit=row.unit)
        return dict(success=True, html=html)
    except Exception as ex:
        return dict(success=False, message=str(ex))


@service.json
@auth.requires(lambda: (auth.has_permission('view', 'view_report')))
def get_indicators_and_stations(*args, **kwargs):
    try:
        html = ''
        station_type = request.vars.type
        province_id = request.vars.province_id
        from_public = request.vars.from_public
        # conditions = (db.station_indicator.station_id == station_id)
        conditions = (db.indicators.id > 0)
        if station_type:
            conditions &= (db.indicators.indicator_type == station_type)
        rows = db(conditions).select(db.indicators.ALL)
        for row in rows:
            html += '<option value="%(value)s" selected>%(name)s (%(unit)s)</option>' % dict(value=row.indicator,
                                                                                             name=row.indicator,
                                                                                             unit=row.unit)

        html2 = ''
        # conditions = (db.stations.id > 0)
        conditions = (db.stations.id > 0)
        if station_type:
            conditions &= (db.stations.station_type == station_type)
        if province_id:
            conditions &= (db.stations.province_id == province_id)
        rows = db(conditions).select(
            db.stations.id,
            db.stations.station_name,
        )
        for row in rows:
            html2 += '<option value="%(value)s" selected>%(name)s</option>' % dict(value=row.id,
                                                                                   name=row.station_name)
        return dict(success=True, html=html, html_2=html2)
    except Exception as ex:
        return dict(success=False, message=str(ex))


################################################################################
# hungdx issue 44 add (co the ap dung cho cac chon tinh, vung ... fix sau
@service.json
def dropdown_content(table, filter_field, get_value_field, get_dsp_field, *args, **kwargs):
    try:
        filter_value = request.vars.filter_value
        filter_value = filter_value.split(';')
        filter_field = filter_field.split('-')
        conditions = (db[table]['id'] > 0)
        t1 = len(filter_field)
        t2 = len(filter_value)
        if t1 == t2:
            for i in range(0, t1):
                if filter_field[i] and filter_value[i] != '':
                    conditions &= (db[table][filter_field[i]] == filter_value[i])
        if current_user:
            if not (auth.has_membership('admin') or auth.has_membership(const.GROUP_MANAGER)):
                list_station_manager = db(db.manager_stations.user_id == current_user.id).select(
                    db.manager_stations.station_id)
                station_ids = [str(item.station_id) for item in list_station_manager]
                conditions &= (db[table]['id'].belongs(station_ids))

        records = db(conditions).select(db[table][get_value_field], db[table][get_dsp_field])

        if records:
            html1 = "<option value=''>%s</option>" % T('-- Select an option --')
            html = ["<option value='%s'>%s</option>" % (item[get_value_field], item[get_dsp_field]) for item in records]
            html = html1 + ''.join(html)
        else:
            html = "<option value=''>%s</option>" % T('-- No data --')

        return dict(success=True, html=html)
    except Exception as ex:
        return dict(success=False, message=str(ex))

